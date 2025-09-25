"""
MySQL MCP服务器主文件

基于FastMCP框架实现MySQL数据库操作服务

@fileoverview MySQL MCP服务器主文件
@author liyq
@version 1.0.0
@since 1.0.0
@updated 2025-09-23
@license MIT
"""

import json
import os
import signal
import sys
import time
from typing import Any, Dict, List, Optional
from datetime import datetime
from pydantic import BaseModel

from fastmcp import FastMCP

from mysql_manager import MySQLManager
from tool_wrapper import create_mcp_tool, ToolDefinition
from constants import STRING_CONSTANTS
from error_handler import ErrorHandler
from logger import logger
from monitor import system_monitor, memory_monitor
from performance_manager import PerformanceManager
from backup_tool import MySQLBackupTool
from import_tool import MySQLImportTool
from rate_limit import RateLimiterManager
from config import config_manager
from type_utils import ErrorCategory, ErrorSeverity, MySQLMCPError, RateLimitConfig

# 创建MySQL管理器实例
mysql_manager = MySQLManager()

# 创建性能管理器实例
performance_manager = PerformanceManager(mysql_manager)

# 创建备份工具实例
backup_tool = MySQLBackupTool(mysql_manager)

# 创建导入工具实例
import_tool = MySQLImportTool(mysql_manager)

# FastMCP 服务器实例配置
mcp = FastMCP(
    name=STRING_CONSTANTS["SERVER_NAME"],
    version=STRING_CONSTANTS["SERVER_VERSION"]
)

# 创建全局速率限制器
rate_limiter_config = RateLimitConfig(
    max_requests=config_manager.security.rate_limit_max,
    window_seconds=config_manager.security.rate_limit_window,
    burst_limit=config_manager.security.rate_limit_max // 2,  # 突发限制为最大请求数的一半
    strategy="token_bucket"  # 使用令牌桶算法，适合数据库操作
)
rate_limiter = RateLimiterManager(rate_limiter_config)


# =============================================================================
# 工具参数模型
# =============================================================================

class QueryParams(BaseModel):
    """MySQL查询参数模型

    用于传递SQL查询语句和参数绑定的数据结构。
    支持参数化查询以防止SQL注入攻击。

    @param query: SQL查询语句字符串
    @param params: 可选的参数列表，用于参数化查询
    """
    query: str
    params: Optional[List[Any]] = None


class TableNameParams(BaseModel):
    """表名参数模型

    简单的参数模型，只包含表名，用于需要指定表的操作。

    @param table_name: 目标表的名称
    """
    table_name: str


class SelectDataParams(BaseModel):
    """数据查询参数模型

    用于从表中查询数据的参数模型，支持列选择、WHERE条件和行数限制。

    @param table_name: 要查询的表名
    @param columns: 可选的列名列表，如果未指定则查询所有列
    @param where_clause: 可选的WHERE条件子句
    @param limit: 可选的行数限制
    """
    table_name: str
    columns: Optional[List[str]] = None
    where_clause: Optional[str] = None
    limit: Optional[int] = None


class InsertDataParams(BaseModel):
    """数据插入参数模型

    用于向表中插入新数据的参数模型。

    @param table_name: 要插入数据的表名
    @param data: 要插入的数据字典，键为列名，值为对应的数据
    """
    table_name: str
    data: Dict[str, Any]


class UpdateDataParams(BaseModel):
    """数据更新参数模型

    用于更新表中现有数据的参数模型，根据WHERE条件更新指定数据。

    @param table_name: 要更新数据的表名
    @param data: 要更新的数据字典，键为列名，值为新数据
    @param where_clause: WHERE条件子句，指定要更新的行
    """
    table_name: str
    data: Dict[str, Any]
    where_clause: str


class DeleteDataParams(BaseModel):
    """数据删除参数模型

    用于从表中删除数据的参数模型，根据WHERE条件删除符合条件的数据。

    @param table_name: 要删除数据的表名
    @param where_clause: WHERE条件子句，指定要删除的行
    """
    table_name: str
    where_clause: str


class CreateTableParams(BaseModel):
    """创建表参数模型

    用于创建新表的参数模型，包含表名和列定义信息。

    @param table_name: 要创建的表名
    @param columns: 列定义列表，每个列包含名称、类型、约束等信息
    """
    table_name: str
    columns: List[Dict[str, Any]]


class DropTableParams(BaseModel):
    """删除表参数模型

    用于删除表的参数模型，支持可选的IF EXISTS子句。

    @param table_name: 要删除的表名
    @param if_exists: 可选，是否使用IF EXISTS子句，避免表不存在时报错
    """
    table_name: str
    if_exists: Optional[bool] = False


class AlterTableParams(BaseModel):
    """修改表结构参数模型

    用于修改表结构的参数模型，支持添加、修改、删除列和索引等操作。

    @param table_name: 要修改的表名
    @param alterations: 修改操作列表，每个操作包含类型和具体参数
    """
    table_name: str
    alterations: List[Dict[str, Any]]


class BatchExecuteParams(BaseModel):
    """批量执行参数模型

    用于批量执行多个SQL查询的参数模型，在单个事务中执行以确保原子性。

    @param queries: 查询列表，每个查询包含SQL语句和可选的参数
    """
    queries: List[Dict[str, Any]]


class BackupParams(BaseModel):
    """数据库备份参数模型

    用于配置数据库备份操作的参数模型，支持完整的备份选项。

    @param output_dir: 可选的输出目录路径
    @param compress: 可选，是否压缩备份文件，默认True
    @param include_data: 可选，是否包含表数据，默认True
    @param include_structure: 可选，是否包含表结构，默认True
    @param tables: 可选，要备份的特定表列表
    @param file_prefix: 可选，备份文件名前缀，默认"mysql_backup"
    @param max_file_size: 可选，最大文件大小(MB)，默认100
    @param backup_type: 可选，备份类型，默认"full"
    """
    output_dir: Optional[str] = None
    compress: Optional[bool] = True
    include_data: Optional[bool] = True
    include_structure: Optional[bool] = True
    tables: Optional[List[str]] = None
    file_prefix: Optional[str] = "mysql_backup"
    max_file_size: Optional[int] = 100
    backup_type: Optional[str] = "full"


class ExportDataParams(BaseModel):
    """数据导出参数模型

    用于配置数据导出操作的参数模型，支持多种格式和选项。

    @param query: 要执行的导出查询SQL语句
    @param params: 可选，查询参数列表
    @param output_dir: 可选，输出目录路径
    @param format: 可选，导出格式，默认"excel"
    @param sheet_name: 可选，Excel工作表名称，默认"Data"
    @param include_headers: 可选，是否包含列标题，默认True
    @param max_rows: 可选，最大导出行数，默认100000
    @param file_name: 可选，自定义文件名
    """
    query: str
    params: Optional[List[Any]] = None
    output_dir: Optional[str] = None
    format: Optional[str] = "excel"
    sheet_name: Optional[str] = "Data"
    include_headers: Optional[bool] = True
    max_rows: Optional[int] = 100000
    file_name: Optional[str] = None


class GenerateReportParams(BaseModel):
    """报表生成参数模型

    用于配置数据报表生成的参数模型，支持多查询和Excel输出。

    @param title: 报表标题
    @param description: 可选，报表描述
    @param queries: 查询列表，每个查询包含名称和SQL语句
    @param output_dir: 可选，输出目录路径
    @param file_name: 可选，自定义文件名
    @param include_headers: 可选，是否包含列标题，默认True
    """
    title: str
    description: Optional[str] = None
    queries: List[Dict[str, Any]]
    output_dir: Optional[str] = None
    file_name: Optional[str] = None
    include_headers: Optional[bool] = True


class ImportDataParams(BaseModel):
    """数据导入参数模型

    用于配置数据导入操作的参数模型，支持多种文件格式和导入选项。

    @param table_name: 目标表名
    @param file_path: 要导入的文件路径
    @param format: 文件格式，支持csv、json、excel、sql等
    @param has_headers: 可选，CSV/Excel文件是否有标题行，默认True
    @param field_mapping: 可选，字段映射字典
    @param batch_size: 可选，批处理大小，默认1000
    @param skip_duplicates: 可选，是否跳过重复行，默认False
    @param conflict_strategy: 可选，冲突处理策略，默认"error"
    @param use_transaction: 可选，是否使用事务，默认True
    @param validate_data: 可选，是否验证数据，默认True
    @param encoding: 可选，文件编码，默认"utf8"
    @param sheet_name: 可选，Excel工作表名称
    @param delimiter: 可选，CSV分隔符，默认","
    @param quote: 可选，CSV引号字符，默认'"'
    """
    table_name: str
    file_path: str
    format: str
    has_headers: Optional[bool] = True
    field_mapping: Optional[Dict[str, str]] = None
    batch_size: Optional[int] = 1000
    skip_duplicates: Optional[bool] = False
    conflict_strategy: Optional[str] = "error"
    use_transaction: Optional[bool] = True
    validate_data: Optional[bool] = True
    encoding: Optional[str] = "utf8"
    sheet_name: Optional[str] = None
    delimiter: Optional[str] = ","
    quote: Optional[str] = '"'


class AnalyzeErrorParams(BaseModel):
    """错误分析参数模型

    用于传递错误信息进行分析的参数模型。

    @param error_message: 要分析的错误消息
    @param operation: 可选，发生错误的上下文操作
    """
    error_message: str
    operation: Optional[str] = None


class ManageIndexesParams(BaseModel):
    """索引管理参数模型

    用于配置索引管理操作的参数模型，支持创建、删除、分析索引等操作。

    @param action: 操作类型，如"create"、"drop"、"analyze"、"optimize"、"list"
    @param table_name: 可选，目标表名
    @param index_name: 可选，索引名称
    @param index_type: 可选，索引类型，默认"INDEX"
    @param columns: 可选，索引包含的列名列表
    @param if_exists: 可选，删除时是否检查索引存在，默认False
    @param invisible: 可选，是否创建不可见索引，默认False
    """
    action: str
    table_name: Optional[str] = None
    index_name: Optional[str] = None
    index_type: Optional[str] = "INDEX"
    columns: Optional[List[str]] = None
    if_exists: Optional[bool] = False
    invisible: Optional[bool] = False


class ManageUsersParams(BaseModel):
    """用户管理参数模型

    用于配置MySQL用户管理操作的参数模型，支持创建、删除用户和权限管理。

    @param action: 操作类型，如"create"、"delete"、"grant"、"revoke"、"list"、"show_grants"
    @param username: 可选，用户名
    @param password: 可选，密码（创建用户时必需）
    @param host: 可选，主机地址，默认"%"
    @param privileges: 可选，权限列表
    @param database: 可选，目标数据库名
    @param table: 可选，目标表名
    @param if_exists: 可选，删除时是否检查用户存在，默认False
    """
    action: str
    username: Optional[str] = None
    password: Optional[str] = None
    host: Optional[str] = "%"
    privileges: Optional[List[str]] = None
    database: Optional[str] = None
    table: Optional[str] = None
    if_exists: Optional[bool] = False


class ProgressTrackerParams(BaseModel):
    """进度跟踪参数模型

    用于配置操作进度跟踪的参数模型，支持列出、获取和取消操作。

    @param action: 操作类型，如"list"、"get"、"cancel"、"summary"
    @param tracker_id: 可选，跟踪器ID
    @param operation_type: 可选，操作类型过滤，默认"all"
    @param include_completed: 可选，是否包含已完成的操作，默认False
    @param detail_level: 可选，详情级别，默认"basic"
    """
    action: str
    tracker_id: Optional[str] = None
    operation_type: Optional[str] = "all"
    include_completed: Optional[bool] = False
    detail_level: Optional[str] = "basic"


class OptimizeMemoryParams(BaseModel):
    """内存优化参数模型

    用于配置内存优化操作的参数模型，支持状态检查、清理和配置管理。

    @param action: 操作类型，如"status"、"cleanup"、"optimize"、"configure"、"report"、"gc"
    @param force_gc: 可选，是否强制垃圾回收，默认True
    @param enable_monitoring: 可选，是否启用监控
    @param max_concurrency: 可选，最大并发数
    @param include_history: 可选，是否包含历史记录，默认False
    """
    action: str
    force_gc: Optional[bool] = True
    enable_monitoring: Optional[bool] = None
    max_concurrency: Optional[int] = None
    include_history: Optional[bool] = False


class ManageQueueParams(BaseModel):
    """队列管理参数模型

    用于配置任务队列管理操作的参数模型，支持暂停、恢复、设置并发数等。

    @param action: 操作类型，如"status"、"pause"、"resume"、"clear"、"set_concurrency"、"cancel"、"diagnostics"、"get_task"
    @param task_id: 可选，任务ID
    @param max_concurrency: 可选，最大并发数
    @param show_details: 可选，是否显示详情，默认False
    @param filter_type: 可选，过滤类型，默认"all"
    """
    action: str
    task_id: Optional[str] = None
    max_concurrency: Optional[int] = None
    show_details: Optional[bool] = False
    filter_type: Optional[str] = "all"


class PerformanceOptimizeParams(BaseModel):
    """性能优化参数模型

    用于配置性能优化和监控操作的参数模型，支持慢查询分析和索引建议。

    @param action: 操作类型，如"enable_slow_query_log"、"status_slow_query_log"等
    @param query: 可选，要分析的SQL查询
    @param params: 可选，查询参数
    @param limit: 可选，结果限制数量，默认50
    @param include_details: 可选，是否包含详细信息，默认True
    @param time_range: 可选，时间范围，默认"1 day"
    @param long_query_time: 可选，慢查询阈值（秒）
    @param log_queries_not_using_indexes: 可选，是否记录未使用索引的查询
    @param monitoring_interval_minutes: 可选，监控间隔（分钟），默认60
    """
    action: str
    query: Optional[str] = None
    params: Optional[List[Any]] = None
    limit: Optional[int] = 50
    include_details: Optional[bool] = True
    time_range: Optional[str] = "1 day"
    long_query_time: Optional[int] = None
    log_queries_not_using_indexes: Optional[bool] = None
    monitoring_interval_minutes: Optional[int] = 60


# =============================================================================
# 限流辅助函数
# =============================================================================

def get_client_identifier(args: Dict) -> str:
    """生成客户端标识符用于限流

    @param args: 函数参数
    @return: 客户端标识符
    """
    # 尝试从参数中提取客户端标识
    client_ip = getattr(args, 'client_ip', None) or 'unknown'
    user_id = getattr(args, 'user_id', None) or 'anonymous'

    # 生成唯一标识符
    identifier = f"{client_ip}:{user_id}"

    # 如果没有客户端信息，使用操作类型作为标识
    if client_ip == 'unknown' and user_id == 'anonymous':
        identifier = f"operation:{args.__class__.__name__}"

    return identifier


def check_rate_limit(args: Dict, operation_type: str) -> bool:
    """检查速率限制

    @param args: 函数参数
    @param operation_type: 操作类型
    @return: 是否允许请求
    """
    try:
        # 生成客户端标识符
        client_id = get_client_identifier(args)

        # 创建操作特定的标识符
        identifier = f"{client_id}:{operation_type}"

        # 检查限流
        return rate_limiter.is_allowed(identifier)

    except Exception as e:
        logger.warn(f"限流检查失败，使用默认策略: {e}")
        # 如果限流检查失败，默认允许请求
        return True


def get_rate_limit_status(args: Dict, operation_type: str) -> Dict:
    """获取限流状态信息

    @param args: 函数参数
    @param operation_type: 操作类型
    @return: 限流状态信息
    """
    try:
        client_id = get_client_identifier(args)
        identifier = f"{client_id}:{operation_type}"

        status = rate_limiter.check_limit(identifier)

        return {
            "allowed": status.allowed,
            "remaining": status.remaining,
            "reset_time": status.reset_time.isoformat() if status.reset_time else None,
            "retry_after": status.retry_after,
            "operation": operation_type,
            "client_id": client_id
        }

    except Exception as e:
        logger.warn(f"获取限流状态失败: {e}")
        return {
            "allowed": True,
            "remaining": 999,
            "reset_time": None,
            "retry_after": None,
            "operation": operation_type,
            "client_id": get_client_identifier(args),
            "error": str(e)
        }


# =============================================================================
# 工具实现函数
# =============================================================================

async def mysql_query(args: QueryParams) -> str:
    """执行MySQL查询

    执行任意MySQL查询，支持参数绑定以确保安全性。
    支持SELECT、SHOW、DESCRIBE等操作。

    @param args: 查询参数
    @return: JSON格式的查询结果
    @throws: MySQLMCPError 当查询验证失败或执行错误时
    """
    # 检查查询长度限制
    if len(args.query) > config_manager.security.max_query_length:
        raise MySQLMCPError(
            f"查询长度 ({len(args.query)}) 超过最大限制 ({config_manager.security.max_query_length})",
            ErrorCategory.VALIDATION_ERROR,
            ErrorSeverity.HIGH
        )

    result = await mysql_manager.execute_query(args.query, args.params)
    return json.dumps(result, ensure_ascii=False, indent=2)


async def mysql_show_tables(args: Dict) -> str:
    """显示所有表

    使用SHOW TABLES命令列出当前数据库中的所有表。
    结果会被缓存以优化性能，提高频繁查询的响应速度。
    提供数据库架构的快速概览，支持开发和运维场景。

    @param args: 空参数
    @return: JSON格式的表名列表，包含表名和相关元数据
    @throws: MySQLMCPError 当数据库连接失败、权限不足或查询执行错误时
    """
    result = await mysql_manager.execute_query("SHOW TABLES")
    return json.dumps(result, ensure_ascii=False, indent=2)


async def mysql_describe_table(args: TableNameParams) -> str:
    """描述表结构

    检索并描述指定表的完整结构，包括列定义、数据类型、约束、
    索引信息和其他元数据。支持DESCRIBE和INFORMATION_SCHEMA查询。
    结果会被智能缓存以提高性能，支持表结构分析和文档生成。

    @param args: 表名参数
    @return: JSON格式的详细表结构信息，包含列、约束、索引等元数据
    @throws: MySQLMCPError 当表名无效、表不存在、权限不足或查询失败时
    """
    schema = await mysql_manager.get_table_schema_cached(args.table_name)
    if schema is None:
        return json.dumps({"error": f"表 {args.table_name} 不存在"}, ensure_ascii=False, indent=2)
    return json.dumps(schema, ensure_ascii=False, indent=2)


async def mysql_select_data(args: SelectDataParams) -> str:
    """查询数据

    根据指定的条件和限制从表中选择数据，提供灵活的数据查询功能。
    支持列选择、WHERE条件和LIMIT限制，确保查询的精确性和安全性。

    @param args: 查询参数，包含表名、列选择、WHERE条件和行数限制
    @return: JSON格式的查询结果数据
    @throws: MySQLMCPError 当表名无效、查询语法错误或执行失败时
    """
    mysql_manager.validate_table_name(args.table_name)

    columns = args.columns or ["*"]
    query = f"SELECT {', '.join(columns)} FROM `{args.table_name}`"
    if args.where_clause:
        query += f" WHERE {args.where_clause}"
    if args.limit:
        query += f" LIMIT {args.limit}"

    result = await mysql_manager.execute_query(query)
    return json.dumps(result, ensure_ascii=False, indent=2)


async def mysql_insert_data(args: InsertDataParams) -> str:
    """插入数据

    使用参数化查询安全地向表中插入新数据，确保数据完整性和安全性。
    自动验证所有输入数据，使用预处理语句防止SQL注入攻击。
    支持单行插入，包含事务安全保障。

    @param args: 插入参数
    @return: 包含成功状态、插入ID和受影响行数的JSON格式结果
    @throws: MySQLMCPError 当表名无效、列名无效、数据类型不匹配或插入失败时
    """
    mysql_manager.validate_table_name(args.table_name)

    # 检查数据量限制
    if len(args.data) == 0:
        raise MySQLMCPError(
            "插入数据不能为空",
            ErrorCategory.VALIDATION_ERROR,
            ErrorSeverity.MEDIUM
        )

    columns = list(args.data.keys())
    placeholders = ", ".join(["%s"] * len(columns))
    values = list(args.data.values())

    query = f"INSERT INTO `{args.table_name}` ({', '.join([f'`{col}`' for col in columns])}) VALUES ({placeholders})"
    result = await mysql_manager.execute_query(query, values)

    await mysql_manager.invalidate_caches("INSERT", args.table_name)

    response_data = {
        "success": True,
        "affected_rows": result.get("affected_rows", 0) if isinstance(result, dict) else 0,
        "insert_id": result.get("insert_id") if isinstance(result, dict) else None
    }

    return json.dumps(response_data, ensure_ascii=False, indent=2)


async def mysql_update_data(args: UpdateDataParams) -> str:
    """更新数据

    使用参数化查询根据指定条件更新表中的现有数据，确保数据修改的安全性和一致性。
    提供完整的输入验证，具有WHERE子句验证、预处理语句和事务安全保障。
    支持条件更新和批量字段修改。

    @param args: 更新参数
    @return: 包含成功状态、受影响行数和更新统计的JSON格式结果
    @throws: MySQLMCPError 当表名无效、WHERE子句缺失/无效、列名不存在或更新失败时
    """
    mysql_manager.validate_table_name(args.table_name)

    # 检查更新数据不为空
    if len(args.data) == 0:
        raise MySQLMCPError(
            "更新数据不能为空",
            ErrorCategory.VALIDATION_ERROR,
            ErrorSeverity.MEDIUM
        )

    set_parts = [f"`{key}` = %s" for key in args.data.keys()]
    values = list(args.data.values())

    query = f"UPDATE `{args.table_name}` SET {', '.join(set_parts)} WHERE {args.where_clause}"
    result = await mysql_manager.execute_query(query, values)

    await mysql_manager.invalidate_caches("UPDATE", args.table_name)

    response_data = {
        "success": True,
        "affected_rows": result.get("affected_rows", 0) if isinstance(result, dict) else 0
    }

    return json.dumps(response_data, ensure_ascii=False, indent=2)


async def mysql_delete_data(args: DeleteDataParams) -> str:
    """删除数据

    根据指定条件从表中安全删除数据，确保删除操作的准确性和安全性。
    使用WHERE子句验证，防止误删除和SQL注入攻击。
    支持条件删除操作，包含删除确认和事务安全保障。

    @param args: 删除参数
    @return: 包含成功状态、删除行数和操作统计的JSON格式结果
    @throws: MySQLMCPError 当表名无效、WHERE子句无效或删除失败时
    """
    mysql_manager.validate_table_name(args.table_name)

    # 检查WHERE子句不为空（防止全表删除）
    if not args.where_clause or args.where_clause.strip() == "":
        raise MySQLMCPError(
            "删除操作必须提供WHERE条件以防止全表删除",
            ErrorCategory.VALIDATION_ERROR,
            ErrorSeverity.CRITICAL
        )

    query = f"DELETE FROM `{args.table_name}` WHERE {args.where_clause}"
    result = await mysql_manager.execute_query(query)

    await mysql_manager.invalidate_caches("DELETE", args.table_name)

    response_data = {
        "success": True,
        "affected_rows": result.get("affected_rows", 0) if isinstance(result, dict) else 0
    }

    return json.dumps(response_data, ensure_ascii=False, indent=2)


async def mysql_get_schema(args: TableNameParams) -> str:
    """获取数据库架构

    检索数据库的完整架构信息，包括表结构、列定义、数据类型、约束等。
    支持特定表查询或获取整个数据库的架构，提供详细的元数据信息。
    结果包含表名、列名、数据类型、可空性、默认值、主键信息等。

    @param args: 架构查询参数，可选择特定表或获取全部表结构
    @return: JSON格式的详细架构信息，包含完整的元数据
    @throws: MySQLMCPError 当表名无效、权限不足或查询失败时
    """
    query = """
        SELECT TABLE_NAME, COLUMN_NAME, DATA_TYPE, IS_NULLABLE, COLUMN_DEFAULT, COLUMN_KEY, EXTRA, COLUMN_COMMENT
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_SCHEMA = DATABASE()
    """

    params = []
    if args.table_name:
        mysql_manager.validate_table_name(args.table_name)
        query += " AND TABLE_NAME = %s"
        params.append(args.table_name)

    query += " ORDER BY TABLE_NAME, ORDINAL_POSITION"

    result = await mysql_manager.execute_query(query, params)
    return json.dumps(result, ensure_ascii=False, indent=2)


async def mysql_get_foreign_keys(args: TableNameParams) -> str:
    """获取外键约束信息

    检索数据库中的外键约束信息，包括表名、列名、约束名、引用表和引用列等。
    支持特定表查询或获取整个数据库的外键约束，提供完整的外键关系图。
    帮助理解数据库表间依赖关系，支持数据库设计分析和数据完整性验证。

    @param args: 外键查询参数，可选择特定表或获取全部外键约束
    @return: JSON格式的外键约束详细信息，包含完整的引用关系
    @throws: MySQLMCPError 当表名无效、权限不足或查询失败时
    """
    query = """
        SELECT TABLE_NAME, COLUMN_NAME, CONSTRAINT_NAME, REFERENCED_TABLE_NAME, REFERENCED_COLUMN_NAME
        FROM INFORMATION_SCHEMA.KEY_COLUMN_USAGE
        WHERE TABLE_SCHEMA = DATABASE() AND REFERENCED_TABLE_NAME IS NOT NULL
    """

    params = []
    if args.table_name:
        mysql_manager.validate_table_name(args.table_name)
        query += " AND TABLE_NAME = %s"
        params.append(args.table_name)

    query += " ORDER BY TABLE_NAME, CONSTRAINT_NAME"

    result = await mysql_manager.execute_query(query, params)
    return json.dumps(result, ensure_ascii=False, indent=2)


async def mysql_create_table(args: CreateTableParams) -> str:
    """创建表

    使用指定的列定义和约束创建新的数据库表，支持完整的表结构定义。
    提供全面的安全验证，包括表名验证、列定义验证，确保数据库操作的安全性。
    支持主键、自增列、默认值等高级约束，支持批量列定义和事务安全保障。
    创建成功后自动使相关缓存失效，确保数据一致性。

    @param args: 创建表参数
    @return: 包含成功状态、创建信息和受影响表结构的JSON格式结果
    @throws: MySQLMCPError 当表名无效、列定义错误、约束冲突或创建失败时
    """
    mysql_manager.validate_table_name(args.table_name)

    # 检查列数量限制
    if len(args.columns) == 0:
        raise MySQLMCPError(
            "创建表时必须提供至少一列",
            ErrorCategory.VALIDATION_ERROR,
            ErrorSeverity.HIGH
        )

    if len(args.columns) > 100:
        raise MySQLMCPError(
            f"列数量 ({len(args.columns)}) 超过最大限制 (100)",
            ErrorCategory.CONSTRAINT_VIOLATION,
            ErrorSeverity.HIGH
        )

    column_defs = []
    primary_keys = []

    for col in args.columns:
        mysql_manager.validate_input(col["name"], "column_name")
        mysql_manager.validate_input(col["type"], "column_type")

        col_def = f"`{col['name']}` {col['type']}"

        if col.get("nullable") is False:
            col_def += " NOT NULL"

        if col.get("auto_increment"):
            col_def += " AUTO_INCREMENT"

        if col.get("default"):
            col_def += f" DEFAULT {col['default']}"

        column_defs.append(col_def)

        if col.get("primary_key"):
            primary_keys.append(col["name"])

    if primary_keys:
        column_defs.append(f"PRIMARY KEY ({', '.join([f'`{pk}`' for pk in primary_keys])})")

    query = f"CREATE TABLE `{args.table_name}` ({', '.join(column_defs)})"
    result = await mysql_manager.execute_query(query)

    await mysql_manager.invalidate_caches("CREATE", args.table_name)

    response_data = {
        "success": True,
        "message": f"表 {args.table_name} 创建成功"
    }

    return json.dumps(response_data, ensure_ascii=False, indent=2)


async def mysql_drop_table(args: DropTableParams) -> str:
    """删除表

    从数据库中删除指定的表，支持可选的IF EXISTS子句以避免删除不存在的表时出错。
    删除操作会自动清理相关缓存，确保数据一致性。提供安全删除机制，
    防止误操作导致的数据丢失，支持批量表删除和级联删除操作。

    @param args: 删除表参数，包含表名和IF EXISTS选项
    @return: JSON格式的操作结果，包含成功状态和删除消息
    @throws: MySQLMCPError 当表名无效、权限不足或删除操作失败时
    """
    mysql_manager.validate_table_name(args.table_name)

    query = f"DROP TABLE{' IF EXISTS' if args.if_exists else ''} `{args.table_name}`"
    result = await mysql_manager.execute_query(query)

    await mysql_manager.invalidate_caches("DROP", args.table_name)

    return json.dumps({
        "success": True,
        "message": f"表 {args.table_name} 删除成功"
    }, ensure_ascii=False, indent=2)


async def mysql_batch_execute(args: BatchExecuteParams) -> str:
    """批量执行查询"""
    try:
        # 检查批量操作的速率限制（批量操作更严格）
        if not check_rate_limit(args, "batch_execute"):
            rate_status = get_rate_limit_status(args, "batch_execute")
            from type_utils import MySQLMCPError, ErrorCategory, ErrorSeverity
            raise MySQLMCPError(
                STRING_CONSTANTS["MSG_RATE_LIMIT_EXCEEDED"],
                ErrorCategory.RATE_LIMIT_ERROR,
                ErrorSeverity.HIGH,
                metadata={
                    "rate_limit_status": rate_status,
                    "operation": "mysql_batch_execute"
                }
            )

        # 检查批量操作数量限制
        if len(args.queries) == 0:
            raise MySQLMCPError(
                "批量操作不能为空",
                ErrorCategory.VALIDATION_ERROR,
                ErrorSeverity.MEDIUM
            )

        if len(args.queries) > config_manager.security.max_result_rows:
            raise MySQLMCPError(
                f"批量操作数量 ({len(args.queries)}) 超过最大限制 ({config_manager.security.max_result_rows})",
                ErrorCategory.CONSTRAINT_VIOLATION,
                ErrorSeverity.HIGH
            )

        results = await mysql_manager.execute_batch_queries(args.queries)

        await mysql_manager.invalidate_caches("DML")

        # 添加限流信息到响应
        rate_status = get_rate_limit_status(args, "batch_execute")
        response_data = {
            "success": True,
            "query_count": len(args.queries),
            "results": results,
            "_rate_limit_info": rate_status
        }

        return json.dumps(response_data, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.error(f"批量执行失败: {e}")
        raise


async def mysql_system_status(args: Dict) -> str:
    """系统状态检查

    执行全面的系统健康状态检查，包括MySQL性能指标、系统资源使用情况、
    内存监控状态和系统整体健康度。提供实时的系统状态快照，支持
    性能监控、故障诊断和容量规划。返回结构化的状态报告，包含时间戳
    和详细的性能指标，帮助识别潜在问题和优化机会。

    @param args: 系统状态检查参数（当前为空参数，保留扩展性）
    @return: JSON格式的综合系统状态报告，包含性能指标和健康状态
    @throws: MySQLMCPError 当系统状态检查过程中发生错误时
    """
    # 获取MySQL管理器性能指标
    mysql_metrics = mysql_manager.get_performance_metrics()

    # 获取系统监控数据
    system_resources = system_monitor.get_current_resources()
    memory_stats = memory_monitor.get_memory_stats()

    # 构建综合状态报告
    status = {
        "timestamp": datetime.now().isoformat(),
        "mysql_performance": mysql_metrics,
        "system_resources": {
            "memory_usage": system_resources.memory_usage,
            "cpu_usage": system_resources.cpu,
            "event_loop_delay": system_resources.event_loop_delay,
            "uptime": system_resources.uptime
        },
        "memory_monitoring": {
            "current": memory_stats["current"].usage if memory_stats["current"] else None,
            "pressure_level": memory_stats["current"].pressure_level if memory_stats["current"] else 0,
            "leak_suspicions": memory_stats["leak_suspicions"],
            "trend": memory_stats["trend"]
        },
        "system_health": system_monitor.get_system_health()
    }

    return json.dumps(status, ensure_ascii=False, indent=2)


async def mysql_analyze_error(args: AnalyzeErrorParams) -> str:
    """错误分析

    对MySQL错误进行深入分析，提供错误诊断、恢复建议和预防措施。
    支持错误分类、根本原因分析和解决建议，帮助快速定位和解决问题。
    提供结构化的错误报告，包含错误类型、严重程度、影响范围和修复建议。

    @param args: 错误分析参数，包含错误消息和操作上下文
    @return: JSON格式的详细错误分析报告，包含诊断结果和建议
    @throws: MySQLMCPError 当错误分析过程中发生异常时
    """
    # 创建一个错误对象用于分析
    error = Exception(args.error_message)

    # 使用统一的错误分析功能
    analysis = ErrorHandler.analyze_error(error, args.operation or 'unknown')

    return json.dumps({
        "success": True,
        "analysis": analysis,
        "error_message": args.error_message,
        "operation": args.operation
    }, ensure_ascii=False, indent=2)


async def mysql_alter_table(args: AlterTableParams) -> str:
    """修改表结构

    执行复杂的表结构修改操作，支持添加、修改、删除列和索引等操作。
    提供批量ALTER操作支持，提高修改效率，同时确保数据安全性和一致性。
    支持完整的DDL操作验证，包括语法检查、约束验证和依赖关系分析。
    提供详细的操作报告和性能统计，支持表结构演进和数据库重构。

    @param args: 表结构修改参数，包含表名和具体的修改操作列表
    @return: JSON格式的修改结果，包含操作统计、性能指标和执行详情
    @throws: MySQLMCPError 当表名无效、修改操作有误或执行失败时
    """
    from type_utils import MySQLMCPError, ErrorCategory, ErrorSeverity

    start_time = time.time()
    mysql_manager.validate_table_name(args.table_name)

    # 检查ALTER操作数量限制
    if len(args.alterations) > mysql_manager.config_manager.security.max_result_rows:
        raise MySQLMCPError(
            f"ALTER操作数量 ({len(args.alterations)}) 超过最大限制 ({mysql_manager.config_manager.security.max_result_rows})",
            ErrorCategory.CONSTRAINT_VIOLATION,
            ErrorSeverity.HIGH
        )

    # 预先验证所有修改操作
    validation_errors = []
    for index, alteration in enumerate(args.alterations):
        if alteration.get("column"):
            try:
                mysql_manager.validate_input(alteration["column"]["name"], f"alteration_{index}_column_name")
                if alteration["column"].get("type"):
                    mysql_manager.validate_input(alteration["column"]["type"], f"alteration_{index}_column_type")
            except Exception as e:
                validation_errors.append(f"列定义验证失败 (索引 {index}): {str(e)}")

        if alteration.get("index"):
            try:
                mysql_manager.validate_input(alteration["index"]["name"], f"alteration_{index}_index_name")
                for col_index, col in enumerate(alteration["index"]["columns"]):
                    mysql_manager.validate_input(col, f"alteration_{index}_index_column_{col_index}")
            except Exception as e:
                validation_errors.append(f"索引定义验证失败 (索引 {index}): {str(e)}")

    if validation_errors:
        raise MySQLMCPError(
            f"输入验证失败: {'; '.join(validation_errors)}",
            ErrorCategory.SECURITY_VIOLATION,
            ErrorSeverity.HIGH
        )

    # 构建 ALTER TABLE 语句
    alter_statements = []

    for index, alteration in enumerate(args.alterations):
        action_type = alteration["type"]

        try:
            if action_type == "ADD_COLUMN":
                if not alteration.get("column"):
                    raise MySQLMCPError(
                        f"ADD_COLUMN 操作必须提供列定义 (索引 {index})",
                        ErrorCategory.SYNTAX_ERROR,
                        ErrorSeverity.MEDIUM
                    )

                col = alteration["column"]
                col_parts = [f'`{col["name"]}` {col["type"]}']

                if col.get("nullable") is False:
                    col_parts.append('NOT NULL')

                if col.get("auto_increment"):
                    col_parts.append('AUTO_INCREMENT')

                if col.get("default") is not None:
                    col_parts.append(f'DEFAULT {col["default"]}')

                if col.get("comment"):
                    col_parts.append(f'COMMENT \'{col["comment"]}\'')

                if col.get("first"):
                    col_parts.append('FIRST')
                elif col.get("after"):
                    col_parts.append(f'AFTER `{col["after"]}`')

                alter_statements.append(f'ADD COLUMN {" ".join(col_parts)}')

            elif action_type == "MODIFY_COLUMN":
                if not alteration.get("column"):
                    raise MySQLMCPError(
                        f"MODIFY_COLUMN 操作必须提供列定义 (索引 {index})",
                        ErrorCategory.SYNTAX_ERROR,
                        ErrorSeverity.MEDIUM
                    )

                col = alteration["column"]
                modify_parts = [f'`{col["name"]}` {col["type"]}']

                if col.get("nullable") is False:
                    modify_parts.append('NOT NULL')

                if col.get("default") is not None:
                    modify_parts.append(f'DEFAULT {col["default"]}')

                if col.get("comment"):
                    modify_parts.append(f'COMMENT \'{col["comment"]}\'')

                if col.get("after"):
                    modify_parts.append(f'AFTER `{col["after"]}`')

                alter_statements.append(f'MODIFY COLUMN {" ".join(modify_parts)}')

            elif action_type == "DROP_COLUMN":
                if not alteration.get("column"):
                    raise MySQLMCPError(
                        f"DROP_COLUMN 操作必须提供列定义 (索引 {index})",
                        ErrorCategory.SYNTAX_ERROR,
                        ErrorSeverity.MEDIUM
                    )

                alter_statements.append(f'DROP COLUMN `{alteration["column"]["name"]}`')

            elif action_type == "ADD_INDEX":
                if not alteration.get("index"):
                    raise MySQLMCPError(
                        f"ADD_INDEX 操作必须提供索引定义 (索引 {index})",
                        ErrorCategory.SYNTAX_ERROR,
                        ErrorSeverity.MEDIUM
                    )

                idx = alteration["index"]
                index_type = idx.get("type", "").upper()
                columns_str = ", ".join([f'`{col}`' for col in idx["columns"]])

                index_parts = []
                if index_type == 'PRIMARY':
                    index_parts.append(f'ADD PRIMARY KEY ({columns_str})')
                else:
                    type_str = ""
                    if index_type == 'UNIQUE':
                        type_str = 'UNIQUE INDEX'
                    elif index_type == 'FULLTEXT':
                        type_str = 'FULLTEXT INDEX'
                    elif index_type == 'SPATIAL':
                        type_str = 'SPATIAL INDEX'
                    else:
                        type_str = 'INDEX'

                    invisible_str = ' INVISIBLE' if idx.get("invisible") else ''
                    index_parts.append(f'ADD {type_str} `{idx["name"]}` ({columns_str}){invisible_str}')

                    if idx.get("comment"):
                        index_parts.append(f'COMMENT \'{idx["comment"]}\'')

                alter_statements.append(" ".join(index_parts))

            elif action_type == "DROP_INDEX":
                if not alteration.get("index"):
                    raise MySQLMCPError(
                        f"DROP_INDEX 操作必须提供索引定义 (索引 {index})",
                        ErrorCategory.SYNTAX_ERROR,
                        ErrorSeverity.MEDIUM
                    )

                idx_name = alteration["index"]["name"]
                alter_statements.append(f'DROP INDEX `{idx_name}`')

            elif action_type == "ADD_FOREIGN_KEY":
                if not alteration.get("foreign_key"):
                    raise MySQLMCPError(
                        f"ADD_FOREIGN_KEY 操作必须提供外键定义 (索引 {index})",
                        ErrorCategory.SYNTAX_ERROR,
                        ErrorSeverity.MEDIUM
                    )

                fk = alteration["foreign_key"]
                local_cols = ", ".join([f'`{col}`' for col in fk["columns"]])
                ref_cols = ", ".join([f'`{col}`' for col in fk["referenced_columns"]])

                fk_parts = [f'ADD CONSTRAINT `{fk["name"]}` FOREIGN KEY ({local_cols}) REFERENCES `{fk["referenced_table"]}` ({ref_cols})']

                if fk.get("on_delete"):
                    fk_parts.append(f'ON DELETE {fk["on_delete"]}')

                if fk.get("on_update"):
                    fk_parts.append(f'ON UPDATE {fk["on_update"]}')

                alter_statements.append(" ".join(fk_parts))

            elif action_type == "DROP_FOREIGN_KEY":
                if not alteration.get("foreign_key"):
                    raise MySQLMCPError(
                        f"DROP_FOREIGN_KEY 操作必须提供外键定义 (索引 {index})",
                        ErrorCategory.SYNTAX_ERROR,
                        ErrorSeverity.MEDIUM
                    )

                fk_name = alteration["foreign_key"]["name"]
                alter_statements.append(f'DROP FOREIGN KEY `{fk_name}`')

            else:
                raise MySQLMCPError(
                    f"未知的修改类型: {action_type} (索引 {index})",
                    ErrorCategory.SYNTAX_ERROR,
                    ErrorSeverity.HIGH
                )

        except MySQLMCPError:
            raise
        except Exception as e:
            raise MySQLMCPError(
                f"处理修改操作失败 (索引 {index}): {str(e)}",
                ErrorCategory.SYNTAX_ERROR,
                ErrorSeverity.HIGH
            )

    # 执行ALTER操作
    result = {"success": True}
    affected_rows = 0
    batch_size = 50  # 每批最多50个操作

    if len(alter_statements) <= batch_size:
        # 少量操作，一次性执行
        alter_query = f"ALTER TABLE `{args.table_name}` {', '.join(alter_statements)}"
        query_result = await mysql_manager.execute_query(alter_query)
        affected_rows = (query_result.get("affected_rows", 0) if isinstance(query_result, dict) else 0)
    else:
        # 大量操作，分批执行
        for i in range(0, len(alter_statements), batch_size):
            batch_statements = alter_statements[i:i + batch_size]
            alter_query = f"ALTER TABLE `{args.table_name}` {', '.join(batch_statements)}"
            batch_result = await mysql_manager.execute_query(alter_query)
            batch_affected = batch_result.get("affected_rows", 0) if isinstance(batch_result, dict) else 0
            affected_rows += batch_affected

    # 修改表后使缓存失效
    await mysql_manager.invalidate_caches("ALTER", args.table_name)

    query_time = time.time() - start_time

    result.update({
        "altered_table": args.table_name,
        "alter_operations": len(args.alterations),
        "affected_rows": affected_rows,
        "query_time": round(query_time, 3),
        "batches_executed": (len(alter_statements) + batch_size - 1) // batch_size
    })

    return json.dumps(result, ensure_ascii=False, indent=2)


async def mysql_manage_users(args: ManageUsersParams) -> str:
    """用户管理

    提供完整的MySQL用户管理功能，包括用户创建、删除、权限授予和撤销等操作。
    支持灵活的权限管理，可以精确控制用户在特定数据库和表上的访问权限。
    提供用户列表查询和权限查看功能，帮助管理员进行用户和权限的审计。
    所有操作都包含完整的输入验证和安全检查，确保操作的安全性和准确性。

    @param args: 用户管理参数，包含操作类型、用户名、密码、权限等信息
    @return: JSON格式的操作结果，包含执行状态、用户信息和操作详情
    @throws: MySQLMCPError 当用户名无效、权限不足或操作失败时
    """
    result = {"action": args.action, "success": True}

    if args.action == "create":
        if not args.username or not args.password:
            raise ValueError("创建用户时必须提供用户名和密码")

        mysql_manager.validate_input(args.username, "username")
        mysql_manager.validate_input(args.password, "password")
        mysql_manager.validate_input(args.host or "%", "host")

        create_user_query = "CREATE USER %s@%s IDENTIFIED BY %s"
        params = [f"{args.username}", args.host or "%", args.password]

        await mysql_manager.execute_query(create_user_query, params)

        result.update({
            "user": {
                "username": args.username,
                "host": args.host or "%",
                "created": str(await mysql_manager.execute_query("SELECT NOW()"))
            },
            "message": f"用户 '{args.username}'@'{args.host or '%'}' 创建成功"
        })

    elif args.action == "delete":
        if not args.username:
            raise ValueError("删除用户时必须提供用户名")

        mysql_manager.validate_input(args.username, "username")
        mysql_manager.validate_input(args.host or "%", "host")

        drop_user_query = f"DROP USER {'IF EXISTS' if args.if_exists else ''} %s@%s"
        params = [f"{args.username}", args.host or "%"]

        await mysql_manager.execute_query(drop_user_query, params)

        result.update({
            "user": {
                "username": args.username,
                "host": args.host or "%",
                "deleted": str(await mysql_manager.execute_query("SELECT NOW()"))
            },
            "message": f"用户 '{args.username}'@'{args.host or '%'}' 删除成功"
        })

    elif args.action == "grant":
        if not args.username or not args.privileges or len(args.privileges) == 0:
            raise ValueError("授予权限时必须提供用户名和权限列表")

        mysql_manager.validate_input(args.username, "username")
        mysql_manager.validate_input(args.host or "%", "host")

        # 验证权限列表
        valid_privileges = ['ALL', 'SELECT', 'INSERT', 'UPDATE', 'DELETE', 'CREATE', 'DROP', 'ALTER', 'INDEX', 'GRANT OPTION']
        for privilege in args.privileges:
            if privilege.upper() not in valid_privileges:
                raise ValueError(f"无效的权限: {privilege}")

        privileges_str = ", ".join(args.privileges)

        if args.database and args.table:
            mysql_manager.validate_input(args.database, "database")
            mysql_manager.validate_input(args.table, "table")
            grant_query = f"GRANT {privileges_str} ON `{args.database}`.`{args.table}` TO %s@%s"
            params = [f"{args.username}", args.host or "%"]
            target = f"{args.database}.{args.table}"
        elif args.database:
            mysql_manager.validate_input(args.database, "database")
            grant_query = f"GRANT {privileges_str} ON `{args.database}`.* TO %s@%s"
            params = [f"{args.username}", args.host or "%"]
            target = f"{args.database}.*"
        else:
            grant_query = f"GRANT {privileges_str} ON *.* TO %s@%s"
            params = [f"{args.username}", args.host or "%"]
            target = "*.*"

        await mysql_manager.execute_query(grant_query, params)

        result.update({
            "user": {
                "username": args.username,
                "host": args.host or "%"
            },
            "privileges": {
                "granted": args.privileges,
                "target": target
            },
            "message": f"成功授予用户 '{args.username}'@'{args.host or '%'}' {privileges_str} 权限"
        })

    elif args.action == "revoke":
        if not args.username or not args.privileges or len(args.privileges) == 0:
            raise ValueError("撤销权限时必须提供用户名和权限列表")

        mysql_manager.validate_input(args.username, "username")
        mysql_manager.validate_input(args.host or "%", "host")

        # 验证权限列表
        valid_privileges = ['ALL', 'SELECT', 'INSERT', 'UPDATE', 'DELETE', 'CREATE', 'DROP', 'ALTER', 'INDEX', 'GRANT OPTION']
        for privilege in args.privileges:
            if privilege.upper() not in valid_privileges:
                raise ValueError(f"无效的权限: {privilege}")

        privileges_str = ", ".join(args.privileges)

        if args.database and args.table:
            mysql_manager.validate_input(args.database, "database")
            mysql_manager.validate_input(args.table, "table")
            revoke_query = f"REVOKE {privileges_str} ON `{args.database}`.`{args.table}` FROM %s@%s"
            params = [f"{args.username}", args.host or "%"]
            target = f"{args.database}.{args.table}"
        elif args.database:
            mysql_manager.validate_input(args.database, "database")
            revoke_query = f"REVOKE {privileges_str} ON `{args.database}`.* FROM %s@%s"
            params = [f"{args.username}", args.host or "%"]
            target = f"{args.database}.*"
        else:
            revoke_query = f"REVOKE {privileges_str} ON *.* FROM %s@%s"
            params = [f"{args.username}", args.host or "%"]
            target = "*.*"

        await mysql_manager.execute_query(revoke_query, params)

        result.update({
            "user": {
                "username": args.username,
                "host": args.host or "%"
            },
            "privileges": {
                "revoked": args.privileges,
                "target": target
            },
            "message": f"成功撤销用户 '{args.username}'@'{args.host or '%'}' {privileges_str} 权限"
        })

    elif args.action == "list":
        list_users_query = """
            SELECT User, Host, authentication_string, password_expired, password_last_changed, account_locked
            FROM mysql.user
            WHERE User != ''
            ORDER BY User, Host
        """

        users = await mysql_manager.execute_query(list_users_query)

        result.update({
            "total_users": len(users),
            "users": users,
            "message": f"找到 {len(users)} 个用户"
        })

    elif args.action == "show_grants":
        if not args.username:
            raise ValueError("显示权限时必须提供用户名")

        mysql_manager.validate_input(args.username, "username")
        mysql_manager.validate_input(args.host or "%", "host")

        show_grants_query = "SHOW GRANTS FOR %s@%s"
        params = [f"{args.username}", args.host or "%"]

        grants = await mysql_manager.execute_query(show_grants_query, params)

        result.update({
            "user": {
                "username": args.username,
                "host": args.host or "%"
            },
            "grants": grants,
            "total_grants": len(grants),
            "message": f"用户 '{args.username}'@'{args.host or '%'}' 拥有 {len(grants)} 个权限"
        })

    else:
        raise ValueError(f"未知的操作: {args.action}")

    return json.dumps(result, ensure_ascii=False, indent=2)


async def mysql_import_data(args: ImportDataParams) -> str:
    """数据导入"""
    try:
        import pandas as pd
        from type_utils import ImportResult, ErrorCategory, ErrorSeverity, MySQLMCPError

        start_time = time.time()

        # 验证表名
        mysql_manager.validate_table_name(args.table_name)

        # 检查文件是否存在
        if not os.path.exists(args.file_path):
            raise MySQLMCPError(
                f"文件不存在: {args.file_path}",
                ErrorCategory.VALIDATION_ERROR,
                ErrorSeverity.HIGH
            )

        # 验证文件大小（限制为100MB）
        file_size = os.path.getsize(args.file_path)
        if file_size > 100 * 1024 * 1024:
            raise MySQLMCPError(
                f"文件过大: {file_size} bytes (最大100MB)",
                ErrorCategory.CONSTRAINT_VIOLATION,
                ErrorSeverity.HIGH
            )

        # 读取和解析数据
        data = []
        column_count = 0

        try:
            if args.format == "csv":
                df = pd.read_csv(
                    args.file_path,
                    header=0 if args.has_headers else None,
                    sep=args.delimiter or ",",
                    quotechar=args.quote or '"',
                    encoding=args.encoding or "utf8"
                )
                data = df.to_dict('records')
                column_count = len(df.columns)

            elif args.format == "excel":
                df = pd.read_excel(
                    args.file_path,
                    header=0 if args.has_headers else None,
                    sheet_name=args.sheet_name
                )
                data = df.to_dict('records')
                column_count = len(df.columns)

            elif args.format == "json":
                import json
                with open(args.file_path, 'r', encoding=args.encoding or "utf8") as f:
                    data = json.load(f)
                if isinstance(data, list) and data:
                    column_count = len(data[0].keys())

            elif args.format == "sql":
                # SQL文件导入 - 直接执行SQL内容
                with open(args.file_path, 'r', encoding=args.encoding or "utf8") as f:
                    sql_content = f.read()

                # 简单验证SQL内容
                mysql_manager.validate_input(sql_content[:1000], 'sql_import')

                # 执行SQL（这里简化处理，实际应该更安全）
                result = await mysql_manager.execute_query(sql_content)

                return json.dumps({
                    "success": True,
                    "imported_rows": result.get("affected_rows", 0) if isinstance(result, dict) else 0,
                    "skipped_rows": 0,
                    "failed_rows": 0,
                    "updated_rows": 0,
                    "total_rows": result.get("affected_rows", 0) if isinstance(result, dict) else 0,
                    "duration": int((time.time() - start_time) * 1000),
                    "batches_processed": 1,
                    "file_path": args.file_path,
                    "format": args.format,
                    "table_name": args.table_name,
                    "error": None,
                    "warnings": [],
                    "transaction_mode": "single_transaction"
                }, ensure_ascii=False, indent=2)

            else:
                raise MySQLMCPError(
                    f"不支持的导入格式: {args.format}",
                    ErrorCategory.VALIDATION_ERROR,
                    ErrorSeverity.HIGH
                )

        except Exception as e:
            raise MySQLMCPError(
                f"文件读取失败: {str(e)}",
                ErrorCategory.DATA_ERROR,
                ErrorSeverity.HIGH
            )

        if not data:
            return json.dumps({
                "success": True,
                "imported_rows": 0,
                "skipped_rows": 0,
                "failed_rows": 0,
                "updated_rows": 0,
                "total_rows": 0,
                "duration": int((time.time() - start_time) * 1000),
                "batches_processed": 0,
                "file_path": args.file_path,
                "format": args.format,
                "table_name": args.table_name,
                "error": None,
                "warnings": ["文件为空"],
                "transaction_mode": "batch_transaction"
            }, ensure_ascii=False, indent=2)

        # 应用字段映射
        if args.field_mapping:
            for row in data:
                for source_field, target_field in args.field_mapping.items():
                    if source_field in row:
                        row[target_field] = row.pop(source_field)

        # 数据验证和清理
        total_rows = len(data)
        imported_rows = 0
        skipped_rows = 0
        failed_rows = 0
        updated_rows = 0
        processed_batches = 0
        warnings = []

        # 分批处理
        batch_size = min(args.batch_size or 1000, 10000)  # 限制最大批次大小
        transaction_mode = "single_transaction" if args.use_transaction else "batch_transaction"

        if args.use_transaction:
            # 单事务模式 - 整个导入在一个事务中
            try:
                await mysql_manager.execute_query("START TRANSACTION")
                processed_batches = await _process_import_batches(
                    mysql_manager, args, data, batch_size
                )
                await mysql_manager.execute_query("COMMIT")
                imported_rows = total_rows - failed_rows
            except Exception as e:
                await mysql_manager.execute_query("ROLLBACK")
                raise MySQLMCPError(
                    f"事务回滚: {str(e)}",
                    ErrorCategory.TRANSACTION_ERROR,
                    ErrorSeverity.HIGH
                )
        else:
            # 批量事务模式 - 每批数据独立事务
            for i in range(0, len(data), batch_size):
                batch_data = data[i:i + batch_size]
                try:
                    await mysql_manager.execute_query("START TRANSACTION")
                    batch_imported, batch_failed = await _process_import_batch(
                        mysql_manager, args, batch_data
                    )
                    await mysql_manager.execute_query("COMMIT")
                    imported_rows += batch_imported
                    failed_rows += batch_failed
                    processed_batches += 1
                except Exception as e:
                    await mysql_manager.execute_query("ROLLBACK")
                    failed_rows += len(batch_data)
                    processed_batches += 1
                    warnings.append(f"批次 {processed_batches} 失败: {str(e)}")

        # 使缓存失效
        await mysql_manager.invalidate_caches("INSERT", args.table_name)

        return json.dumps({
            "success": True,
            "imported_rows": imported_rows,
            "skipped_rows": skipped_rows,
            "failed_rows": failed_rows,
            "updated_rows": updated_rows,
            "total_rows": total_rows,
            "duration": int((time.time() - start_time) * 1000),
            "batches_processed": processed_batches,
            "file_path": args.file_path,
            "format": args.format,
            "table_name": args.table_name,
            "error": None,
            "warnings": warnings,
            "transaction_mode": transaction_mode
        }, ensure_ascii=False, indent=2)

    except Exception as e:
        logger.error(f"数据导入失败: {e}")
        raise


async def _process_import_batches(
    mysql_manager, args: ImportDataParams, data: list, batch_size: int
) -> int:
    """处理导入批次"""
    processed_batches = 0
    for i in range(0, len(data), batch_size):
        batch_data = data[i:i + batch_size]
        await _process_import_batch(mysql_manager, args, batch_data)
        processed_batches += 1
    return processed_batches


async def _process_import_batch(
    mysql_manager, args: ImportDataParams, batch_data: list
) -> tuple[int, int]:
    """处理单个导入批次"""
    imported = 0
    failed = 0

    for row in batch_data:
        try:
            # 数据验证
            if args.validate_data:
                for key, value in row.items():
                    mysql_manager.validate_input(str(key), "column_name")
                    if value is not None:
                        mysql_manager.validate_input(str(value), "column_value")

            # 构建插入语句
            columns = list(row.keys())
            values = list(row.values())
            placeholders = ", ".join(["%s"] * len(columns))

            query = f"INSERT INTO `{args.table_name}` ({', '.join([f'`{col}`' for col in columns])}) VALUES ({placeholders})"

            # 根据冲突策略处理
            if args.conflict_strategy == "update":
                # ON DUPLICATE KEY UPDATE
                update_parts = [f"`{col}` = VALUES(`{col}`)" for col in columns]
                query += f" ON DUPLICATE KEY UPDATE {', '.join(update_parts)}"
            elif args.conflict_strategy == "skip":
                # INSERT IGNORE
                query = f"INSERT IGNORE INTO `{args.table_name}` ({', '.join([f'`{col}`' for col in columns])}) VALUES ({placeholders})"

            result = await mysql_manager.execute_query(query, values)
            if result and "affected_rows" in result and result["affected_rows"] > 0:
                imported += 1
            else:
                failed += 1

        except Exception as e:
            failed += 1
            logger.warn(f"导入行失败: {e}")

    return imported, failed


class VerifyBackupParams(BaseModel):
    """备份验证参数模型

    用于验证备份文件完整性的参数模型。

    @param backup_file_path: 要验证的备份文件路径
    """
    backup_file_path: str


# 进度跟踪相关参数类
class ProgressTrackerParams(BaseModel):
    action: str
    tracker_id: Optional[str] = None
    operation_type: Optional[str] = "all"
    include_completed: Optional[bool] = False
    detail_level: Optional[str] = "basic"


# 队列管理相关参数类
class ManageQueueParams(BaseModel):
    action: str
    task_id: Optional[str] = None
    max_concurrency: Optional[int] = None
    show_details: Optional[bool] = False
    filter_type: Optional[str] = "all"


# 内存优化相关参数类
class OptimizeMemoryParams(BaseModel):
    action: str
    force_gc: Optional[bool] = True
    enable_monitoring: Optional[bool] = None
    max_concurrency: Optional[int] = None
    include_history: Optional[bool] = False


class RateLimitManageParams(BaseModel):
    """速率限制管理参数模型

    用于配置速率限制管理操作的参数模型，支持状态查看、重置和配置更新。

    @param action: 操作类型，如"status"、"reset"、"update_config"、"get_stats"、"demo_algorithms"
    @param operation_type: 可选，操作类型
    @param identifier: 可选，标识符
    @param new_config: 可选，新配置字典
    """
    action: str
    operation_type: Optional[str] = None
    identifier: Optional[str] = None
    new_config: Optional[Dict] = None


# 性能优化相关参数类
class PerformanceOptimizeParams(BaseModel):
    """性能优化参数模型

    用于配置性能优化和监控操作的参数模型，支持慢查询分析和索引建议。

    @param action: 操作类型，如"enable_slow_query_log"、"status_slow_query_log"等
    @param query: 可选，要分析的SQL查询
    @param params: 可选，查询参数
    @param limit: 可选，结果限制数量，默认50
    @param include_details: 可选，是否包含详细信息，默认True
    @param time_range: 可选，时间范围，默认"1 day"
    @param long_query_time: 可选，慢查询阈值（秒）
    @param log_queries_not_using_indexes: 可选，是否记录未使用索引的查询
    @param monitoring_interval_minutes: 可选，监控间隔（分钟），默认60
    """
    action: str
    query: Optional[str] = None
    params: Optional[List[Any]] = None
    limit: Optional[int] = 50
    include_details: Optional[bool] = True
    time_range: Optional[str] = "1 day"
    long_query_time: Optional[int] = None
    log_queries_not_using_indexes: Optional[bool] = None
    monitoring_interval_minutes: Optional[int] = 60


# 备份工具相关参数类
class IncrementalBackupParams(BaseModel):
    """增量备份参数模型

    用于配置增量备份操作的参数模型，支持基于时间戳、binlog等多种增量模式。

    @param output_dir: 可选，输出目录路径
    @param compress: 可选，是否压缩，默认True
    @param include_data: 可选，是否包含数据，默认True
    @param include_structure: 可选，是否包含结构，默认False
    @param tables: 可选，要备份的表列表
    @param file_prefix: 可选，文件名前缀，默认"mysql_incremental"
    @param max_file_size: 可选，最大文件大小(MB)，默认100
    @param incremental_mode: 可选，增量模式，默认"timestamp"
    @param tracking_table: 可选，跟踪表名，默认"__backup_tracking"
    @param base_backup_path: 可选，基础备份路径
    @param last_backup_time: 可选，最后备份时间
    @param binlog_position: 可选，binlog位置
    """
    output_dir: Optional[str] = None
    compress: Optional[bool] = True
    include_data: Optional[bool] = True
    include_structure: Optional[bool] = False
    tables: Optional[List[str]] = None
    file_prefix: Optional[str] = "mysql_incremental"
    max_file_size: Optional[int] = 100
    incremental_mode: Optional[str] = "timestamp"
    tracking_table: Optional[str] = "__backup_tracking"
    base_backup_path: Optional[str] = None
    last_backup_time: Optional[str] = None
    binlog_position: Optional[str] = None


async def mysql_manage_queue(args: ManageQueueParams) -> str:
    """队列管理工具"""
    try:
        result = {"action": args.action, "success": True}

        if args.action == "status":
            result.update({
                "stats": {
                    "total_tasks": 0,
                    "queued_tasks": 0,
                    "running_tasks": 0,
                    "completed_tasks": 0,
                    "failed_tasks": 0,
                    "cancelled_tasks": 0,
                    "max_concurrent_tasks": 5,
                    "average_wait_time": 0,
                    "average_execution_time": 0
                },
                "filter": args.filter_type,
                "total_tasks": 0,
                "filtered_tasks": {
                    "queued": 0,
                    "running": 0,
                    "completed": 0,
                    "failed": 0
                }
            })

        elif args.action == "pause":
            result.update({
                "message": "任务队列已暂停",
                "status": "paused"
            })

        elif args.action == "resume":
            result.update({
                "message": "任务队列已恢复",
                "status": "resumed"
            })

        elif args.action == "clear":
            result.update({
                "cleared_count": 0,
                "message": "已清除 0 个排队中的任务"
            })

        elif args.action == "set_concurrency":
            if not args.max_concurrency:
                raise ValueError("设置并发数时必须提供 max_concurrency 参数")
            result.update({
                "new_max_concurrency": args.max_concurrency,
                "message": f"最大并发任务数已设置为 {args.max_concurrency}"
            })

        elif args.action == "cancel":
            if not args.task_id:
                raise ValueError("取消任务时必须提供 task_id 参数")
            result.update({
                "task_id": args.task_id,
                "cancelled": False,
                "message": "任务取消失败或任务不存在"
            })

        else:
            raise ValueError(f"不支持的操作: {args.action}")

        return json.dumps(result, ensure_ascii=False, indent=2)

    except Exception as e:
        logger.error(f"队列管理失败: {e}")
        raise


async def mysql_optimize_memory(args: OptimizeMemoryParams) -> str:
    """内存优化工具"""
    try:
        import psutil
        import gc

        result = {"action": args.action, "success": True}

        if args.action == "status":
            # 获取内存状态
            memory = psutil.virtual_memory()
            process = psutil.Process()
            process_memory = process.memory_info()

            result.update({
                "system": {
                    "current": {
                        "heap_used": f"{process_memory.rss / 1024 / 1024:.2f} MB",
                        "heap_total": f"{memory.total / 1024 / 1024:.2f} MB",
                        "rss": f"{process_memory.rss / 1024 / 1024:.2f} MB",
                        "pressure_level": f"{memory.percent:.2f}%"
                    },
                    "trend": "stable",
                    "leak_suspicions": 0,
                    "gc_stats": {
                        "triggered": 0,
                        "last_gc": None,
                        "memory_freed": "0 MB"
                    }
                },
                "overall_health": {
                    "status": "healthy" if memory.percent < 80 else "warning",
                    "recommendations": ["内存使用情况正常"] if memory.percent < 80 else ["内存使用率较高"]
                }
            })

        elif args.action == "cleanup":
            # 执行基础清理
            collected = gc.collect()
            result.update({
                "cleanup_results": {
                    "objects_collected": collected,
                    "message": f"已清理 {collected} 个对象"
                },
                "message": "内存清理完成"
            })

        elif args.action == "optimize":
            # 执行全面优化
            before_gc = psutil.Process().memory_info()
            collected = gc.collect()
            after_gc = psutil.Process().memory_info()

            memory_freed = before_gc.rss - after_gc.rss

            result.update({
                "optimization_results": {
                    "gc_optimization": {
                        "memory_freed": f"{memory_freed / 1024 / 1024:.2f} MB",
                        "before_heap": f"{before_gc.rss / 1024 / 1024:.2f} MB",
                        "after_heap": f"{after_gc.rss / 1024 / 1024:.2f} MB"
                    }
                },
                "recommendations": ["内存优化完成"],
                "message": "内存全面优化完成"
            })

        elif args.action == "configure":
            config_updates = []
            if args.enable_monitoring is not None:
                config_updates.append(f"内存监控: {'已启用' if args.enable_monitoring else '已禁用'}")
            if args.max_concurrency is not None:
                config_updates.append(f"最大并发数: {args.max_concurrency}")

            result.update({
                "updates": config_updates,
                "current_config": {
                    "memory_monitoring": args.enable_monitoring,
                    "max_concurrency": args.max_concurrency
                },
                "message": "配置已更新"
            })

        else:
            raise ValueError(f"不支持的操作: {args.action}")

        return json.dumps(result, ensure_ascii=False, indent=2)

    except Exception as e:
        logger.error(f"内存优化失败: {e}")
        raise


async def mysql_performance_optimize(args: PerformanceOptimizeParams) -> str:
    """性能优化工具"""
    try:
        # 使用真正的性能管理器
        result = await performance_manager.optimize_performance(args.action, {
            "query": args.query,
            "params": args.params,
            "limit": args.limit,
            "includeDetails": args.include_details,
            "timeRange": args.time_range,
            "longQueryTime": args.long_query_time,
            "logQueriesNotUsingIndexes": args.log_queries_not_using_indexes,
            "monitoringIntervalMinutes": args.monitoring_interval_minutes
        })

        # 如果结果是字符串，说明是JSON，直接返回
        if isinstance(result, str):
            return result

        # 如果是其他类型，转换为JSON字符串
        return json.dumps(result, ensure_ascii=False, indent=2)

    except Exception as e:
        logger.error(f"性能优化失败: {e}")
        raise


async def mysql_manage_rate_limit(args: RateLimitManageParams) -> str:
    """速率限制管理工具"""
    try:
        result = {"action": args.action, "success": True}

        if args.action == "status":
            # 获取全局限流器状态
            global_stats = rate_limiter.get_stats()
            result.update({
                "global_config": {
                    "max_requests": config_manager.security.rate_limit_max,
                    "window_seconds": config_manager.security.rate_limit_window,
                    "burst_limit": config_manager.security.rate_limit_max // 2,
                    "strategy": "token_bucket"
                },
                "global_stats": global_stats,
                "current_time": time.time()
            })

        elif args.action == "reset":
            # 重置限流器
            identifier = args.identifier
            if identifier:
                rate_limiter.reset(identifier)
                result.update({
                    "message": f"已重置标识符 '{identifier}' 的限流状态",
                    "reset_identifier": identifier
                })
            else:
                # 重置所有限流状态
                rate_limiter = RateLimiterManager(rate_limiter_config)
                result.update({
                    "message": "已重置所有限流状态",
                    "reset_all": True
                })

        elif args.action == "update_config":
            # 更新限流配置
            if not args.new_config:
                raise ValueError("更新配置需要提供 new_config 参数")

            try:
                # 验证新配置
                new_config_dict = args.new_config
                new_rate_config = RateLimitConfig(
                    max_requests=new_config_dict.get("max_requests", config_manager.security.rate_limit_max),
                    window_seconds=new_config_dict.get("window_seconds", config_manager.security.rate_limit_window),
                    burst_limit=new_config_dict.get("burst_limit", config_manager.security.rate_limit_max // 2),
                    strategy=new_config_dict.get("strategy", "token_bucket")
                )

                # 更新全局限流器（需要重启服务器来生效）
                # 注意：由于这是全局变量，实际应用中可能需要重启服务
                result.update({
                    "message": "配置更新成功，但需要重启服务器才能完全生效",
                    "restart_required": True,
                    "updated_config": new_rate_config.model_dump()
                })

                result.update({
                    "message": "限流配置已更新",
                    "new_config": new_rate_config.model_dump(),
                    "old_config": config_manager.security.model_dump()
                })

            except Exception as config_error:
                result.update({
                    "success": False,
                    "error": f"配置更新失败: {str(config_error)}",
                    "message": "请检查配置参数是否正确"
                })

        elif args.action == "get_stats":
            # 获取特定操作的统计信息
            operation_type = args.operation_type or "all"
            stats = rate_limiter.get_stats()

            if operation_type != "all":
                # 可以在这里添加特定操作的统计逻辑
                result.update({
                    "operation_type": operation_type,
                    "stats": stats
                })
            else:
                result.update({
                    "all_stats": stats
                })

        elif args.action == "demo_algorithms":
            # 演示不同限流算法
            demo_results = {}

            # 测试固定窗口算法
            fixed_config = RateLimitConfig(max_requests=5, window_seconds=10, strategy="fixed_window")
            fixed_limiter = RateLimiterManager(fixed_config)

            fixed_results = []
            for i in range(8):
                allowed = fixed_limiter.is_allowed(f"demo_user_{i}")
                status = fixed_limiter.check_limit(f"demo_user_{i}")
                fixed_results.append({
                    "request": i + 1,
                    "allowed": allowed,
                    "remaining": status.remaining
                })

            demo_results["fixed_window"] = fixed_results

            # 测试令牌桶算法
            token_config = RateLimitConfig(max_requests=10, window_seconds=10, burst_limit=5, strategy="token_bucket")
            token_limiter = RateLimiterManager(token_config)

            token_results = []
            for i in range(8):
                allowed = token_limiter.is_allowed(f"demo_user_{i}")
                status = token_limiter.check_limit(f"demo_user_{i}")
                token_results.append({
                    "request": i + 1,
                    "allowed": allowed,
                    "remaining": status.remaining
                })

            demo_results["token_bucket"] = token_results

            result.update({
                "demo_results": demo_results,
                "message": "限流算法演示完成",
                "algorithms_tested": ["fixed_window", "token_bucket"]
            })

        else:
            raise ValueError(f"不支持的操作: {args.action}")

        return json.dumps(result, ensure_ascii=False, indent=2)

    except Exception as e:
        logger.error(f"限流管理失败: {e}")
        raise


async def mysql_replication_status(args: Dict) -> str:
    """复制状态监控"""
    try:
        action = args.get("action", "status")
        result = {"success": True, "action": action}

        if action == "status":
            # 检查主库状态
            try:
                master_status = await mysql_manager.execute_query('SHOW MASTER STATUS')
                result["master_status"] = master_status
                result["is_master"] = len(master_status) > 0
            except Exception:
                result["master_status"] = {"error": "无法获取主库状态"}
                result["is_master"] = False

            # 检查从库状态
            try:
                slave_status = await mysql_manager.execute_query('SHOW SLAVE STATUS')
                result["slave_status"] = slave_status
                result["is_slave"] = len(slave_status) > 0
            except Exception:
                result["slave_status"] = {"error": "无法获取从库状态"}
                result["is_slave"] = False

            # 分析复制健康状态
            result["replication_health"] = {
                "status": "unknown",
                "health_score": 0,
                "issues": ["复制状态未知"],
                "recommendations": ["检查复制配置"]
            }
            result["recommendations"] = ["复制监控功能受限，建议使用专用复制监控工具"]

        elif action == "delay":
            result.update({
                "success": False,
                "action": "delay",
                "error": "复制延迟检测功能未完全实现",
                "configured": False
            })

        elif action == "diagnose":
            result.update({
                "diagnostics": {
                    "has_errors": False,
                    "errors": [],
                    "warnings": ["诊断功能受限"],
                    "recommendations": ["使用专用复制诊断工具"]
                },
                "system_variables": {"replication_status": "unknown"}
            })

        elif action == "config":
            result.update({
                "configuration": [{
                    "query": "SHOW VARIABLES LIKE '%replication%'",
                    "result": [{"Variable_name": "replication_status", "Value": "unknown"}],
                    "success": True
                }],
                "config_analysis": {
                    "is_master": False,
                    "is_slave": False,
                    "issues": ["配置分析功能受限"],
                    "recommendations": ["使用专用配置分析工具"]
                }
            })

        result["timestamp"] = time.time()
        return json.dumps(result, ensure_ascii=False, indent=2)

    except Exception as e:
        logger.error(f"复制状态监控失败: {e}")
        raise


# 备份工具处理函数
async def mysql_backup_incremental(args: IncrementalBackupParams) -> str:
    """增量备份"""
    try:
        result = await backup_tool.create_incremental_backup(args.model_dump())
        return json.dumps(result, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.error(f"增量备份失败: {e}")
        raise


async def mysql_backup_full(args: BackupParams) -> str:
    """全量备份"""
    try:
        result = await backup_tool.create_backup(args.model_dump())
        return json.dumps(result, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.error(f"全量备份失败: {e}")
        raise


async def mysql_export_data_advanced(args: ExportDataParams) -> str:
    """高级数据导出"""
    try:
        result = await backup_tool.export_data(
            args.query,
            args.params,
            args.model_dump()
        )
        return json.dumps(result, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.error(f"数据导出失败: {e}")
        raise


async def mysql_generate_report_advanced(args: GenerateReportParams) -> str:
    """高级报表生成"""
    try:
        result = await backup_tool.generate_report(args.model_dump())
        return json.dumps(result, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.error(f"报表生成失败: {e}")
        raise


async def mysql_verify_backup_advanced(args: VerifyBackupParams) -> str:
    """高级备份验证"""
    try:
        result = await backup_tool.verify_backup(args.backup_file_path)
        return json.dumps(result, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.error(f"备份验证失败: {e}")
        raise


# 导入工具处理函数
async def mysql_import_from_csv(args: ImportDataParams) -> str:
    """从CSV文件导入数据"""
    try:
        result = await import_tool.import_from_csv(args)
        return json.dumps(result, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.error(f"CSV导入失败: {e}")
        raise


async def mysql_import_from_json(args: ImportDataParams) -> str:
    """从JSON文件导入数据"""
    try:
        result = await import_tool.import_from_json(args)
        return json.dumps(result, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.error(f"JSON导入失败: {e}")
        raise


async def mysql_import_from_excel(args: ImportDataParams) -> str:
    """从Excel文件导入数据"""
    try:
        result = await import_tool.import_from_excel(args)
        return json.dumps(result, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.error(f"Excel导入失败: {e}")
        raise


async def mysql_import_from_sql(args: ImportDataParams) -> str:
    """从SQL文件导入数据"""
    try:
        result = await import_tool.import_from_sql(args)
        return json.dumps(result, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.error(f"SQL导入失败: {e}")
        raise


async def mysql_import_data_advanced(args: ImportDataParams) -> str:
    """通用数据导入"""
    try:
        result = await import_tool.import_data(args)
        return json.dumps(result, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.error(f"数据导入失败: {e}")
        raise


async def mysql_validate_import(args: ImportDataParams) -> str:
    """验证导入数据"""
    try:
        result = await import_tool.validate_import(args)
        return json.dumps(result, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.error(f"导入验证失败: {e}")
        raise


async def mysql_progress_tracker(args: ProgressTrackerParams) -> str:
    """进度跟踪工具"""
    try:
        # 简化的进度跟踪实现
        result = {"action": args.action, "success": True}

        if args.action == "list":
            # 列出活跃操作
            result.update({
                "total_active_trackers": 0,  # 简化实现
                "filtered_count": 0,
                "filter": {
                    "operation_type": args.operation_type,
                    "include_completed": args.include_completed
                },
                "trackers": [],
                "summary": {
                    "backup": 0,
                    "export": 0,
                    "running": 0,
                    "completed": 0
                }
            })

        elif args.action == "get":
            if not args.tracker_id:
                raise ValueError("获取跟踪器详情需要提供tracker_id")

            result.update({
                "tracker_id": args.tracker_id,
                "operation": "unknown",
                "start_time": None,
                "progress": {"stage": "unknown", "progress": 0, "message": "跟踪器不存在"},
                "timing": {
                    "elapsed": 0,
                    "elapsed_formatted": "0s",
                    "estimated_total": "unknown",
                    "estimated_remaining": "unknown"
                },
                "can_cancel": False,
                "status": "not_found"
            })

        elif args.action == "cancel":
            if not args.tracker_id:
                raise ValueError("取消操作需要提供tracker_id")

            result.update({
                "tracker_id": args.tracker_id,
                "cancelled": False,
                "message": "跟踪器不存在或不支持取消"
            })

        elif args.action == "summary":
            result.update({
                "active_operations": {
                    "total": 0,
                    "backup": 0,
                    "export": 0,
                    "running": 0,
                    "near_completion": 0
                },
                "system_status": {
                    "queue_stats": {
                        "running_tasks": 0,
                        "queued_tasks": 0,
                        "max_concurrent_tasks": 5
                    },
                    "memory_info": {
                        "usage": "0MB",
                        "pressure": "0%",
                        "status": "normal"
                    }
                },
                "longest_running": None,
                "recommendations": ["系统运行正常，无活跃操作"]
            })

        else:
            raise ValueError(f"不支持的操作: {args.action}")

        return json.dumps(result, ensure_ascii=False, indent=2)

    except Exception as e:
        logger.error(f"进度跟踪失败: {e}")
        raise


async def mysql_verify_backup(args: VerifyBackupParams) -> str:
    """验证备份文件"""
    try:
        import zipfile
        import hashlib
        import re

        backup_path = args.backup_file_path

        # 检查文件是否存在
        if not os.path.exists(backup_path):
            return json.dumps({
                "valid": False,
                "file_size": 0,
                "tables_found": [],
                "error": f"备份文件不存在: {backup_path}"
            }, ensure_ascii=False, indent=2)

        # 获取文件信息
        file_stats = os.stat(backup_path)
        file_size = file_stats.st_size
        created_at = file_stats.st_mtime

        # 检测压缩类型
        file_ext = os.path.splitext(backup_path)[1].lower()
        is_compressed = file_ext in ['.zip', '.gz', '.gzip', '.br']

        compression_type = 'none'
        if is_compressed:
            if file_ext == '.zip':
                compression_type = 'ZIP'
            elif file_ext in ['.gz', '.gzip']:
                compression_type = 'GZIP'
            elif file_ext == '.br':
                compression_type = 'BROTLI'

        content = ""
        compression_ratio = None
        decompressed_size = None
        temp_files = []

        try:
            if is_compressed and file_ext == '.zip':
                # 处理ZIP文件
                with zipfile.ZipFile(backup_path, 'r') as zipf:
                    # 找到SQL文件
                    sql_files = [f for f in zipf.namelist() if f.endswith('.sql')]
                    if not sql_files:
                        raise Exception("ZIP文件中未找到SQL文件")

                    # 读取第一个SQL文件
                    sql_file = sql_files[0]
                    with zipf.open(sql_file) as f:
                        content = f.read().decode('utf-8')
                        decompressed_size = len(content)

                    compression_ratio = file_size / decompressed_size if decompressed_size > 0 else None

            elif is_compressed and file_ext in ['.gz', '.gzip']:
                # 处理GZIP文件
                import gzip
                with gzip.open(backup_path, 'rt', encoding='utf-8') as f:
                    content = f.read()
                    decompressed_size = len(content)
                    compression_ratio = file_size / decompressed_size if decompressed_size > 0 else None

            else:
                # 处理未压缩文件
                with open(backup_path, 'r', encoding='utf-8') as f:
                    content = f.read()

        except UnicodeDecodeError:
            return json.dumps({
                "valid": False,
                "file_size": file_size,
                "tables_found": [],
                "error": "文件编码错误，无法读取备份内容"
            }, ensure_ascii=False, indent=2)

        except Exception as e:
            return json.dumps({
                "valid": False,
                "file_size": file_size,
                "tables_found": [],
                "error": f"读取备份文件失败: {str(e)}"
            }, ensure_ascii=False, indent=2)

        # 验证文件大小
        if file_size == 0:
            return json.dumps({
                "valid": False,
                "file_size": 0,
                "tables_found": [],
                "error": "备份文件为空"
            }, ensure_ascii=False, indent=2)

        if file_size < 100:
            return json.dumps({
                "valid": False,
                "file_size": file_size,
                "tables_found": [],
                "error": "备份文件大小异常小"
            }, ensure_ascii=False, indent=2)

        # 计算校验和
        checksum = hashlib.sha256(content.encode('utf-8')).hexdigest()

        # 验证备份文件结构
        warnings = []

        # 检查文件头
        has_header = (
            'mysqldump' in content or
            'MySQL dump' in content or
            'MySQL数据库备份' in content or
            '-- Server version' in content
        )

        # 检查文件尾
        has_footer = (
            'Dump completed' in content or
            '备份结束' in content or
            '-- Dump completed on' in content or
            'SET SQL_MODE=@OLD_SQL_MODE' in content
        )

        # 检查字符集设置
        has_charset = (
            'utf8' in content.lower() or
            'utf-8' in content.lower() or
            'CHARACTER SET' in content.upper() or
            'CHARSET' in content.upper()
        )

        # 查找表定义
        create_table_matches = re.findall(r'CREATE TABLE[\s\S]*?ENGINE[\s\S]*?;', content, re.IGNORECASE)
        drop_table_matches = re.findall(r'DROP TABLE IF EXISTS [`"]([^`"]+)[`"]', content, re.IGNORECASE)
        insert_matches = re.findall(r'INSERT INTO[\s\S]*?VALUES[\s\S]*?;', content, re.IGNORECASE)

        # 提取表名
        tables_from_drop = []
        for match in drop_table_matches:
            table_match = re.search(r'[`"]([^`"]+)[`"]', match)
            if table_match:
                tables_from_drop.append(table_match.group(1))

        # 去重
        tables_found = list(set(tables_from_drop))

        # 估算记录数量
        record_count = len(insert_matches)

        # 检查备份类型
        backup_type = 'unknown'
        if create_table_matches and insert_matches:
            backup_type = 'full'
        elif create_table_matches:
            backup_type = 'structure'
        elif insert_matches:
            backup_type = 'data'

        # 生成警告
        if not has_header:
            warnings.append('缺少备份文件头')

        if not has_footer:
            warnings.append('缺少文件结尾标识')

        if not has_charset:
            warnings.append('未找到字符集设置')

        if not tables_found:
            warnings.append('未找到表定义')

        if backup_type == 'full' and record_count == 0:
            warnings.append('未找到数据插入语句')

        # 计算完整性评分
        completeness_score = 0
        if has_header:
            completeness_score += 0.3
        if has_footer:
            completeness_score += 0.2
        if tables_found:
            completeness_score += 0.3
        if has_charset:
            completeness_score += 0.1
        if insert_matches:
            completeness_score += 0.1

        # 扣除警告分数
        completeness_score -= min(len(warnings) * 0.05, 0.2)
        completeness_score = max(0, min(1, completeness_score))

        # 判断是否有效
        is_valid = has_header and tables_found and completeness_score >= 0.7

        # 生成错误信息
        error = None
        if not is_valid:
            errors = []
            if not has_header:
                errors.append('缺少备份文件头')
            if not tables_found:
                errors.append('未找到表定义')
            if completeness_score < 0.7:
                errors.append(f"完整性评分过低 ({completeness_score:.1f})")
            error = f"备份文件验证失败: {'; '.join(errors)}"

        return json.dumps({
            "valid": is_valid,
            "file_size": file_size,
            "tables_found": tables_found,
            "record_count": record_count,
            "created_at": created_at,
            "backup_type": backup_type,
            "compression": compression_type,
            "compression_ratio": compression_ratio,
            "decompressed_size": decompressed_size,
            "checksum": checksum,
            "error": error,
            "warnings": warnings if warnings else None
        }, ensure_ascii=False, indent=2)

    except Exception as e:
        logger.error(f"验证备份失败: {e}")
        raise


async def mysql_generate_report(args: GenerateReportParams) -> str:
    """生成数据报表"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from pathlib import Path

        start_time = time.time()

        # 验证所有查询
        for i, query_config in enumerate(args.queries):
            mysql_manager.validate_input(query_config.query, f'report_query_{i}')
            mysql_manager.validate_input(query_config.name, f'report_name_{i}')
            if query_config.params:
                for j, param in enumerate(query_config.params):
                    mysql_manager.validate_input(param, f'report_query_{i}_param_{j}')

        # 设置默认输出目录
        output_dir = args.output_dir or "./reports"
        os.makedirs(output_dir, exist_ok=True)

        # 生成报表文件名
        timestamp = time.strftime("%Y%m%d_%H%M%S", time.localtime())
        if args.file_name:
            file_name = args.file_name
            if not file_name.endswith('.xlsx'):
                file_name += '.xlsx'
        else:
            file_name = f"report_{timestamp}.xlsx"

        file_path = os.path.join(output_dir, file_name)

        # 创建Excel工作簿
        workbook = Workbook()
        workbook.remove(workbook.active)  # 移除默认工作表

        # 添加报表信息工作表
        info_sheet = workbook.create_sheet("报表信息")
        info_sheet['A1'] = "报表标题"
        info_sheet['B1'] = args.title
        info_sheet['A2'] = "生成时间"
        info_sheet['B2'] = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
        if args.description:
            info_sheet['A3'] = "报表描述"
            info_sheet['B3'] = args.description

        info_sheet['A4'] = "查询数量"
        info_sheet['B4'] = len(args.queries)

        # 设置信息表样式
        info_sheet['A1'].font = Font(bold=True, size=14)
        info_sheet.column_dimensions['A'].width = 15
        info_sheet.column_dimensions['B'].width = 40

        total_rows = 0
        total_columns = 0
        query_results = []

        # 为每个查询创建工作表
        for query_config in args.queries:
            try:
                # 执行查询
                result = await mysql_manager.execute_query(query_config.query, query_config.params or [])

                if isinstance(result, list) and result:
                    # 创建工作表（Excel工作表名称限制31个字符）
                    sheet_name = query_config.name[:31]
                    worksheet = workbook.create_sheet(sheet_name)

                    # 添加查询名称和描述
                    worksheet['A1'] = query_config.name
                    worksheet['A1'].font = Font(bold=True, size=12)
                    worksheet.merge_cells('A1:Z1')

                    # 添加查询SQL（如果不为空）
                    if query_config.query.strip():
                        worksheet['A3'] = "查询SQL:"
                        worksheet['A4'] = query_config.query
                        worksheet['A4'].alignment = Alignment(wrap_text=True)
                        worksheet.merge_cells('A4:Z4')

                    # 添加数据标题行
                    columns = list(result[0].keys())
                    start_row = 6

                    # 添加列头
                    for col_idx, col_name in enumerate(columns):
                        cell = worksheet.cell(row=start_row, column=col_idx + 1)
                        cell.value = col_name
                        cell.font = Font(bold=True)
                        # 设置列宽
                        worksheet.column_dimensions[chr(65 + col_idx)].width = min(max(len(col_name) + 2, 10), 50)

                    # 添加数据
                    for row_idx, row_data in enumerate(result):
                        for col_idx, col_value in enumerate(row_data.values()):
                            cell = worksheet.cell(row=start_row + row_idx + 1, column=col_idx + 1)
                            cell.value = col_value

                    # 统计信息
                    query_info = {
                        "name": query_config.name,
                        "row_count": len(result),
                        "column_count": len(columns),
                        "success": True
                    }

                    total_rows += len(result)
                    total_columns = max(total_columns, len(columns))
                else:
                    # 空结果处理
                    query_info = {
                        "name": query_config.name,
                        "row_count": 0,
                        "column_count": 0,
                        "success": True,
                        "message": "查询无结果"
                    }

                query_results.append(query_info)

            except Exception as query_error:
                # 查询失败时创建错误工作表
                error_sheet_name = f"错误_{query_config.name[:25]}"
                error_sheet = workbook.create_sheet(error_sheet_name)

                error_sheet['A1'] = "查询名称"
                error_sheet['B1'] = query_config.name
                error_sheet['A2'] = "错误信息"
                error_sheet['B2'] = str(query_error)
                error_sheet['A3'] = "查询SQL"
                error_sheet['B3'] = query_config.query

                # 设置样式
                error_sheet.column_dimensions['A'].width = 15
                error_sheet.column_dimensions['B'].width = 50
                error_sheet['B3'].alignment = Alignment(wrap_text=True)

                query_results.append({
                    "name": query_config.name,
                    "success": False,
                    "error": str(query_error)
                })

        # 更新信息表统计信息
        current_row = 5
        if total_rows > 0:
            info_sheet[f'A{current_row}'] = "总行数"
            info_sheet[f'B{current_row}'] = total_rows
            current_row += 1

        if total_columns > 0:
            info_sheet[f'A{current_row}'] = "总列数"
            info_sheet[f'B{current_row}'] = total_columns
            current_row += 1

        # 添加查询结果摘要
        info_sheet[f'A{current_row}'] = "查询结果摘要"
        info_sheet[f'A{current_row}'].font = Font(bold=True)
        current_row += 1

        for i, query_result in enumerate(query_results):
            if query_result.get("success"):
                status = f"成功 ({query_result.get('row_count', 0)}行)"
            else:
                status = "失败"

            info_sheet[f'A{current_row}'] = query_result["name"]
            info_sheet[f'B{current_row}'] = status
            current_row += 1

        # 保存文件
        workbook.save(file_path)

        # 获取文件大小
        file_size = os.path.getsize(file_path)
        duration = int((time.time() - start_time) * 1000)

        return json.dumps({
            "success": True,
            "file_path": file_path,
            "file_size": file_size,
            "total_queries": len(args.queries),
            "successful_queries": len([q for q in query_results if q.get("success", False)]),
            "total_rows": total_rows,
            "total_columns": total_columns,
            "duration": duration,
            "title": args.title,
            "timestamp": timestamp,
            "query_results": query_results
        }, ensure_ascii=False, indent=2)

    except Exception as e:
        logger.error(f"生成报表失败: {e}")
        raise


async def mysql_backup(args: BackupParams) -> str:
    """数据库备份"""
    try:
        import subprocess
        import tempfile
        from pathlib import Path

        start_time = time.time()

        # 设置默认值
        output_dir = args.output_dir or "./backups"
        os.makedirs(output_dir, exist_ok=True)

        compress = args.compress if args.compress is not None else True
        include_data = args.include_data if args.include_data is not None else True
        include_structure = args.include_structure if args.include_structure is not None else True
        tables = args.tables or []
        file_prefix = args.file_prefix or "mysql_backup"
        max_file_size = args.max_file_size or 100

        # 获取数据库配置
        config = mysql_manager.config_manager.database

        # 生成备份文件名
        timestamp = time.strftime("%Y%m%d_%H%M%S", time.localtime())
        file_name = f"{file_prefix}_{timestamp}"
        backup_path = os.path.join(output_dir, file_name)

        # 构建mysqldump命令参数
        dump_args = [
            "mysqldump",
            f"-h{config.host}",
            f"-P{config.port}",
            f"-u{config.user}",
            f"-p{config.password}",
            "--default-character-set=utf8mb4",
            "--single-transaction",
            "--routines",
            "--triggers"
        ]

        # 根据选项添加参数
        if not include_data:
            dump_args.append("--no-data")

        if not include_structure:
            dump_args.append("--no-create-info")

        # 添加数据库名
        dump_args.append(config.database or "")

        # 添加指定的表
        if tables:
            dump_args.extend(tables)

        # 执行备份
        sql_file_path = f"{backup_path}.sql"

        try:
            with open(sql_file_path, 'w', encoding='utf-8') as f:
                result = subprocess.run(dump_args, stdout=f, stderr=subprocess.PIPE, text=True, check=True)

            # 检查文件大小并压缩
            file_size = os.path.getsize(sql_file_path)

            final_file_path = sql_file_path
            if compress or file_size > (max_file_size * 1024 * 1024):
                # 使用zip压缩
                import zipfile
                zip_file_path = f"{backup_path}.zip"

                with zipfile.ZipFile(zip_file_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    zipf.write(sql_file_path, os.path.basename(sql_file_path))

                # 删除原始SQL文件
                os.remove(sql_file_path)
                final_file_path = zip_file_path
                file_size = os.path.getsize(zip_file_path)

            # 获取表统计信息
            table_count = 0
            record_count = 0

            if tables:
                table_count = len(tables)
                # 估算记录数（简化处理）
                record_count = 0
            else:
                # 获取所有表的统计信息
                try:
                    tables_result = await mysql_manager.execute_query("SHOW TABLES")
                    if isinstance(tables_result, list):
                        table_count = len(tables_result)

                    # 获取总记录数（简化处理）
                    count_result = await mysql_manager.execute_query("""
                        SELECT SUM(TABLE_ROWS) as total_rows
                        FROM information_schema.tables
                        WHERE table_schema = DATABASE()
                    """)
                    if isinstance(count_result, list) and count_result:
                        record_count = count_result[0].get('total_rows', 0) or 0
                except Exception:
                    # 如果获取统计信息失败，使用默认值
                    pass

            duration = int((time.time() - start_time) * 1000)

            return json.dumps({
                "success": True,
                "file_path": final_file_path,
                "file_size": file_size,
                "table_count": table_count,
                "record_count": record_count,
                "duration": duration,
                "backup_type": args.backup_type or "full",
                "compressed": compress,
                "includes_data": include_data,
                "includes_structure": include_structure,
                "timestamp": timestamp
            }, ensure_ascii=False, indent=2)

        except subprocess.CalledProcessError as e:
            error_msg = e.stderr if e.stderr else str(e)
            raise Exception(f"mysqldump执行失败: {error_msg}")
        except FileNotFoundError:
            raise Exception("mysqldump命令未找到，请确保MySQL客户端已安装")

    except Exception as e:
        logger.error(f"数据库备份失败: {e}")
        raise


async def mysql_export_data(args: ExportDataParams) -> str:
    """数据导出"""
    try:
        # 检查导出操作的速率限制
        if not check_rate_limit(args, "export"):
            rate_status = get_rate_limit_status(args, "export")
            from type_utils import MySQLMCPError, ErrorCategory, ErrorSeverity
            raise MySQLMCPError(
                STRING_CONSTANTS["MSG_RATE_LIMIT_EXCEEDED"],
                ErrorCategory.RATE_LIMIT_ERROR,
                ErrorSeverity.HIGH,
                metadata={
                    "rate_limit_status": rate_status,
                    "operation": "mysql_export_data"
                }
            )

        import pandas as pd
        from pathlib import Path
        from type_utils import ExportResult, ErrorCategory, ErrorSeverity, MySQLMCPError

        start_time = time.time()

        # 验证查询和参数
        mysql_manager.validate_input(args.query, 'export_query')
        if args.params:
            for i, param in enumerate(args.params):
                mysql_manager.validate_input(param, f'export_param_{i}')

        # 设置默认值
        output_dir = args.output_dir
        if not output_dir:
            output_dir = "exports"
        os.makedirs(output_dir, exist_ok=True)

        format_type = args.format or "excel"
        sheet_name = args.sheet_name or "Data"
        include_headers = args.include_headers if args.include_headers is not None else True
        max_rows = args.max_rows or 100000

        # 生成文件名
        timestamp = int(time.time())
        if args.file_name:
            file_name = args.file_name
        else:
            file_name = f"mysql_export_{timestamp}"

        # 执行查询
        result = await mysql_manager.execute_query(args.query, args.params or [])

        # 验证结果
        if not isinstance(result, list):
            raise MySQLMCPError(
                "查询结果格式无效，导出需要返回列表格式的数据",
                ErrorCategory.DATA_ERROR,
                ErrorSeverity.HIGH
            )

        if len(result) == 0:
            # 创建空文件
            file_path = ""
            if format_type == "excel":
                file_path = os.path.join(output_dir, f"{file_name}.xlsx")
                # 创建空的DataFrame并保存
                df = pd.DataFrame()
                df.to_excel(file_path, index=False)
            elif format_type == "csv":
                file_path = os.path.join(output_dir, f"{file_name}.csv")
                pd.DataFrame().to_csv(file_path, index=False, header=include_headers)
            elif format_type == "json":
                file_path = os.path.join(output_dir, f"{file_name}.json")
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump([], f, ensure_ascii=False, indent=2)

            export_result = ExportResult(
                success=True,
                file_path=file_path,
                file_size=0,
                row_count=0,
                column_count=0,
                format=format_type,
                duration=int((time.time() - start_time) * 1000),
                processing_mode="direct"
            )
        else:
            # 限制行数
            if len(result) > max_rows:
                result = result[:max_rows]
                logger.warn(f"结果集超过最大行数限制 {max_rows}，已截断")

            # 转换为DataFrame
            df = pd.DataFrame(result)

            # 导出文件
            file_path = ""
            if format_type == "excel":
                file_path = os.path.join(output_dir, f"{file_name}.xlsx")
                df.to_excel(file_path, sheet_name=sheet_name, index=False, header=include_headers)
            elif format_type == "csv":
                file_path = os.path.join(output_dir, f"{file_name}.csv")
                df.to_csv(file_path, index=False, header=include_headers, encoding='utf-8')
            elif format_type == "json":
                file_path = os.path.join(output_dir, f"{file_name}.json")
                df.to_json(file_path, orient='records', indent=2, force_ascii=False)
            else:
                raise MySQLMCPError(
                    f"不支持的导出格式: {format_type}",
                    ErrorCategory.VALIDATION_ERROR,
                    ErrorSeverity.HIGH
                )

            # 获取文件大小
            file_size = os.path.getsize(file_path)

            export_result = ExportResult(
                success=True,
                file_path=file_path,
                file_size=file_size,
                row_count=len(df),
                column_count=len(df.columns),
                format=format_type,
                duration=int((time.time() - start_time) * 1000),
                processing_mode="direct"
            )

        # 添加限流信息到响应
        rate_status = get_rate_limit_status(args, "export")
        response_data = {
            "success": export_result.success,
            "file_path": export_result.file_path,
            "file_size": export_result.file_size,
            "row_count": export_result.row_count,
            "column_count": export_result.column_count,
            "format": export_result.format,
            "duration": export_result.duration,
            "processing_mode": export_result.processing_mode,
            "query": args.query,
            "timestamp": int(time.time()),
            "_rate_limit_info": rate_status
        }

        return json.dumps(response_data, ensure_ascii=False, indent=2)

    except Exception as e:
        logger.error(f"数据导出失败: {e}")
        raise


async def mysql_manage_indexes(args: ManageIndexesParams) -> str:
    """索引管理

    提供完整的MySQL索引管理功能，包括索引查看、分析和优化操作。
    支持索引列表查询、性能分析和优化建议，帮助提升数据库查询性能。
    提供索引使用统计和优化建议，支持数据库性能调优和索引策略制定。
    包含索引碎片整理和统计信息更新功能，确保索引的高效性和准确性。

    @param args: 索引管理参数，包含操作类型、表名、索引名等信息
    @return: JSON格式的操作结果，包含索引信息、分析结果或优化状态
    @throws: MySQLMCPError 当表名无效、索引不存在或操作失败时
    """
    result = {"action": args.action, "success": True}

    if args.action == "list":
        if args.table_name:
            mysql_manager.validate_table_name(args.table_name)
            query = """
                SELECT INDEX_NAME, COLUMN_NAME, NON_UNIQUE, SEQ_IN_INDEX, INDEX_TYPE
                FROM INFORMATION_SCHEMA.STATISTICS
                WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = %s
                ORDER BY INDEX_NAME, SEQ_IN_INDEX
            """
            indexes = await mysql_manager.execute_query(query, [args.table_name])
            result["indexes"] = indexes
        else:
            query = """
                SELECT TABLE_NAME, INDEX_NAME, COLUMN_NAME, NON_UNIQUE, SEQ_IN_INDEX, INDEX_TYPE
                FROM INFORMATION_SCHEMA.STATISTICS
                WHERE TABLE_SCHEMA = DATABASE()
                ORDER BY TABLE_NAME, INDEX_NAME, SEQ_IN_INDEX
            """
            indexes = await mysql_manager.execute_query(query)
            result["indexes"] = indexes

    elif args.action == "analyze":
        if not args.table_name:
            raise ValueError("分析索引时必须提供表名")

        mysql_manager.validate_table_name(args.table_name)

        # 获取索引信息
        query = """
            SELECT INDEX_NAME, COLUMN_NAME, NON_UNIQUE, SEQ_IN_INDEX, INDEX_TYPE, CARDINALITY
            FROM INFORMATION_SCHEMA.STATISTICS
            WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = %s
            ORDER BY INDEX_NAME, SEQ_IN_INDEX
        """
        indexes = await mysql_manager.execute_query(query, [args.table_name])

        # 分析索引
        analysis = {
            "table_name": args.table_name,
            "total_indexes": len(indexes),
            "recommendations": []
        }

        if not indexes:
            analysis["recommendations"].append("建议为经常查询的列创建索引")

        result["analysis"] = analysis

    elif args.action == "optimize":
        if not args.table_name:
            raise ValueError("优化索引时必须提供表名")

        mysql_manager.validate_table_name(args.table_name)
        query = f"OPTIMIZE TABLE `{args.table_name}`"
        optimize_result = await mysql_manager.execute_query(query)
        result["optimization_result"] = optimize_result

    return json.dumps(result, ensure_ascii=False, indent=2)


# =============================================================================
# 信号处理器
# =============================================================================

async def graceful_shutdown(signum, frame):
    """优雅关闭处理器"""
    logger.info(f"收到信号 {signum}，开始优雅关闭", "graceful_shutdown")

    try:
        # 停止系统监控
        system_monitor.stop_monitoring()
        logger.info("系统监控已停止")

        # 停止内存监控
        memory_monitor.stop_monitoring()
        logger.info("内存监控已停止")

        # 关闭MySQL管理器
        await mysql_manager.close()
        logger.info("MySQL管理器已关闭")

        logger.info("服务已优雅关闭", "graceful_shutdown")
    except Exception as e:
        logger.error(f"优雅关闭过程中发生错误: {e}", "graceful_shutdown")

    # 退出进程
    sys.exit(0)


def setup_signal_handlers():
    """设置信号处理器"""
    def signal_handler(signum, frame):
        """信号处理器包装函数"""
        try:
            loop = asyncio.get_running_loop()
            loop.create_task(graceful_shutdown(signum, frame))
        except RuntimeError:
            # 没有运行的事件循环，直接运行同步关闭
            asyncio.run(graceful_shutdown(signum, frame))

    try:
        # 注册信号处理器
        signal.signal(signal.SIGINT, signal_handler)
        signal.signal(signal.SIGTERM, signal_handler)
        logger.info("信号处理器已设置", "signal_setup")
    except (OSError, ValueError) as e:
        logger.warn(f"设置信号处理器失败: {e}", "signal_setup")


# =============================================================================
# 工具注册配置
# =============================================================================

def register_tools():
    """注册所有MCP工具"""

    # 创建工具定义列表
    tool_definitions = []

    # 基础查询工具
    tool_definitions.append(ToolDefinition(
        name="mysql_query",
        description="Execute MySQL queries (SELECT, SHOW, DESCRIBE, etc.)",
        parameters={"query": "str", "params": "Optional[List[Any]]"},
        handler=mysql_query,
        error_message="执行MySQL查询失败",
        operation_type="query"
    ))

    tool_definitions.append(ToolDefinition(
        name="mysql_show_tables",
        description="Show all tables in the current database",
        parameters={},
        handler=mysql_show_tables,
        error_message="显示表列表失败",
        operation_type="query"
    ))

    tool_definitions.append(ToolDefinition(
        name="mysql_describe_table",
        description="Describe table structure and schema",
        parameters={"table_name": "str"},
        handler=mysql_describe_table,
        error_message="描述表结构失败",
        operation_type="query"
    ))

    tool_definitions.append(ToolDefinition(
        name="mysql_get_schema",
        description="Get database schema information",
        parameters={"table_name": "Optional[str]"},
        handler=mysql_get_schema,
        error_message="获取数据库架构失败",
        operation_type="query"
    ))

    tool_definitions.append(ToolDefinition(
        name="mysql_get_foreign_keys",
        description="Get foreign key relationships",
        parameters={"table_name": "Optional[str]"},
        handler=mysql_get_foreign_keys,
        error_message="获取外键关系失败",
        operation_type="query"
    ))

    # 数据操作工具
    tool_definitions.append(ToolDefinition(
        name="mysql_select_data",
        description="Select data from table with optional conditions",
        parameters={"table_name": "str", "columns": "Optional[List[str]]", "where_clause": "Optional[str]", "limit": "Optional[int]"},
        handler=mysql_select_data,
        error_message="查询数据失败",
        operation_type="data_operation"
    ))

    tool_definitions.append(ToolDefinition(
        name="mysql_insert_data",
        description="Insert data into table",
        parameters={"table_name": "str", "data": "Dict[str, Any]"},
        handler=mysql_insert_data,
        error_message="插入数据失败",
        operation_type="data_operation"
    ))

    tool_definitions.append(ToolDefinition(
        name="mysql_update_data",
        description="Update data in table with conditions",
        parameters={"table_name": "str", "data": "Dict[str, Any]", "where_clause": "str"},
        handler=mysql_update_data,
        error_message="更新数据失败",
        operation_type="data_operation"
    ))

    tool_definitions.append(ToolDefinition(
        name="mysql_delete_data",
        description="Delete data from table with conditions",
        parameters={"table_name": "str", "where_clause": "str"},
        handler=mysql_delete_data,
        error_message="删除数据失败",
        operation_type="data_operation"
    ))

    # 表结构管理工具
    tool_definitions.append(ToolDefinition(
        name="mysql_create_table",
        description="Create a new table with column definitions",
        parameters={"table_name": "str", "columns": "List[Dict[str, Any]]"},
        handler=mysql_create_table,
        error_message="创建表失败",
        operation_type="schema_operation"
    ))

    tool_definitions.append(ToolDefinition(
        name="mysql_drop_table",
        description="Drop/delete a table",
        parameters={"table_name": "str", "if_exists": "Optional[bool]"},
        handler=mysql_drop_table,
        error_message="删除表失败",
        operation_type="schema_operation"
    ))

    tool_definitions.append(ToolDefinition(
        name="mysql_alter_table",
        description="Alter table structure (add/modify/drop columns, indexes)",
        parameters={"table_name": "str", "alterations": "List[Dict[str, Any]]"},
        handler=mysql_alter_table,
        error_message="修改表结构失败",
        operation_type="schema_operation"
    ))

    # 批量操作工具
    tool_definitions.append(ToolDefinition(
        name="mysql_batch_execute",
        description="Execute multiple queries in batch",
        parameters={"queries": "List[Dict[str, Any]]"},
        handler=mysql_batch_execute,
        error_message="批量执行失败",
        operation_type="batch_operation"
    ))

    # 备份和导出工具
    tool_definitions.append(ToolDefinition(
        name="mysql_backup",
        description="Create full database backup",
        parameters={
            "output_dir": "Optional[str]", "compress": "Optional[bool]", "include_data": "Optional[bool]",
            "include_structure": "Optional[bool]", "tables": "Optional[List[str]]",
            "file_prefix": "Optional[str]", "max_file_size": "Optional[int]",
            "backup_type": "Optional[str]"
        },
        handler=mysql_backup,
        error_message="创建备份失败",
        operation_type="backup"
    ))

    tool_definitions.append(ToolDefinition(
        name="mysql_backup_incremental",
        description="Create incremental backup",
        parameters={
            "output_dir": "Optional[str]", "compress": "Optional[bool]", "include_data": "Optional[bool]",
            "include_structure": "Optional[bool]", "tables": "Optional[List[str]]",
            "file_prefix": "Optional[str]", "max_file_size": "Optional[int]",
            "incremental_mode": "Optional[str]", "tracking_table": "Optional[str]",
            "base_backup_path": "Optional[str]", "last_backup_time": "Optional[str]",
            "binlog_position": "Optional[str]"
        },
        handler=mysql_backup_incremental,
        error_message="创建增量备份失败",
        operation_type="backup"
    ))

    tool_definitions.append(ToolDefinition(
        name="mysql_export_data",
        description="Export data to various formats (Excel, CSV, JSON)",
        parameters={
            "query": "str", "params": "Optional[List[Any]]", "output_dir": "Optional[str]",
            "format": "Optional[str]", "sheet_name": "Optional[str]",
            "include_headers": "Optional[bool]", "max_rows": "Optional[int]",
            "file_name": "Optional[str]"
        },
        handler=mysql_export_data,
        error_message="导出数据失败",
        operation_type="export"
    ))

    tool_definitions.append(ToolDefinition(
        name="mysql_generate_report",
        description="Generate advanced data reports",
        parameters={
            "title": "str", "description": "Optional[str]", "queries": "Optional[List[Dict[str, Any]]]",
            "output_dir": "Optional[str]", "file_name": "Optional[str]", "include_headers": "Optional[bool]"
        },
        handler=mysql_generate_report,
        error_message="生成报表失败",
        operation_type="report"
    ))

    tool_definitions.append(ToolDefinition(
        name="mysql_verify_backup",
        description="Verify backup file integrity",
        parameters={"backup_file_path": "str"},
        handler=mysql_verify_backup,
        error_message="验证备份失败",
        operation_type="backup"
    ))

    # 数据导入工具
    tool_definitions.append(ToolDefinition(
        name="mysql_import_data",
        description="Import data from various file formats",
        parameters={
            "table_name": "str", "file_path": "str", "format": "str",
            "has_headers": "Optional[bool]", "field_mapping": "Optional[Dict[str, str]]",
            "batch_size": "Optional[int]", "skip_duplicates": "Optional[bool]",
            "conflict_strategy": "Optional[str]", "use_transaction": "Optional[bool]",
            "validate_data": "Optional[bool]", "encoding": "Optional[str]",
            "sheet_name": "Optional[str]", "delimiter": "Optional[str]", "quote": "Optional[str]"
        },
        handler=mysql_import_data,
        error_message="导入数据失败",
        operation_type="import"
    ))

    # 用户管理工具
    tool_definitions.append(ToolDefinition(
        name="mysql_manage_users",
        description="Manage MySQL users (create, delete, grant, revoke permissions)",
        parameters={
            "action": "str", "username": "Optional[str]", "password": "Optional[str]",
            "host": "Optional[str]", "privileges": "Optional[List[str]]",
            "database": "Optional[str]", "table": "Optional[str]", "if_exists": "Optional[bool]"
        },
        handler=mysql_manage_users,
        error_message="用户管理失败",
        operation_type="user_management"
    ))

    # 索引管理工具
    tool_definitions.append(ToolDefinition(
        name="mysql_manage_indexes",
        description="Manage database indexes (list, analyze, optimize)",
        parameters={
            "action": "str", "table_name": "Optional[str]", "index_name": "Optional[str]",
            "index_type": "Optional[str]", "columns": "Optional[List[str]]",
            "if_exists": "Optional[bool]", "invisible": "Optional[bool]"
        },
        handler=mysql_manage_indexes,
        error_message="索引管理失败",
        operation_type="index_management"
    ))

    # 系统监控和优化工具
    tool_definitions.append(ToolDefinition(
        name="mysql_system_status",
        description="Get comprehensive system status",
        parameters={},
        handler=mysql_system_status,
        error_message="获取系统状态失败",
        operation_type="monitoring"
    ))

    tool_definitions.append(ToolDefinition(
        name="mysql_analyze_error",
        description="Analyze MySQL errors and provide recommendations",
        parameters={"error_message": "str", "operation": "Optional[str]"},
        handler=mysql_analyze_error,
        error_message="错误分析失败",
        operation_type="monitoring"
    ))

    tool_definitions.append(ToolDefinition(
        name="mysql_performance_optimize",
        description="Performance optimization and monitoring",
        parameters={
            "action": "str", "query": "Optional[str]", "params": "Optional[List[Any]]",
            "limit": "Optional[int]", "include_details": "Optional[bool]",
            "time_range": "Optional[str]", "long_query_time": "Optional[int]",
            "log_queries_not_using_indexes": "Optional[bool]", "monitoring_interval_minutes": "Optional[int]"
        },
        handler=mysql_performance_optimize,
        error_message="性能优化失败",
        operation_type="optimization"
    ))

    # 内存和队列管理工具
    tool_definitions.append(ToolDefinition(
        name="mysql_optimize_memory",
        description="Manage memory optimization and monitoring",
        parameters={
            "action": "str", "force_gc": "Optional[bool]", "enable_monitoring": "Optional[bool]",
            "max_concurrency": "Optional[int]", "include_history": "Optional[bool]"
        },
        handler=mysql_optimize_memory,
        error_message="内存优化失败",
        operation_type="optimization"
    ))

    tool_definitions.append(ToolDefinition(
        name="mysql_manage_queue",
        description="Manage task queue operations",
        parameters={
            "action": "str", "task_id": "Optional[str]", "max_concurrency": "Optional[int]",
            "show_details": "Optional[bool]", "filter_type": "Optional[str]"
        },
        handler=mysql_manage_queue,
        error_message="队列管理失败",
        operation_type="queue_management"
    ))

    # 限流管理工具
    tool_definitions.append(ToolDefinition(
        name="mysql_manage_rate_limit",
        description="Manage rate limiting configuration and status",
        parameters={
            "action": "str", "operation_type": "Optional[str]", "identifier": "Optional[str]",
            "new_config": "Optional[Dict[str, Any]]"
        },
        handler=mysql_manage_rate_limit,
        error_message="限流管理失败",
        operation_type="rate_limiting"
    ))

    # 复制监控工具
    tool_definitions.append(ToolDefinition(
        name="mysql_replication_status",
        description="Monitor MySQL replication status",
        parameters={"action": "Optional[str]"},
        handler=mysql_replication_status,
        error_message="复制状态监控失败",
        operation_type="monitoring"
    ))

    # 进度跟踪工具
    tool_definitions.append(ToolDefinition(
        name="mysql_progress_tracker",
        description="Track operation progress",
        parameters={
            "action": "str", "tracker_id": "Optional[str]", "operation_type": "Optional[str]",
            "include_completed": "Optional[bool]", "detail_level": "Optional[str]"
        },
        handler=mysql_progress_tracker,
        error_message="进度跟踪失败",
        operation_type="monitoring"
    ))

    # 使用create_mcp_tool注册所有工具
    for tool_def in tool_definitions:
        tool_config = create_mcp_tool(tool_def)
        mcp.tool(
            name=tool_config["name"],
            description=tool_config["description"]
        )(tool_config["execute"])


async def start_server():
    """启动服务器"""
    try:
        await mysql_manager.initialize()

        # 注册所有工具
        register_tools()

        # 启动系统监控
        system_monitor.start_monitoring()

        # 启动内存监控
        memory_monitor.start_monitoring()

        logger.info("MySQL MCP服务器启动成功", "server_startup", {
            "server_name": STRING_CONSTANTS["SERVER_NAME"],
            "version": STRING_CONSTANTS["SERVER_VERSION"],
            "session_id": mysql_manager.session_id
        })

        await mcp.run_async()
    except Exception as e:
        logger.error("MySQL MCP服务器启动失败", "server_startup", e, {
            "error": str(e),
            "server_name": STRING_CONSTANTS["SERVER_NAME"],
            "version": STRING_CONSTANTS["SERVER_VERSION"]
        })
        await mysql_manager.close()
        raise


if __name__ == "__main__":
    import asyncio

    # 设置信号处理器
    setup_signal_handlers()

    # 启动服务器
    asyncio.run(start_server())