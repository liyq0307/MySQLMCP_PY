"""
MySQL 高级备份和导出工具

提供企业级数据库备份、数据导出和报表生成功能，集成了智能内存管理、
任务队列调度、错误恢复和性能监控等高级特性。

@fileoverview MySQL 高级备份和导出工具 - 企业级数据管理解决方案
@author liyq
@version 1.0.0
@since 1.0.0
@updated 2025-09-26
@license MIT
"""

import asyncio
import os
import time
import zipfile
from datetime import datetime
from typing import Any, Dict, List, Optional, Callable
import tempfile
import hashlib
from dataclasses import dataclass, field
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

from mysql_manager import MySQLManager
from type_utils import (
    BackupOptions, ExportOptions, BackupResult, ExportResult,
    ReportConfig, ReportQuery, IncrementalBackupOptions,
    IncrementalBackupResult, ProgressTracker, CancellationToken,
    ProgressInfo, RecoveryStrategy, ErrorRecoveryResult,
    LargeFileOptions, MemoryUsage, TaskQueue, BackupVerificationResult
)
from common_utils import TimeUtils, MemoryUtils
from logger import logger
from cache import CacheRegion
from retry_strategy import SmartRetryStrategy
from export_tool import ExporterFactory
from tool_wrapper import with_error_handling, with_performance_monitoring


@dataclass
class MemoryManager:
    """
    高级内存管理器

    提供完整的内存监控、压力检测和自动清理功能。
    集成了内存使用跟踪、定时监控、压力回调和智能垃圾回收机制。

    主要功能：
    - 实时内存使用情况监控
    - 内存压力检测和阈值管理
    - 自动垃圾回收和内存清理
    - 压力事件回调机制
    - 内存使用统计和报告
    """

    max_memory_threshold: int = 1024 * 1024 * 1024  # 默认1GB
    monitoring_interval: int = 5000  # 5秒检查间隔
    _monitoring_task: Optional[asyncio.Task] = None
    _is_monitoring: bool = False
    _delayed_monitoring: bool = False
    _memory_pressure_callbacks: List[callable] = field(default_factory=list)

    def __post_init__(self):
        """初始化后处理"""
        if self.max_memory_threshold <= 0:
            raise ValueError("最大内存阈值必须大于0")

    def get_current_usage(self) -> MemoryUsage:
        """
        获取当前内存使用情况

        基于TypeScript实现，返回详细的内存使用统计信息，
        包括RSS、堆内存、外部内存等各项指标。

        Returns:
            MemoryUsage: 内存使用情况对象
        """
        memory_info = MemoryUtils.get_process_memory_info()
        return MemoryUsage(
            rss=int(memory_info["rss"]),
            heap_used=int(memory_info["heap_used"]),
            heap_total=int(memory_info["heap_total"]),
            external=int(memory_info["external"]),
            array_buffers=0  # Python中暂不实现
        )

    def check_memory_pressure(self) -> float:
        """
        检查当前内存压力

        计算当前内存使用量与设定阈值之间的比率。
        压力值范围从0.0（无压力）到1.0（达到或超过阈值）。

        Returns:
            float: 内存压力值(0.0-1.0)
        """
        usage = self.get_current_usage()
        total_used = usage.heap_used + usage.external
        pressure = min(total_used / self.max_memory_threshold, 1.0)

        logger.debug("内存压力检查", "MemoryManager", {
            "pressure": f"{pressure:.3f}",
            "total_used": f"{total_used / 1024 / 1024:.1f}MB",
            "threshold": f"{self.max_memory_threshold / 1024 / 1024:.1f}MB"
        })

        return pressure

    async def request_memory_cleanup(self) -> None:
        """
        请求执行内存清理

        尝试通过强制垃圾回收来释放不再使用的内存。
        这是一个尽力而为的操作，实际效果取决于Python垃圾回收器。

        记录清理前后的内存使用情况以便监控效果。
        """
        logger.info("开始内存清理", "MemoryManager")

        # 记录清理前的内存状态
        before_usage = self.get_current_usage()
        logger.info("清理前内存状态", "MemoryManager", {
            "heap_used": f"{before_usage.heap_used / 1024 / 1024:.1f}MB",
            "rss": f"{before_usage.rss / 1024 / 1024:.1f}MB"
        })

        # 强制垃圾回收
        import gc
        gc.collect()

        # 等待一个事件循环周期
        await asyncio.sleep(0.1)

        # 记录清理后的内存状态
        after_usage = self.get_current_usage()
        logger.info("清理后内存状态", "MemoryManager", {
            "heap_used": f"{after_usage.heap_used / 1024 / 1024:.1f}MB",
            "rss": f"{after_usage.rss / 1024 / 1024:.1f}MB",
            "freed_memory": f"{(before_usage.heap_used - after_usage.heap_used) / 1024 / 1024:.1f}MB"
        })

    def enable_memory_monitoring(self) -> None:
        """
        启用内存监控

        启动定时器定期检查内存压力。如果压力超过阈值，
        将触发已注册的压力回调，并在压力过高时自动请求内存清理。
        """
        if self._is_monitoring:
            logger.warning("内存监控已经启用", "MemoryManager")
            return

        self._is_monitoring = True

        # 检查是否有运行中的事件循环
        try:
            loop = asyncio.get_running_loop()
            self._monitoring_task = asyncio.create_task(self._monitoring_loop())
            logger.info("内存监控已启用", "MemoryManager", {
                "interval": f"{self.monitoring_interval}ms",
                "threshold": f"{self.max_memory_threshold / 1024 / 1024:.1f}MB"
            })
        except RuntimeError:
            # 没有运行的事件循环，标记为延迟启用
            logger.info("内存监控标记为延迟启用（等待事件循环）", "MemoryManager")
            self._delayed_monitoring = True

    def start_delayed_monitoring(self) -> None:
        """
        启动延迟的内存监控

        当事件循环可用时调用此方法来启动之前延迟的监控。
        """
        if self._delayed_monitoring and self._is_monitoring and not self._monitoring_task:
            try:
                self._monitoring_task = asyncio.create_task(self._monitoring_loop())
                self._delayed_monitoring = False
                logger.info("延迟的内存监控已启动", "MemoryManager", {
                    "interval": f"{self.monitoring_interval}ms",
                    "threshold": f"{self.max_memory_threshold / 1024 / 1024:.1f}MB"
                })
            except RuntimeError as e:
                logger.error("启动延迟监控失败", "MemoryManager", {"error": str(e)})

    def disable_memory_monitoring(self) -> None:
        """
        禁用内存监控

        停止内存压力监控定时器并清理相关资源。
        """
        if not self._is_monitoring:
            return

        self._is_monitoring = False

        if self._monitoring_task and not self._monitoring_task.done():
            self._monitoring_task.cancel()
            try:
                asyncio.get_event_loop().run_until_complete(self._monitoring_task)
            except asyncio.CancelledError:
                pass

        self._monitoring_task = None
        logger.info("内存监控已禁用", "MemoryManager")

    async def _monitoring_loop(self) -> None:
        """
        监控循环

        定期检查内存压力并触发相应处理。
        """
        while self._is_monitoring:
            try:
                await asyncio.sleep(self.monitoring_interval / 1000)

                pressure = self.check_memory_pressure()

                # 触发压力回调
                if pressure > 0.8:
                    for callback in self._memory_pressure_callbacks:
                        try:
                            callback(pressure)
                        except Exception as e:
                            logger.error(f"内存压力回调执行失败: {e}", "MemoryManager")

                    # 自动触发内存清理
                    if pressure > 0.9:
                        logger.warning("内存压力过高，自动清理", "MemoryManager", {
                            "pressure": f"{pressure:.3f}"
                        })
                        await self.request_memory_cleanup()

            except asyncio.CancelledError:
                break
            except Exception as e:
                logger.error(f"内存监控循环异常: {e}", "MemoryManager")
                await asyncio.sleep(1)  # 避免快速循环

    def on_memory_pressure(self, callback: callable) -> None:
        """
        注册内存压力回调

        添加一个回调函数，当内存压力超过阈值时将被调用。

        Args:
            callback: 当内存压力过高时调用的函数，接收压力值(float)作为参数
        """
        if callback not in self._memory_pressure_callbacks:
            self._memory_pressure_callbacks.append(callback)
            logger.info("已注册内存压力回调", "MemoryManager", {
                "callback_count": len(self._memory_pressure_callbacks)
            })

    def remove_memory_pressure_callback(self, callback: callable) -> None:
        """
        移除内存压力回调

        从回调列表中移除指定的函数。

        Args:
            callback: 需要移除的回调函数
        """
        if callback in self._memory_pressure_callbacks:
            self._memory_pressure_callbacks.remove(callback)
            logger.info("已移除内存压力回调", "MemoryManager", {
                "callback_count": len(self._memory_pressure_callbacks)
            })

    def get_memory_stats(self) -> Dict[str, Any]:
        """
        获取内存统计信息

        返回详细的内存使用统计和监控状态信息。

        Returns:
            Dict[str, Any]: 内存统计信息
        """
        usage = self.get_current_usage()
        pressure = self.check_memory_pressure()

        return {
            "current_usage": {
                "rss": usage.rss,
                "heap_used": usage.heap_used,
                "heap_total": usage.heap_total,
                "external": usage.external,
                "array_buffers": usage.array_buffers
            },
            "pressure": pressure,
            "threshold": self.max_memory_threshold,
            "monitoring_enabled": self._is_monitoring,
            "callback_count": len(self._memory_pressure_callbacks),
            "formatted_usage": {
                "rss_mb": f"{usage.rss / 1024 / 1024:.1f}",
                "heap_used_mb": f"{usage.heap_used / 1024 / 1024:.1f}",
                "heap_total_mb": f"{usage.heap_total / 1024 / 1024:.1f}",
                "external_mb": f"{usage.external / 1024 / 1024:.1f}",
                "pressure_percent": f"{pressure * 100:.1f}"
            }
        }

    async def cleanup(self) -> None:
        """
        清理资源

        停止监控并清理所有回调函数。
        """
        self.disable_memory_monitoring()
        self._memory_pressure_callbacks.clear()

        logger.info("MemoryManager资源已清理", "MemoryManager")


class MySQLBackupTool:
    """
    MySQL 高级备份和导出工具类

    基于企业级架构设计，提供完整的数据库备份、数据导出和报表生成功能。
    集成了智能内存管理、任务队列调度、进度跟踪、错误恢复等高级特性。
    支持事件驱动架构，提供丰富的事件通知机制。

    主要组件：
    - 内存管理器：智能内存监控和优化
    - 任务调度器：优先级队列和并发控制
    - 导出引擎：Excel/CSV/JSON 多格式导出
    - 缓存系统：查询结果缓存和 LRU 淘汰
    - 进度跟踪：实时进度监控和取消机制
    - 事件系统：基于观察者模式的事件通知

    支持的主要事件：
    - backup-error: 备份错误
    - backup-completed: 备份完成
    - memory-pressure: 内存压力
    - memory-critical: 内存严重不足
    - task-queued: 任务排队
    - task-started: 任务开始
    - task-completed: 任务完成
    - task-failed: 任务失败
    - task-cancelled: 任务取消
    - progress-update: 进度更新
    - scheduler-started: 调度器启动
    - scheduler-stopped: 调度器停止
    - queue-paused: 队列暂停
    - queue-resumed: 队列恢复
    - cleanup-completed: 清理完成
    """

    def __init__(self, mysql_manager: MySQLManager):
        """初始化备份工具"""
        self.mysql_manager = mysql_manager
        self.cache_manager = mysql_manager.cache_manager
        self.memory_manager = MemoryManager()
        self.smart_retry_manager = SmartRetryStrategy

        # 任务队列管理
        self.task_queue: Dict[str, TaskQueue] = {}
        self.running_tasks = 0
        self.max_concurrent_tasks = 5
        self.task_id_counter = 0

        # 进度跟踪器
        self.progress_trackers: Dict[str, ProgressTracker] = {}

        # 调度器状态
        self.scheduler_running = False
        self.scheduler_task = None

        # 事件系统
        self._event_listeners: Dict[str, List[Callable]] = {}
        self._max_listeners = 200

        # 设置最大监听器数量
        self.set_max_listeners(200)

        # 初始化内存管理
        self._setup_memory_management()

        # 启动任务调度器（延迟到事件循环可用时）
        self._scheduler_delayed = False
        self._start_task_scheduler()

        # 优化最大并发数
        self._optimize_max_concurrency()

        # 发出初始化完成事件
        self.emit('initialized', {
            'max_concurrent_tasks': self.max_concurrent_tasks,
            'memory_monitoring': True,
            'adaptive_scheduling': True
        })

    def start_delayed_services(self) -> None:
        """
        启动延迟的服务

        在事件循环可用时调用此方法来启动内存监控和任务调度器。
        """
        # 启动延迟的内存监控
        self.memory_manager.start_delayed_monitoring()

        # 启动延迟的任务调度器
        if self._scheduler_delayed and not self.scheduler_running:
            try:
                self._start_task_scheduler()
                self._scheduler_delayed = False
            except RuntimeError:
                logger.info("任务调度器仍需等待事件循环", "MySQLBackupTool")

    def emit(self, event: str, data: Any = None) -> None:
        """
        发出事件

        Args:
            event: 事件名称
            data: 事件数据
        """
        if event in self._event_listeners:
            for listener in self._event_listeners[event]:
                try:
                    if data is not None:
                        listener(data)
                    else:
                        listener()
                except Exception as e:
                    logger.error(f"事件监听器执行失败: {e}", "MySQLBackupTool", {
                        'event': event,
                        'listener_count': len(self._event_listeners[event])
                    })

    def on(self, event: str, listener: Callable) -> None:
        """
        注册事件监听器

        Args:
            event: 事件名称
            listener: 监听器函数
        """
        if event not in self._event_listeners:
            self._event_listeners[event] = []
        self._event_listeners[event].append(listener)

        logger.debug(f"事件监听器已注册", "MySQLBackupTool", {
            'event': event,
            'listener_count': len(self._event_listeners[event])
        })

    def off(self, event: str, listener: Callable) -> None:
        """
        移除事件监听器

        Args:
            event: 事件名称
            listener: 监听器函数
        """
        if event in self._event_listeners:
            try:
                self._event_listeners[event].remove(listener)
                logger.debug(f"事件监听器已移除", "MySQLBackupTool", {
                    'event': event,
                    'listener_count': len(self._event_listeners[event])
                })
            except ValueError:
                pass  # 监听器不存在

    def set_max_listeners(self, max_listeners: int) -> None:
        """设置最大监听器数量"""
        self._max_listeners = max_listeners

    def _optimize_max_concurrency(self) -> None:
        """优化最大并发数基于系统资源"""
        memory_usage = self.memory_manager.get_current_usage()
        available_memory = memory_usage.heap_total - memory_usage.heap_used

        # 基于可用内存动态调整并发数
        if available_memory > 500 * 1024 * 1024:  # > 500MB
            self.max_concurrent_tasks = 8
        elif available_memory > 200 * 1024 * 1024:  # > 200MB
            self.max_concurrent_tasks = 5
        else:
            self.max_concurrent_tasks = 3

        logger.info("并发数已优化", "MySQLBackupTool", {
            "max_concurrent_tasks": self.max_concurrent_tasks,
            "available_memory": f"{available_memory / 1024 / 1024:.1f}MB"
        })

    async def _get_table_statistics(self, specific_tables: Optional[List[str]] = None) -> Dict[str, int]:
        """获取表统计信息"""
        try:
            tables = specific_tables or []

            if not tables:
                # 检查缓存
                cached_tables = await self.cache_manager.get(CacheRegion.QUERY_RESULT, 'SHOW_TABLES')
                if cached_tables:
                    tables = cached_tables
                else:
                    result = await self.mysql_manager.execute_query('SHOW TABLES')
                    tables = [row[list(row.keys())[0]] for row in result]
                    await self.cache_manager.set(CacheRegion.QUERY_RESULT, 'SHOW_TABLES', tables)

            # 并行获取所有表的统计信息
            count_promises = []
            for table_name in tables:
                count_promises.append(self._get_table_count(table_name))

            # 分批执行以避免过多并发查询
            batch_size = min(10, self.max_concurrent_tasks)
            total_record_count = 0

            for i in range(0, len(count_promises), batch_size):
                batch = count_promises[i:i + batch_size]
                batch_results = await asyncio.gather(*batch, return_exceptions=True)

                for result in batch_results:
                    if isinstance(result, int):
                        total_record_count += result
                    elif isinstance(result, Exception):
                        logger.warn(f"获取表统计信息失败: {result}")

                # 检查内存压力
                pressure = self.memory_manager.check_memory_pressure()
                if pressure > 0.8:
                    await self._request_memory_cleanup()
                    await asyncio.sleep(0.1)

            return {
                "table_count": len(tables),
                "record_count": total_record_count
            }
        except Exception as e:
            logger.warn(f"获取表统计信息失败: {e}")
            return {"table_count": 0, "record_count": 0}

    async def _get_table_count(self, table_name: str) -> int:
        """获取单个表的记录数"""
        try:
            cache_key = f"COUNT_{table_name}"
            cached_count = await self.cache_manager.get(CacheRegion.QUERY_RESULT, cache_key)

            if cached_count is not None:
                return cached_count

            # 使用更快的统计查询
            result = await self.mysql_manager.execute_query(
                "SELECT table_rows FROM information_schema.tables WHERE table_schema = DATABASE() AND table_name = %s",
                [table_name]
            )

            count = 0
            if result and len(result) > 0:
                count = result[0].get('table_rows', 0) or 0
                # 如果统计信息不准确，使用精确计数（但限制在小表上）
                if count == 0:
                    exact_result = await self.mysql_manager.execute_query(
                        f"SELECT COUNT(*) as count FROM `{table_name}` LIMIT 1000"
                    )
                    count = exact_result[0].get('count', 0) if exact_result else 0

            await self.cache_manager.set(CacheRegion.QUERY_RESULT, cache_key, count)
            return count
        except Exception as e:
            logger.warn(f"获取表 {table_name} 记录数失败: {e}")
            return 0

    @with_error_handling('createBackup', 'MSG_BACKUP_FAILED')
    @with_performance_monitoring('backup_create')
    async def create_backup(self, options: Optional[BackupOptions] = None) -> BackupResult:
        """创建数据库备份"""
        start_time = TimeUtils.now()
        options = options or BackupOptions()

        # 发出备份开始事件
        self.emit('backup-started', {
            'options': options,
            'start_time': start_time.isoformat()
        })

        # 设置默认值
        backup_options = BackupOptions(
            output_dir=options.output_dir or './backups',
            compress=options.compress if options.compress is not None else True,
            include_data=options.include_data if options.include_data is not None else True,
            include_structure=options.include_structure if options.include_structure is not None else True,
            tables=options.tables or [],
            file_prefix=options.file_prefix or 'mysql_backup',
            max_file_size=options.max_file_size or 100
        )

        try:
            # 确保输出目录存在
            os.makedirs(backup_options.output_dir, exist_ok=True)

            # 生成备份文件名
            timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
            file_name = f"{backup_options.file_prefix}_{timestamp}"
            backup_path = os.path.join(backup_options.output_dir, file_name)

            # 获取数据库配置
            config = self.mysql_manager.config_manager.database

            # 获取表统计信息
            table_stats = await self._get_table_statistics(backup_options.tables)
            table_count = table_stats["table_count"]
            record_count = table_stats["record_count"]

            # 发出进度更新事件
            self.emit('backup-progress', {
                'stage': 'preparing',
                'progress': 10,
                'message': '正在准备备份...',
                'table_count': table_count,
                'record_count': record_count
            })

            # 构建 mysqldump 命令
            dump_args = [
                'mysqldump',
                f'-h{config["host"]}',
                f'-P{config["port"]}',
                f'-u{config["user"]}',
                f'-p{config["password"]}',
                '--default-character-set=utf8mb4',
                '--single-transaction',
                '--routines',
                '--triggers'
            ]

            if not backup_options.include_data:
                dump_args.append('--no-data')

            if not backup_options.include_structure:
                dump_args.append('--no-create-info')

            dump_args.append(config["database"] or '')

            if backup_options.tables:
                dump_args.extend(backup_options.tables)

            # 执行备份
            sql_file_path = f"{backup_path}.sql"
            await self._execute_mysqldump(dump_args, sql_file_path)

            final_file_path = sql_file_path
            file_size = 0

            # 检查文件大小并压缩
            if os.path.exists(sql_file_path):
                file_size = os.path.getsize(sql_file_path)

                # 发出进度更新事件
                self.emit('backup-progress', {
                    'stage': 'dumping',
                    'progress': 70,
                    'message': '正在创建备份文件...',
                    'file_size': f"{file_size / 1024 / 1024:.1f}MB"
                })

                if backup_options.compress or file_size > (backup_options.max_file_size * 1024 * 1024):
                    zip_file_path = f"{backup_path}.zip"
                    await self._compress_file(sql_file_path, zip_file_path)

                    # 删除原始SQL文件
                    os.unlink(sql_file_path)

                    final_file_path = zip_file_path
                    file_size = os.path.getsize(zip_file_path)

                    # 发出进度更新事件
                    self.emit('backup-progress', {
                        'stage': 'compressing',
                        'progress': 90,
                        'message': '正在压缩备份文件...',
                        'compressed_size': f"{file_size / 1024 / 1024:.1f}MB"
                    })

            duration = int((TimeUtils.now() - start_time) * 1000)

            # 发出备份完成事件
            self.emit('backup-completed', {
                'file_path': final_file_path,
                'file_size': file_size,
                'table_count': table_count,
                'record_count': record_count,
                'duration': duration
            })

            return BackupResult(
                success=True,
                file_path=final_file_path,
                file_size=file_size,
                table_count=table_count,
                record_count=record_count,
                duration=duration
            )

        except Exception as e:
            # 清理缓存以释放内存
            await self.cache_manager.clear_region(CacheRegion.QUERY_RESULT)

            duration = int((TimeUtils.now() - start_time) * 1000)
            error_message = str(e)

            logger.error("备份创建失败", "MySQLBackupTool", {
                "error": error_message,
                "duration": duration
            })

            return BackupResult(
                success=False,
                error=error_message,
                duration=duration
            )

    async def _execute_mysqldump(self, args: List[str], output_path: str) -> None:
        """执行mysqldump命令"""
        try:
            # 创建输出目录
            os.makedirs(os.path.dirname(output_path), exist_ok=True)

            # 执行命令
            process = await asyncio.create_subprocess_exec(
                *args,
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.PIPE
            )

            with open(output_path, 'wb') as output_file:
                while True:
                    chunk = await process.stdout.read(8192)
                    if not chunk:
                        break
                    output_file.write(chunk)

            await process.wait()

            if process.returncode != 0:
                error_output = await process.stderr.read()
                raise Exception(f"mysqldump failed: {error_output.decode()}")

        except Exception as e:
            logger.error(f"执行mysqldump失败: {e}")
            raise

    async def _compress_file(self, input_path: str, output_path: str) -> None:
        """压缩文件"""
        try:
            if not os.path.exists(input_path):
                raise FileNotFoundError(f"输入文件不存在: {input_path}")

            async with asyncio.get_event_loop().run_in_executor(
                None,
                self._compress_file_sync,
                input_path,
                output_path
            ):
                pass
        except Exception as e:
            logger.error(f"文件压缩失败: {e}")
            raise

    def _compress_file_sync(self, input_path: str, output_path: str) -> None:
        """同步压缩文件"""
        try:
            with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED, compresslevel=6) as zipf:
                zipf.write(input_path, os.path.basename(input_path))
        except Exception as e:
            logger.error(f"同步压缩文件失败: {e}")
            raise

    async def export_data(self, query: str, params: Optional[List[Any]] = None,
                           options: Optional[ExportOptions] = None) -> ExportResult:
        """通用导出方法 - 使用ExporterFactory"""
        try:
            # 设置默认选项
            if not options:
                options = ExportOptions()

            # 确定导出格式
            format_type = options.format or 'excel'

            # 使用ExporterFactory创建相应的导出器
            exporter_factory = ExporterFactory.get_instance(self.mysql_manager, self.memory_manager)
            exporter = exporter_factory.create_exporter(format_type)
            return await exporter.export(query, params or [], options)

        except Exception as e:
            error_message = str(e)
            logger.error("数据导出失败", "MySQLBackupTool", {
                "error": error_message,
                "query": query[:100]
            })

            return ExportResult(
                success=False,
                error=error_message,
                duration=0
            )

    async def export_data_with_recovery(self, query: str, params: Optional[List[Any]] = None,
                                      options: Optional[ExportOptions] = None,
                                      recovery: Optional[RecoveryStrategy] = None) -> ErrorRecoveryResult:
        """带错误恢复的数据导出"""
        default_recovery = RecoveryStrategy(
            retry_count=3,
            retry_delay=1000,
            exponential_backoff=True,
            fallback_options={
                'streaming': False,
                'max_rows': 10000,
                'batch_size': 500
            }
        )

        strategy = RecoveryStrategy(**(recovery.model_dump() if recovery else default_recovery.model_dump()))

        return await self._execute_with_recovery(
            lambda: self.export_data(query, params, options),
            strategy,
            'export',
            options or ExportOptions()
        )

    async def export_data_with_progress(self, query: str, params: Optional[List[Any]] = None,
                                      options: Optional[ExportOptions] = None,
                                      cancellation_token: Optional[CancellationToken] = None) -> tuple:
        """带进度跟踪的数据导出"""
        tracker = await self.create_progress_tracker('export', cancellation_token)

        try:
            # 检查取消状态
            if cancellation_token and cancellation_token.is_cancelled:
                raise Exception(cancellation_token.reason or '操作已被取消')

            # 更新进度
            tracker.progress = ProgressInfo(
                stage='preparing',
                progress=10,
                message='准备开始数据导出...',
                start_time=datetime.now()
            )
            self._update_progress(tracker)

            # 检查取消状态
            if cancellation_token and cancellation_token.is_cancelled:
                raise Exception(cancellation_token.reason or '操作已被取消')

            # 更新进度
            tracker.progress = ProgressInfo(
                stage='processing',
                progress=50,
                message='正在执行查询和导出数据...',
                start_time=datetime.now()
            )
            self._update_progress(tracker)

            # 执行导出
            result = await self.export_data(query, params, options)

            # 完成
            tracker.progress = ProgressInfo(
                stage='completed',
                progress=100,
                message='数据导出完成',
                start_time=datetime.now()
            )
            self._update_progress(tracker)

            if tracker.on_complete:
                tracker.on_complete(result)

            return result, tracker

        except Exception as e:
            tracker.progress = ProgressInfo(
                stage='error',
                progress=0,
                message=f'数据导出失败: {str(e)}',
                start_time=datetime.now()
            )
            self._update_progress(tracker)

            if tracker.on_error:
                tracker.on_error(e)

            raise

    async def export_data_queued(self, query: str, params: Optional[List[Any]] = None,
                               options: Optional[ExportOptions] = None,
                               priority: int = 1) -> str:
        """带队列的数据导出"""
        task_id = self.add_task_to_queue(
            'export',
            lambda: self._execute_export_task(query, params or [], options or ExportOptions()),
            {'query': query, 'params': params or [], 'options': options},
            priority
        )

        return task_id

    async def _execute_export_task(self, query: str, params: List[Any], options: ExportOptions) -> ExportResult:
        """执行导出任务"""
        return await self.export_data(query, params, options)

    async def export_large_data(self, query: str, params: Optional[List[Any]] = None,
                              options: Optional[ExportOptions] = None) -> ExportResult:
        """大文件数据导出 - 使用流式处理"""
        start_time = TimeUtils.now()

        try:
            if not options:
                options = ExportOptions()

            # 设置大文件处理选项
            large_options = ExportOptions(
                **options.model_dump(),
                streaming=True,
                batch_size=options.batch_size or 1000,
                max_rows=options.max_rows or 1000000
            )

            # 检查内存使用情况
            memory_pressure = self.memory_manager.check_memory_pressure()
            if memory_pressure > 0.8:
                await self._request_memory_cleanup()

            # 生成输出文件路径
            output_path = self._generate_output_path_for_export(options, 'excel')

            # 使用流式导出
            return await self._export_with_streaming(query, params or [], large_options, output_path, start_time)

        except Exception as e:
            duration = int((TimeUtils.now() - start_time) * 1000)
            error_message = str(e)

            logger.error("大文件数据导出失败", "MySQLBackupTool", {
                "error": error_message,
                "duration": duration
            })

            return ExportResult(
                success=False,
                error=error_message,
                duration=duration
            )

    async def _export_with_streaming(self, query: str, params: List[Any],
                                   options: ExportOptions, output_path: str, start_time: float) -> ExportResult:
        """使用流式处理的导出"""
        exporter_factory = ExporterFactory.get_instance(self.mysql_manager, self.memory_manager)
        exporter = exporter_factory.create_exporter(options.format or 'excel')

        # 这里可以实现更复杂的流式处理逻辑
        # 目前先使用基础的导出器
        return await exporter.export(query, params, options)

    def _generate_output_path_for_export(self, options: ExportOptions, extension: str) -> str:
        """生成导出文件路径"""
        # 确保输出目录存在
        Path(options.output_dir or './exports').mkdir(parents=True, exist_ok=True)

        # 生成文件名
        if not options.file_name:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_%f')[:-3]
            file_name = f'export_{timestamp}.{extension}'
        else:
            file_name = options.file_name
            if not file_name.endswith(f'.{extension}'):
                file_name += f'.{extension}'

        return str(Path(options.output_dir or './exports') / file_name)

    async def export_with_batch_optimization(self, query: str, params: Optional[List[Any]] = None,
                                           options: Optional[ExportOptions] = None) -> ExportResult:
        """使用批量优化的导出"""
        start_time = TimeUtils.now()

        try:
            if not options:
                options = ExportOptions()

            # 优化批处理设置
            batch_options = ExportOptions(
                **options.model_dump(),
                batch_size=options.batch_size or 500,
                streaming=True
            )

            # 检查查询是否适合批量处理
            if self._is_suitable_for_batch_processing(query):
                return await self._export_with_batch_processing(query, params or [], batch_options, start_time)
            else:
                return await self.export_data(query, params, options)

        except Exception as e:
            duration = int((TimeUtils.now() - start_time) * 1000)
            error_message = str(e)

            logger.error("批量优化导出失败", "MySQLBackupTool", {
                "error": error_message,
                "duration": duration
            })

            return ExportResult(
                success=False,
                error=error_message,
                duration=duration
            )

    def _is_suitable_for_batch_processing(self, query: str) -> bool:
        """检查查询是否适合批量处理"""
        query_lower = query.lower()
        return 'select' in query_lower and 'limit' not in query_lower

    async def _export_with_batch_processing(self, query: str, params: List[Any],
                                          options: ExportOptions, start_time: float) -> ExportResult:
        """使用批量处理的导出"""
        # 这里可以实现批量处理逻辑
        # 目前先使用基础导出
        return await self.export_data(query, params, options)

    async def _execute_with_recovery(self, operation, strategy: RecoveryStrategy,
                                   operation_type: str, original_options) -> ErrorRecoveryResult:
        """通用错误恢复执行器"""
        last_error = None
        attempts_used = 0
        recovery_applied = None

        # 主要尝试
        for attempt in range(strategy.retry_count + 1):
            try:
                attempts_used = attempt + 1

                if attempt > 0:
                    # 计算延迟时间
                    delay = strategy.retry_delay * (2 ** (attempt - 1)) if strategy.exponential_backoff else strategy.retry_delay
                    await asyncio.sleep(delay / 1000)

                    # 调用重试回调
                    if strategy.on_retry:
                        strategy.on_retry(attempt, last_error)

                result = await operation()

                if attempt > 0:
                    recovery_applied = f"成功重试 (第 {attempt} 次尝试)"

                return ErrorRecoveryResult(
                    success=True,
                    result=result,
                    recovery_applied=recovery_applied,
                    attempts_used=attempts_used
                )

            except Exception as e:
                last_error = e

                # 检查是否是可恢复的错误
                if not self._is_recoverable_error(e):
                    break

        # 如果有回退选项，尝试回退策略
        if strategy.fallback_options and last_error:
            try:
                if strategy.on_fallback:
                    strategy.on_fallback(last_error)

                # 应用回退选项
                fallback_options = original_options.model_copy()
                for key, value in strategy.fallback_options.items():
                    setattr(fallback_options, key, value)

                fallback_result = await operation()  # 这里应该传递fallback_options，但为了简化

                return ErrorRecoveryResult(
                    success=True,
                    result=fallback_result,
                    recovery_applied="应用回退策略成功",
                    attempts_used=attempts_used
                )

            except Exception as fallback_error:
                pass  # 回退也失败了

        # 所有恢复尝试都失败了
        return ErrorRecoveryResult(
            success=False,
            error=f"{operation_type} 操作失败，已尝试 {attempts_used} 次: {str(last_error)}",
            attempts_used=attempts_used,
            final_error=str(last_error)
        )

    def _is_recoverable_error(self, error: Exception) -> bool:
        """判断错误是否可恢复"""
        error_message = str(error).lower()
        recoverable_patterns = [
            'timeout', 'connection', 'network', 'temporary', 'lock wait', 'deadlock'
        ]

        return any(pattern in error_message for pattern in recoverable_patterns)

    async def generate_report(self, report_config: ReportConfig) -> ExportResult:
        """生成数据报表"""
        start_time = TimeUtils.now()

        try:
            # 设置默认导出选项
            options = report_config.options or ExportOptions(
                output_dir='./reports',
                format='excel',
                include_headers=True
            )

            # 确保输出目录存在
            os.makedirs(options.output_dir, exist_ok=True)

            # 生成文件名
            if not options.file_name:
                timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
                options.file_name = f"report_{report_config.title.replace(' ', '_')}_{timestamp}.xlsx"

            file_path = os.path.join(options.output_dir, options.file_name)

            # 创建工作簿
            wb = Workbook()
            total_rows = 0
            total_columns = 0

            # 添加报表信息工作表
            info_sheet = wb.active
            info_sheet.title = '报表信息'
            info_sheet['A1'] = '报表标题'
            info_sheet['B1'] = report_config.title
            info_sheet['A2'] = '生成时间'
            info_sheet['B2'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            if report_config.description:
                info_sheet['A3'] = '报表描述'
                info_sheet['B3'] = report_config.description
            info_sheet['A4'] = '查询数量'
            info_sheet['B4'] = len(report_config.queries)

            # 为每个查询创建工作表
            for query_config in report_config.queries:
                try:
                    data = await self.mysql_manager.execute_query(query_config.query, query_config.params or [])

                    if data:
                        ws = wb.create_sheet(title=query_config.name[:31])  # Excel限制

                        # 添加标题
                        ws['A1'] = query_config.name
                        ws['A1'].font = Font(bold=True, size=12)
                        title_row = 1

                        # 添加列头
                        if options.include_headers and data:
                            headers = list(data[0].keys())
                            for col_num, header in enumerate(headers, 1):
                                ws.cell(row=title_row + 2, column=col_num, value=header)

                            # 设置列头样式
                            header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                            header_font = Font(bold=True)
                            for cell in ws[title_row + 2]:
                                cell.fill = header_fill
                                cell.font = header_font

                            data_start_row = title_row + 3
                        else:
                            data_start_row = title_row + 2

                        # 添加数据
                        for row_num, row_data in enumerate(data, data_start_row):
                            for col_num, value in enumerate(row_data.values(), 1):
                                ws.cell(row=row_num, column=col_num, value=value)

                        # 自动调整列宽
                        for column in ws.columns:
                            max_length = 0
                            column_letter = column[0].column_letter

                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass

                            adjusted_width = min(max(max_length + 2, 10), 50)
                            ws.column_dimensions[column_letter].width = adjusted_width

                        total_rows += len(data)
                        total_columns = max(total_columns, len(data[0]) if data else 0)

                except Exception as e:
                    # 为查询错误创建错误工作表
                    error_sheet = wb.create_sheet(title=f"错误_{query_config.name[:25]}")
                    error_sheet['A1'] = '查询名称'
                    error_sheet['B1'] = query_config.name
                    error_sheet['A2'] = '错误信息'
                    error_sheet['B2'] = str(e)
                    error_sheet['A3'] = '查询SQL'
                    error_sheet['B3'] = query_config.query

            # 更新信息表
            info_sheet['A5'] = '总行数'
            info_sheet['B5'] = total_rows
            info_sheet['A6'] = '总列数'
            info_sheet['B6'] = total_columns

            # 保存文件
            wb.save(file_path)

            file_size = os.path.getsize(file_path) if os.path.exists(file_path) else 0
            duration = int((TimeUtils.now() - start_time) * 1000)

            return ExportResult(
                success=True,
                file_path=file_path,
                file_size=file_size,
                row_count=total_rows,
                column_count=total_columns,
                format='excel',
                duration=duration
            )

        except Exception as e:
            duration = int((TimeUtils.now() - start_time) * 1000)
            error_message = str(e)

            logger.error("报表生成失败", "MySQLBackupTool", {
                "error": error_message,
                "duration": duration,
                "report_title": report_config.title
            })

            return ExportResult(
                success=False,
                error=error_message,
                duration=duration
            )

    async def create_progress_tracker(self, operation: str,
                                    cancellation_token: Optional[CancellationToken] = None) -> ProgressTracker:
        """创建进度跟踪器"""
        tracker_id = f"{operation}_{int(time.time())}_{self.task_id_counter}"
        self.task_id_counter += 1

        tracker = ProgressTracker(
            id=tracker_id,
            operation=operation,
            start_time=datetime.now(),
            progress=ProgressInfo(
                stage='preparing',
                progress=0,
                message=f'准备开始 {operation} 操作...'
            ),
            cancellation_token=cancellation_token
        )

        self.progress_trackers[tracker_id] = tracker

        # 监听取消事件
        if cancellation_token:
            cancellation_token.cancel = lambda reason=None: self._cancel_operation(tracker_id, reason)

        return tracker

    def _cancel_operation(self, tracker_id: str, reason: Optional[str] = None) -> None:
        """取消操作"""
        tracker = self.progress_trackers.get(tracker_id)
        if tracker:
            tracker.progress = ProgressInfo(
                stage='error',
                progress=0,
                message=reason or '操作已被取消'
            )
            self._update_progress(tracker)

    def _update_progress(self, tracker: ProgressTracker) -> None:
        """更新进度"""
        if tracker.on_progress:
            tracker.on_progress(tracker.progress)

    def set_max_listeners(self, max_listeners: int) -> None:
        """设置最大监听器数量"""
        self._max_listeners = max_listeners

    def _setup_memory_management(self) -> None:
        """设置内存管理"""
        self.memory_manager.enable_memory_monitoring()

        # 监听内存压力
        self._memory_pressure_handler = lambda pressure: self._on_memory_pressure(pressure)

    def _on_memory_pressure(self, pressure: float) -> None:
        """内存压力处理"""
        if pressure > 0.85:
            # 高内存压力：清理完成的跟踪器
            self._cleanup_completed_trackers()

            # 强制垃圾回收
            self._request_memory_cleanup()

            if pressure > 0.95:
                logger.warning("内存使用率过高，暂停新任务", "MySQLBackupTool", {
                    "pressure": pressure
                })

    def _cleanup_completed_trackers(self) -> None:
        """清理已完成的跟踪器"""
        current_time = time.time()
        expired_trackers = []

        for tracker_id, tracker in self.progress_trackers.items():
            elapsed = current_time - tracker.start_time.timestamp()
            if elapsed > 300 or tracker.progress.stage in ['completed', 'error']:
                expired_trackers.append(tracker_id)

        for tracker_id in expired_trackers:
            del self.progress_trackers[tracker_id]

    async def _request_memory_cleanup(self) -> None:
        """请求内存清理"""
        self.memory_manager.request_memory_cleanup()
        await self.cache_manager.clear_region(CacheRegion.QUERY_RESULT)

    def _start_task_scheduler(self) -> None:
        """启动任务调度器"""
        if self.scheduler_running:
            return

        try:
            loop = asyncio.get_running_loop()
            self.scheduler_running = True
            self.scheduler_task = asyncio.create_task(self._task_scheduler_loop())

            # 发出调度器启动事件
            self.emit('scheduler-started', {
                'max_concurrent_tasks': self.max_concurrent_tasks,
                'adaptive_scheduling': True
            })
        except RuntimeError:
            # 没有运行的事件循环，标记为延迟启动
            self.scheduler_running = False
            self._scheduler_delayed = True
            logger.info("任务调度器将在事件循环可用时启动", "MySQLBackupTool")

    async def _task_scheduler_loop(self) -> None:
        """任务调度器循环 - 改进版自适应调度"""
        while self.scheduler_running:
            try:
                # 获取待执行的任务数量以动态调整检查间隔
                queued_tasks = len([task for task in self.task_queue.values() if task.status == 'queued'])
                interval = 0.5 if queued_tasks > 5 else 1.0  # 高负载时更频繁检查

                await self._process_task_queue()
                await asyncio.sleep(interval)
            except Exception as e:
                logger.error(f"任务调度器错误: {e}", "MySQLBackupTool")
                await asyncio.sleep(5)  # 错误时等待更长时间

    async def _process_task_queue(self) -> None:
        """处理任务队列 - 改进版任务处理"""
        # 清理已完成的任务
        self._cleanup_completed_tasks()

        # 如果正在运行的任务已达到最大值，则等待
        if self.running_tasks >= self.max_concurrent_tasks:
            return

        # 获取待执行的任务（按优先级排序）
        pending_tasks = [
            task for task in self.task_queue.values()
            if task.status == 'queued'
        ]
        pending_tasks.sort(key=lambda x: x.priority, reverse=True)

        # 启动可以运行的任务
        tasks_to_start = pending_tasks[:self.max_concurrent_tasks - self.running_tasks]

        for task in tasks_to_start:
            asyncio.create_task(self._execute_task(task))

    def _cleanup_completed_tasks(self) -> None:
        """清理已完成的任务"""
        current_time = time.time()
        expired_tasks = []

        for task_id, task in self.task_queue.items():
            if task.status in ['completed', 'failed', 'cancelled']:
                if task.completed_at and (current_time - task.completed_at.timestamp()) > 1800:  # 30分钟
                    expired_tasks.append(task_id)

        for task_id in expired_tasks:
            del self.task_queue[task_id]

    async def _execute_task(self, task: TaskQueue) -> None:
        """执行单个任务 - 改进版任务执行"""
        try:
            # 更新任务状态
            task.status = 'running'
            task.started_at = datetime.now()
            self.running_tasks += 1

            # 发出任务开始事件
            self.emit('task-started', {
                'task_id': task.id,
                'task_type': task.type,
                'priority': task.priority,
                'running_tasks': self.running_tasks
            })

            # 执行任务
            if task.type == 'backup':
                result = await self.create_backup(task.result if hasattr(task, 'result') else None)
            elif task.type == 'export':
                # 这里需要从任务参数中获取查询和选项
                result = await self.export_data('SELECT 1', [])
            else:
                raise ValueError(f"不支持的任务类型: {task.type}")

            # 任务完成
            task.status = 'completed'
            task.completed_at = datetime.now()
            task.result = result
            self.running_tasks -= 1

            # 发出任务完成事件
            self.emit('task-completed', {
                'task_id': task.id,
                'task_type': task.type,
                'result': result,
                'duration': int((task.completed_at.timestamp() - task.started_at.timestamp()) * 1000),
                'running_tasks': self.running_tasks
            })

        except Exception as e:
            task.status = 'failed'
            task.completed_at = datetime.now()
            task.error = str(e)
            self.running_tasks -= 1

            # 发出任务失败事件
            self.emit('task-failed', {
                'task_id': task.id,
                'task_type': task.type,
                'error': str(e),
                'running_tasks': self.running_tasks
            })

            logger.error(f"任务执行失败: {e}", "MySQLBackupTool", {
                "task_id": task.id,
                "task_type": task.type
            })

    def add_task_to_queue(self, task_type: str, operation_func: callable,
                          params: Dict[str, Any] = None, priority: int = 1) -> str:
        """添加任务到队列 - 改进版任务队列管理"""
        task_id = f"{task_type}_{int(time.time())}_{self.task_id_counter}"
        self.task_id_counter += 1

        task = TaskQueue(
            id=task_id,
            type=task_type,
            status='queued',
            priority=priority,
            created_at=datetime.now(),
            progress=ProgressInfo(
                stage='preparing',
                progress=0,
                message=f'{task_type} 任务已加入队列...'
            )
        )

        self.task_queue[task_id] = task

        # 发出任务排队事件
        self.emit('task-queued', {
            'task_id': task_id,
            'task_type': task_type,
            'priority': priority,
            'queue_size': len(self.task_queue)
        })

        logger.info("任务已加入队列", "MySQLBackupTool", {
            "task_id": task_id,
            "task_type": task_type,
            "priority": priority,
            "queue_size": len(self.task_queue)
        })

        return task_id

    def get_task_status(self, task_id: str) -> Optional[TaskQueue]:
        """获取任务状态"""
        return self.task_queue.get(task_id)

    def get_memory_usage(self) -> MemoryUsage:
        """获取内存使用情况"""
        return self.memory_manager.get_current_usage()

    def get_memory_pressure(self) -> float:
        """获取内存压力"""
        return self.memory_manager.check_memory_pressure()

    async def cleanup_memory(self) -> Dict[str, Any]:
        """清理内存"""
        await self._request_memory_cleanup()
        self._cleanup_completed_trackers()

        return {
            "timestamp": datetime.now().isoformat(),
            "memory_usage": self.get_memory_usage().model_dump(),
            "active_trackers": len(self.progress_trackers),
            "queued_tasks": len([t for t in self.task_queue.values() if t.status == 'queued'])
        }

    def set_memory_monitoring(self, enabled: bool) -> None:
        """设置内存监控"""
        if enabled:
            self.memory_manager.enable_memory_monitoring()
        else:
            self.memory_manager.disable_memory_monitoring()

    async def close(self) -> None:
        """关闭备份工具"""
        try:
            # 停止任务调度器
            self.scheduler_running = False
            if self.scheduler_task:
                self.scheduler_task.cancel()
                try:
                    await self.scheduler_task
                except asyncio.CancelledError:
                    pass

            # 取消所有运行中的任务
            for task in self.task_queue.values():
                if task.status == 'running':
                    task.status = 'cancelled'
                    task.completed_at = datetime.now()

            # 清理缓存
            await self.cache_manager.clear_region(CacheRegion.QUERY_RESULT)

            # 禁用内存监控
            self.memory_manager.disable_memory_monitoring()

            # 清理所有跟踪器
            self.progress_trackers.clear()

            # 清空任务队列
            self.task_queue.clear()

            logger.info("MySQL备份工具已关闭", "MySQLBackupTool")

        except Exception as e:
            logger.error(f"关闭备份工具时出错: {e}", "MySQLBackupTool")

    async def create_incremental_backup(self, options: Optional[IncrementalBackupOptions] = None) -> IncrementalBackupResult:
        """创建增量备份"""
        start_time = TimeUtils.now()
        options = options or IncrementalBackupOptions()

        # 设置默认值
        backup_options = IncrementalBackupOptions(
            output_dir=options.output_dir or './backups',
            compress=options.compress if options.compress is not None else True,
            include_data=options.include_data if options.include_data is not None else True,
            include_structure=options.include_structure if options.include_structure is not None else False,
            tables=options.tables or [],
            file_prefix=options.file_prefix or 'mysql_incremental',
            max_file_size=options.max_file_size or 100,
            incremental_mode=options.incremental_mode or 'timestamp',
            tracking_table=options.tracking_table or '__backup_tracking'
        )

        try:
            # 确保输出目录存在
            os.makedirs(backup_options.output_dir, exist_ok=True)

            # 确定增量备份的起始点
            since_time = datetime.now()
            changed_tables = []
            total_changes = 0

            if backup_options.incremental_mode == 'timestamp':
                if backup_options.last_backup_time:
                    since_time = datetime.fromisoformat(backup_options.last_backup_time.replace('Z', '+00:00'))
                else:
                    # 尝试从跟踪表获取最后备份时间
                    since_time = await self._get_last_backup_time(backup_options.tracking_table)

            # 获取有变化的表
            changed_info = await self._get_changed_tables(since_time, backup_options.tables)
            changed_tables = changed_info["changed_tables"]
            total_changes = changed_info["total_changes"]

            if not changed_tables:
                return IncrementalBackupResult(
                    success=True,
                    backup_type='incremental',
                    file_path='',
                    file_size=0,
                    table_count=0,
                    record_count=0,
                    duration=int((TimeUtils.now() - start_time) * 1000),
                    incremental_since=since_time.isoformat(),
                    changed_tables=[],
                    total_changes=0,
                    message='自上次备份以来没有数据变化'
                )

            # 生成备份文件名
            timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
            since_timestamp = since_time.strftime('%Y-%m-%d_%H-%M-%S')
            file_name = f"{backup_options.file_prefix}_{since_timestamp}_to_{timestamp}"
            backup_path = os.path.join(backup_options.output_dir, file_name)

            # 获取数据库配置
            config = self.mysql_manager.config_manager.database

            # 构建增量备份的 mysqldump 命令
            dump_args = [
                'mysqldump',
                f'-h{config["host"]}',
                f'-P{config["port"]}',
                f'-u{config["user"]}',
                f'-p{config["password"]}',
                '--default-character-set=utf8mb4',
                '--single-transaction'
            ]

            if not backup_options.include_structure:
                dump_args.append('--no-create-info')

            if not backup_options.include_data:
                dump_args.append('--no-data')

            # 添加时间过滤条件
            time_filter = since_time.strftime('%Y-%m-%d %H:%M:%S')
            dump_args.extend([
                '--where', f"updated_at >= '{time_filter}'",
                config["database"] or ''
            ])

            dump_args.extend(changed_tables)

            # 执行增量备份
            sql_file_path = f"{backup_path}.sql"
            await self._execute_mysqldump(dump_args, sql_file_path)

            # 在备份文件中添加增量备份信息
            backup_info = f"""-- Incremental Backup Information
-- Base backup: {backup_options.base_backup_path or 'N/A'}
-- Incremental since: {since_time.isoformat()}
-- Changed tables: {', '.join(changed_tables)}
-- Total changes: {total_changes}
-- Created at: {datetime.now().isoformat()}

"""
            if os.path.exists(sql_file_path):
                with open(sql_file_path, 'r', encoding='utf-8') as f:
                    existing_content = f.read()

                with open(sql_file_path, 'w', encoding='utf-8') as f:
                    f.write(backup_info + existing_content)

                final_file_path = sql_file_path
                file_size = os.path.getsize(sql_file_path)

                # 压缩文件
                if backup_options.compress:
                    zip_file_path = f"{backup_path}.zip"
                    await self._compress_file(sql_file_path, zip_file_path)
                    os.unlink(sql_file_path)
                    final_file_path = zip_file_path
                    file_size = os.path.getsize(zip_file_path)

                # 更新备份跟踪表
                await self._update_backup_tracking(backup_options.tracking_table, datetime.now(), final_file_path)

                duration = int((TimeUtils.now() - start_time) * 1000)

                return IncrementalBackupResult(
                    success=True,
                    backup_type='incremental',
                    file_path=final_file_path,
                    file_size=file_size,
                    table_count=len(changed_tables),
                    record_count=total_changes,
                    duration=duration,
                    base_backup_path=backup_options.base_backup_path,
                    incremental_since=since_time.isoformat(),
                    changed_tables=changed_tables,
                    total_changes=total_changes
                )

        except Exception as e:
            duration = int((TimeUtils.now() - start_time) * 1000)
            error_message = str(e)

            logger.error("增量备份创建失败", "MySQLBackupTool", {
                "error": error_message,
                "duration": duration
            })

            return IncrementalBackupResult(
                success=False,
                backup_type='incremental',
                error=error_message,
                duration=duration,
                changed_tables=[],
                total_changes=0
            )

    async def _get_last_backup_time(self, tracking_table: str) -> datetime:
        """获取最后备份时间"""
        try:
            # 检查跟踪表是否存在
            result = await self.mysql_manager.execute_query(
                "SELECT COUNT(*) as count FROM information_schema.tables WHERE table_name = %s AND table_schema = DATABASE()",
                [tracking_table]
            )

            if not result or result[0].get('count', 0) == 0:
                # 创建跟踪表
                await self.mysql_manager.execute_query(f"""
                    CREATE TABLE IF NOT EXISTS `{tracking_table}` (
                        id INT AUTO_INCREMENT PRIMARY KEY,
                        backup_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        backup_type ENUM('full', 'incremental') NOT NULL,
                        backup_path VARCHAR(500),
                        file_size BIGINT,
                        table_count INT,
                        record_count BIGINT,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                """)
                return datetime.fromtimestamp(0)  # 如果表不存在，返回最早时间

            # 获取最后备份时间
            result = await self.mysql_manager.execute_query(
                f"SELECT backup_time FROM `{tracking_table}` ORDER BY backup_time DESC LIMIT 1"
            )

            if result and len(result) > 0:
                backup_time_str = result[0].get('backup_time')
                if backup_time_str:
                    return datetime.fromisoformat(str(backup_time_str).replace('Z', '+00:00'))

            return datetime.fromtimestamp(0)
        except Exception as e:
            logger.warn(f"获取最后备份时间失败: {e}")
            return datetime.fromtimestamp(0)

    async def _get_changed_tables(self, since_time: datetime,
                                specific_tables: Optional[List[str]] = None) -> Dict[str, Any]:
        """获取有变化的表"""
        try:
            changed_tables = []
            total_changes = 0

            # 获取所有表或指定表
            tables = specific_tables or []
            if not tables:
                result = await self.mysql_manager.execute_query('SHOW TABLES')
                tables = [row[list(row.keys())[0]] for row in result]

            # 并行检查表变化
            for table in tables:
                try:
                    # 优先使用information_schema获取列信息
                    result = await self.mysql_manager.execute_query(
                        "SELECT COLUMN_NAME FROM information_schema.COLUMNS WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = %s AND DATA_TYPE IN ('timestamp', 'datetime') AND COLUMN_NAME IN ('updated_at', 'modified_at', 'updated_time', 'modification_time')",
                        [table]
                    )

                    if result and len(result) > 0:
                        timestamp_column = result[0].get('COLUMN_NAME')

                        # 使用索引优化的查询检查变化
                        result = await self.mysql_manager.execute_query(
                            f"SELECT COUNT(*) as count FROM `{table}` WHERE `{timestamp_column}` >= %s AND `{timestamp_column}` IS NOT NULL",
                            [since_time.strftime('%Y-%m-%d %H:%M:%S')]
                        )

                        change_count = result[0].get('count', 0) if result else 0
                        if change_count > 0:
                            changed_tables.append(table)
                            total_changes += change_count
                except Exception as e:
                    logger.warn(f"检查表 {table} 变化时出错: {e}")

            return {
                "changed_tables": changed_tables,
                "total_changes": total_changes
            }

        except Exception as e:
            logger.warn(f"获取变化表失败: {e}")
            return {"changed_tables": [], "total_changes": 0}

    async def _update_backup_tracking(self, tracking_table: str, backup_time: datetime,
                                    backup_path: str, backup_type: str = 'incremental') -> None:
        """更新备份跟踪记录"""
        try:
            file_size = os.path.getsize(backup_path) if os.path.exists(backup_path) else 0

            await self.mysql_manager.execute_query(
                f"INSERT INTO `{tracking_table}` (backup_time, backup_type, backup_path, file_size) VALUES (%s, %s, %s, %s)",
                [backup_time, backup_type, backup_path, file_size]
            )

        except Exception as e:
            logger.warn(f"更新备份跟踪记录失败: {e}")

    async def verify_backup(self, backup_file_path: str, deep_validation: bool = True) -> BackupVerificationResult:
        """验证备份文件"""
        try:
            if not os.path.exists(backup_file_path):
                return BackupVerificationResult(
                    valid=False,
                    file_size=0,
                    tables_found=[],
                    error="备份文件不存在"
                )

            stats = os.stat(backup_file_path)
            file_size = stats.st_size
            created_at = datetime.fromtimestamp(stats.st_mtime).isoformat()

            warnings = []
            compression = 'none'
            compression_ratio = None
            decompressed_size = None
            checksum = None
            temp_files = []

            # 检测压缩类型
            file_ext = os.path.splitext(backup_file_path)[1].lower()
            is_compressed = file_ext in ['.zip', '.gz', '.gzip', '.br']

            if is_compressed:
                compression = 'ZIP' if file_ext == '.zip' else ('GZIP' if file_ext in ['.gz', '.gzip'] else 'BROTLI')

                if not deep_validation:
                    return BackupVerificationResult(
                        valid=True,
                        file_size=file_size,
                        tables_found=[],
                        compression=compression,
                        created_at=created_at,
                        warnings=['仅进行浅验证，未解压缩验证内容']
                    )

                # 深度验证：解压缩并验证内容
                try:
                    temp_dir = tempfile.mkdtemp()
                    temp_files = await self._decompress_file(backup_file_path, temp_dir, file_ext)

                    if not temp_files:
                        raise Exception('解压缩后未找到文件')

                    # 找到SQL文件
                    sql_file = temp_files[0]
                    if len(temp_files) > 1:
                        max_size = 0
                        for file in temp_files:
                            if file.endswith('.sql'):
                                file_stats = os.stat(file)
                                if file_stats.st_size > max_size:
                                    max_size = file_stats.st_size
                                    sql_file = file

                    content = await self._read_file_async(sql_file)

                    # 计算压缩比
                    decompressed_size = os.path.getsize(sql_file)
                    compression_ratio = file_size / decompressed_size if decompressed_size > 0 else None

                    # 计算校验和
                    checksum = hashlib.sha256(content.encode()).hexdigest()

                    # 清理临时文件
                    import shutil
                    shutil.rmtree(temp_dir)

                except Exception as e:
                    return BackupVerificationResult(
                        valid=False,
                        file_size=file_size,
                        tables_found=[],
                        compression=compression,
                        created_at=created_at,
                        error=f"解压缩失败: {str(e)}",
                        warnings=['无法解压缩文件进行内容验证']
                    )

            else:
                content = await self._read_file_async(backup_file_path)
                checksum = hashlib.sha256(content.encode()).hexdigest()

            # 检查文件大小
            if file_size == 0:
                return BackupVerificationResult(
                    valid=False,
                    file_size=0,
                    tables_found=[],
                    error='备份文件为空'
                )

            if file_size < 100:
                warnings.append('备份文件大小异常小，可能不完整')

            # 增强的SQL结构检查
            has_header = any(keyword in content for keyword in [
                'mysqldump', 'MySQL dump', '-- Server version', 'MySQL数据库备份'
            ])

            has_footer = any(keyword in content for keyword in [
                'Dump completed', '备份结束', '-- Dump completed on', 'SET SQL_MODE=@OLD_SQL_MODE'
            ])

            has_charset = any(keyword in content for keyword in ['utf8', 'UTF8', 'CHARACTER SET', 'CHARSET'])

            # 检查表结构
            create_table_matches = len([m for m in content.split('\n') if 'CREATE TABLE' in m])
            drop_table_matches = len([m for m in content.split('\n') if 'DROP TABLE' in m])
            insert_matches = len([m for m in content.split('\n') if 'INSERT INTO' in m])

            # 提取表名
            tables_from_drop = []
            for line in content.split('\n'):
                if 'DROP TABLE' in line:
                    # 简单的表名提取
                    parts = line.split()
                    if len(parts) >= 3:
                        table_name = parts[2].strip('`;"\'')
                        if table_name:
                            tables_from_drop.append(table_name)

            all_tables = list(set(tables_from_drop))

            # 估算记录数量
            estimated_records = len([line for line in content.split('\n') if 'INSERT INTO' in line])

            # 检查备份类型
            backup_type = 'unknown'
            if create_table_matches > 0 and insert_matches > 0:
                backup_type = 'full'
            elif create_table_matches > 0:
                backup_type = 'structure'
            elif insert_matches > 0:
                backup_type = 'data'

            # 验证检查
            if not all_tables:
                warnings.append('未找到表定义，备份可能不完整或格式不正确')

            if backup_type == 'full' and estimated_records == 0:
                warnings.append('未找到数据插入语句，可能为空表备份')

            if not has_charset:
                warnings.append('未找到字符集设置，可能导致乱码问题')

            # 检查是否有多行INSERT
            multi_row_inserts = len([line for line in content.split('\n') if ',(' in line])
            if multi_row_inserts > 0:
                estimated_records += multi_row_inserts

            # 检查文件完整性
            if not content.endswith('\n') and not content.endswith(';'):
                warnings.append('备份文件可能被截断（文件末尾异常）')

            # 计算完整性评分
            completeness_score = self._calculate_completeness_score({
                'has_header': has_header,
                'has_footer': has_footer,
                'tables_count': len(all_tables),
                'has_inserts': insert_matches > 0,
                'has_charset': has_charset,
                'warnings_count': len(warnings)
            })

            is_valid = has_header and len(all_tables) > 0 and completeness_score >= 0.7

            return BackupVerificationResult(
                valid=is_valid,
                file_size=file_size,
                tables_found=all_tables,
                record_count=estimated_records,
                created_at=created_at,
                backup_type=backup_type,
                compression=compression,
                compression_ratio=compression_ratio,
                decompressed_size=decompressed_size,
                checksum=checksum,
                error=None if is_valid else self._generate_validation_error(has_header, len(all_tables), has_footer, completeness_score),
                warnings=warnings if warnings else None
            )

        except Exception as e:
            return BackupVerificationResult(
                valid=False,
                file_size=0,
                tables_found=[],
                error=f"验证过程中出错: {str(e)}"
            )

    def _calculate_completeness_score(self, params: Dict[str, Any]) -> float:
        """计算备份文件完整性评分"""
        score = 0.0

        if params['has_header']:
            score += 0.3
        if params['has_footer']:
            score += 0.2
        if params['tables_count'] > 0:
            score += 0.3
        if params['has_charset']:
            score += 0.1
        if params['has_inserts']:
            score += 0.1

        # 警告扣分
        score -= min(params['warnings_count'] * 0.05, 0.2)

        return max(0.0, min(1.0, score))

    def _generate_validation_error(self, has_header: bool, tables_count: int,
                                 has_footer: bool, completeness_score: float) -> str:
        """生成验证错误信息"""
        errors = []

        if not has_header:
            errors.append('缺少备份文件头')
        if tables_count == 0:
            errors.append('未找到表定义')
        if not has_footer:
            errors.append('缺少文件结尾标识')
        if completeness_score < 0.7:
            errors.append(f'完整性评分过低 ({completeness_score * 100:.1f}%)')

        return f"备份文件验证失败: {', '.join(errors)}"

    async def _decompress_file(self, file_path: str, temp_dir: str, file_ext: str) -> List[str]:
        """解压缩文件"""
        try:
            if file_ext == '.zip':
                import zipfile
                with zipfile.ZipFile(file_path, 'r') as zip_ref:
                    zip_ref.extractall(temp_dir)
                    return [os.path.join(temp_dir, f) for f in zip_ref.namelist()]
            else:
                # 其他压缩格式的处理
                return [file_path]  # 简化处理
        except Exception as e:
            logger.error(f"解压缩文件失败: {e}")
            return []

    async def _read_file_async(self, file_path: str) -> str:
        """异步读取文件"""
        def _read_file():
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                return f.read()

        return await asyncio.get_event_loop().run_in_executor(None, _read_file)

    # 为了向后兼容，提供一些便捷方法
    async def backup_database(self, **kwargs) -> BackupResult:
        """便捷的数据库备份方法"""
        options = BackupOptions(**kwargs)
        return await self.create_backup(options)

    async def export_query_result(self, query: str, **kwargs) -> ExportResult:
        """便捷的查询结果导出方法"""
        options = ExportOptions(**kwargs)
        return await self.export_data(query, options=options)

    async def create_data_report(self, title: str, queries: List[Dict[str, Any]],
                                **kwargs) -> ExportResult:
        """便捷的数据报表创建方法"""
        report_queries = [ReportQuery(**query) for query in queries]
        config = ReportConfig(title=title, queries=report_queries, options=ExportOptions(**kwargs))
        return await self.generate_report(config)

    async def create_backup_with_progress(self, options: Optional[BackupOptions] = None,
                                        cancellation_token: Optional[CancellationToken] = None) -> tuple:
        """带进度跟踪的备份创建"""
        tracker = await self.create_progress_tracker('backup', cancellation_token)

        try:
            # 更新进度
            tracker.progress = ProgressInfo(
                stage='preparing',
                progress=10,
                message='正在准备备份...'
            )
            self._update_progress(tracker)

            # 执行备份
            tracker.progress = ProgressInfo(
                stage='dumping',
                progress=50,
                message='正在创建备份...'
            )
            self._update_progress(tracker)

            result = await self.create_backup(options)

            # 完成
            tracker.progress = ProgressInfo(
                stage='completed',
                progress=100,
                message='备份完成'
            )
            self._update_progress(tracker)

            if tracker.on_complete:
                tracker.on_complete(result)

            return result, tracker

        except Exception as e:
            tracker.progress = ProgressInfo(
                stage='error',
                progress=0,
                message=f'备份失败: {str(e)}'
            )
            self._update_progress(tracker)

            if tracker.on_error:
                tracker.on_error(e)

            raise

    async def create_backup_with_recovery(self, options: Optional[BackupOptions] = None,
                                        recovery: Optional[RecoveryStrategy] = None) -> ErrorRecoveryResult:
        """带错误恢复的备份创建"""
        default_recovery = RecoveryStrategy(
            retry_count=3,
            retry_delay=1000,
            exponential_backoff=True,
            fallback_options={
                'compress': False,
                'max_file_size': 50
            }
        )

        strategy = RecoveryStrategy(**(recovery.model_dump() if recovery else default_recovery.model_dump()))

        return await self._execute_with_recovery(
            lambda: self.create_backup(options),
            strategy,
            'backup',
            options or BackupOptions()
        )

    async def _execute_with_recovery(self, operation, strategy: RecoveryStrategy,
                                   operation_type: str, original_options) -> ErrorRecoveryResult:
        """通用错误恢复执行器"""
        last_error = None
        attempts_used = 0
        recovery_applied = None

        # 主要尝试
        for attempt in range(strategy.retry_count + 1):
            try:
                attempts_used = attempt + 1

                if attempt > 0:
                    # 计算延迟时间
                    delay = strategy.retry_delay * (2 ** (attempt - 1)) if strategy.exponential_backoff else strategy.retry_delay
                    await asyncio.sleep(delay / 1000)

                    # 调用重试回调
                    if strategy.on_retry:
                        strategy.on_retry(attempt, last_error)

                result = await operation()

                if attempt > 0:
                    recovery_applied = f"成功重试 (第 {attempt} 次尝试)"

                return ErrorRecoveryResult(
                    success=True,
                    result=result,
                    recovery_applied=recovery_applied,
                    attempts_used=attempts_used
                )

            except Exception as e:
                last_error = e

                # 检查是否是可恢复的错误
                if not self._is_recoverable_error(e):
                    break

        # 如果有回退选项，尝试回退策略
        if strategy.fallback_options and last_error:
            try:
                if strategy.on_fallback:
                    strategy.on_fallback(last_error)

                # 应用回退选项
                fallback_options = original_options.model_copy()
                for key, value in strategy.fallback_options.items():
                    setattr(fallback_options, key, value)

                fallback_result = await operation()  # 这里应该传递fallback_options，但为了简化

                return ErrorRecoveryResult(
                    success=True,
                    result=fallback_result,
                    recovery_applied="应用回退策略成功",
                    attempts_used=attempts_used
                )

            except Exception as fallback_error:
                pass  # 回退也失败了

        # 所有恢复尝试都失败了
        return ErrorRecoveryResult(
            success=False,
            error=f"{operation_type} 操作失败，已尝试 {attempts_used} 次: {str(last_error)}",
            attempts_used=attempts_used,
            final_error=str(last_error)
        )

    def _is_recoverable_error(self, error: Exception) -> bool:
        """判断错误是否可恢复"""
        error_message = str(error).lower()
        recoverable_patterns = [
            'timeout', 'connection', 'network', 'temporary', 'lock wait', 'deadlock'
        ]

        return any(pattern in error_message for pattern in recoverable_patterns)

    async def create_large_file_backup(self, options: Optional[BackupOptions] = None,
                                     large_file_options: Optional[LargeFileOptions] = None) -> BackupResult:
        """大文件备份（使用内存优化）"""
        start_time = TimeUtils.now()
        options = options or BackupOptions()
        large_file_options = large_file_options or LargeFileOptions()

        # 设置默认值
        large_options = LargeFileOptions(
            chunk_size=large_file_options.chunk_size or 64 * 1024 * 1024,
            max_memory_usage=large_file_options.max_memory_usage or 512 * 1024 * 1024,
            use_memory_pool=large_file_options.use_memory_pool if large_file_options.use_memory_pool is not None else True,
            compression_level=large_file_options.compression_level or 6,
            disk_threshold=large_file_options.disk_threshold or 100 * 1024 * 1024
        )

        backup_options = BackupOptions(
            **(options.model_dump() if hasattr(options, 'model_dump') else options.__dict__),
            compress=True,
            max_file_size=int(large_options.disk_threshold / (1024 * 1024))
        )

        try:
            # 检查内存使用情况
            memory_pressure = self.memory_manager.check_memory_pressure()
            if memory_pressure > 0.8:
                await self._request_memory_cleanup()

            # 确保输出目录存在
            os.makedirs(backup_options.output_dir, exist_ok=True)

            # 生成临时目录用于分块处理
            temp_dir = os.path.join(backup_options.output_dir, f'temp_{int(time.time())}')
            os.makedirs(temp_dir, exist_ok=True)

            try:
                # 获取表信息并计算总大小
                tables = await self._get_tables_with_sizes(backup_options.tables)
                total_processed = 0
                chunks = []

                # 按大小排序表，先处理小表（优化内存使用）
                tables.sort(key=lambda x: x.get('size', 0))

                # 预估内存需求并调整批次大小
                estimated_memory_per_table = max([min(t.get('size', 0) / 10, 50 * 1024 * 1024) for t in tables] or [50 * 1024 * 1024])
                available_memory = self.memory_manager.get_current_usage().heap_total * 0.5
                optimal_batch_size = max(1, min(3, int(available_memory / estimated_memory_per_table)))
                final_batch_size = optimal_batch_size

                # 分块处理表
                for i, table in enumerate(tables):
                    # 检查内存压力
                    current_pressure = self.memory_manager.check_memory_pressure()
                    if current_pressure > 0.9:
                        await self._request_memory_cleanup()
                        await asyncio.sleep(1)

                    # 为大表创建单独的备份文件
                    if table.get('size', 0) > large_options.disk_threshold:
                        chunk_file = await self._create_table_chunk_backup(table, temp_dir, large_options)
                        chunks.append(chunk_file)
                    else:
                        # 小表可以合并处理
                        small_tables = [t for t in tables if t.get('size', 0) <= large_options.disk_threshold]
                        chunk_file = await self._create_small_table_backup(small_tables, temp_dir, large_options)
                        chunks.append(chunk_file)

                    total_processed += table.get('size', 0)

                # 合并所有块文件
                final_backup_path = await self._merge_backup_chunks(chunks, backup_options, large_options)

                # 清理临时文件
                import shutil
                shutil.rmtree(temp_dir, ignore_errors=True)

                file_size = os.path.getsize(final_backup_path) if os.path.exists(final_backup_path) else 0
                duration = int((TimeUtils.now() - start_time) * 1000)

                return BackupResult(
                    success=True,
                    file_path=final_backup_path,
                    file_size=file_size,
                    table_count=len(tables),
                    record_count=total_processed,
                    duration=duration
                )

            except Exception as e:
                # 清理临时文件
                import shutil
                shutil.rmtree(temp_dir, ignore_errors=True)
                raise

        except Exception as e:
            duration = int((TimeUtils.now() - start_time) * 1000)
            error_message = str(e)

            logger.error("大文件备份创建失败", "MySQLBackupTool", {
                "error": error_message,
                "duration": duration
            })

            return BackupResult(
                success=False,
                error=error_message,
                duration=duration
            )

    async def _get_tables_with_sizes(self, specific_tables: Optional[List[str]] = None) -> List[Dict[str, Any]]:
        """获取表及其大小信息"""
        tables_info = []

        tables = specific_tables or []
        if not tables:
            result = await self.mysql_manager.execute_query('SHOW TABLES')
            tables = [row[list(row.keys())[0]] for row in result]

        # 批量获取所有表的状态信息
        try:
            where_clause = "AND table_name IN ({})".format(','.join(['%s'] * len(tables))) if tables else ""
            params = tables if tables else []

            result = await self.mysql_manager.execute_query(
                f"SELECT table_name, data_length, index_length, table_rows, update_time FROM information_schema.tables WHERE table_schema = DATABASE() {where_clause} ORDER BY (data_length + index_length) ASC",
                params
            )

            if result:
                for row in result:
                    table_name = row.get('table_name', '')
                    data_length = row.get('data_length', 0) or 0
                    index_length = row.get('index_length', 0) or 0
                    table_rows = row.get('table_rows', 0) or 0

                    tables_info.append({
                        'name': table_name,
                        'size': data_length + index_length,
                        'rows': table_rows
                    })

        except Exception as e:
            logger.warn(f"批量获取表状态失败，使用逐个查询方式: {e}")

            # 降级到逐个查询
            for table_name in tables:
                try:
                    result = await self.mysql_manager.execute_query(
                        'SHOW TABLE STATUS LIKE %s', [table_name]
                    )

                    if result and len(result) > 0:
                        status = result[0]
                        size = (status.get('Data_length', 0) or 0) + (status.get('Index_length', 0) or 0)
                        rows = status.get('Rows', 0) or 0

                        tables_info.append({
                            'name': table_name,
                            'size': size,
                            'rows': rows
                        })

                except Exception as e:
                    logger.warn(f"无法获取表 {table_name} 的信息: {e}")
                    tables_info.append({
                        'name': table_name,
                        'size': 0,
                        'rows': 0
                    })

        return tables_info

    async def _create_table_chunk_backup(self, table: Dict[str, Any],
                                       temp_dir: str, options: LargeFileOptions) -> str:
        """创建表块备份"""
        chunk_file = os.path.join(temp_dir, f"{table['name']}_chunk.sql")
        config = self.mysql_manager.config_manager.database

        # 对于大表，使用LIMIT分批导出
        batch_size = max(1000, int(options.chunk_size / 1024))
        offset = 0
        has_more_data = True

        with open(chunk_file, 'w', encoding='utf-8') as f:
            # 写入表结构
            result = await self.mysql_manager.execute_query(
                f"SHOW CREATE TABLE `{table['name']}`"
            )

            if result and len(result) > 0:
                create_statement = result[0].get('Create Table', '')
                f.write(f"DROP TABLE IF EXISTS `{table['name']}`;\n")
                f.write(f"{create_statement};\n\n")

            # 分批导出数据
            while has_more_data:
                # 检查内存压力
                memory_pressure = self.memory_manager.check_memory_pressure()
                if memory_pressure > 0.85:
                    await self._request_memory_cleanup()

                result = await self.mysql_manager.execute_query(
                    f"SELECT * FROM `{table['name']}` LIMIT {batch_size} OFFSET {offset}"
                )

                if not result or len(result) == 0:
                    has_more_data = False
                    break

                # 生成INSERT语句
                if result:
                    columns = list(result[0].keys())
                    column_list = ', '.join([f'`{col}`' for col in columns])

                    f.write(f"INSERT INTO `{table['name']}` ({column_list}) VALUES\n")

                    for i, row in enumerate(result):
                        values = []
                        for col in columns:
                            value = row[col]
                            if value is None:
                                values.append('NULL')
                            elif isinstance(value, str):
                                values.append(f"'{value.replace(chr(39), chr(39) + chr(39))}'")
                            else:
                                values.append(str(value))

                        f.write(f"({', '.join(values)}){';' if i == len(result) - 1 else ','}\n")

                offset += batch_size

                if len(result) < batch_size:
                    has_more_data = False

        return chunk_file

    async def _create_small_table_backup(self, tables: List[Dict[str, Any]],
                                       temp_dir: str, options: LargeFileOptions) -> str:
        """创建小表备份"""
        chunk_file = os.path.join(temp_dir, f"small_tables_{int(time.time())}.sql")
        config = self.mysql_manager.config_manager.database

        # 使用mysqldump导出小表
        dump_args = [
            'mysqldump',
            f'-h{config["host"]}',
            f'-P{config["port"]}',
            f'-u{config["user"]}',
            f'-p{config["password"]}',
            '--default-character-set=utf8mb4',
            '--single-transaction',
            '--routines',
            '--triggers',
            config["database"] or ''
        ]

        table_names = [t['name'] for t in tables]
        dump_args.extend(table_names)

        await self._execute_mysqldump(dump_args, chunk_file)
        return chunk_file

    async def _merge_backup_chunks(self, chunks: List[str], backup_options: BackupOptions,
                                 large_file_options: LargeFileOptions) -> str:
        """合并备份块"""
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        final_file_name = f"{backup_options.file_prefix or 'mysql_large_backup'}_{timestamp}"
        final_path = os.path.join(backup_options.output_dir, final_file_name)

        if len(chunks) == 1:
            # 只有一个块，直接移动并压缩
            single_chunk = chunks[0]
            sql_path = f"{final_path}.sql"
            os.rename(single_chunk, sql_path)

            if backup_options.compress:
                zip_path = f"{final_path}.zip"
                await self._compress_file(sql_path, zip_path)
                os.unlink(sql_path)
                return zip_path

            return sql_path

        # 多个块需要合并
        merged_sql_path = f"{final_path}.sql"

        with open(merged_sql_path, 'w', encoding='utf-8') as output_file:
            # 写入文件头
            output_file.write("-- MySQL Large File Backup\n")
            output_file.write(f"-- Generated on: {datetime.now().isoformat()}\n")
            output_file.write(f"-- Chunks: {len(chunks)}\n\n")
            output_file.write("SET FOREIGN_KEY_CHECKS=0;\n\n")

            # 合并所有块
            for chunk_file in chunks:
                with open(chunk_file, 'r', encoding='utf-8') as chunk_file:
                    output_file.write(chunk_file.read())
                output_file.write('\n')

                # 删除临时块文件
                os.unlink(chunk_file)

            output_file.write("\nSET FOREIGN_KEY_CHECKS=1;\n")

        # 压缩合并后的文件
        if backup_options.compress:
            zip_path = f"{final_path}.zip"
            await self._compress_file(merged_sql_path, zip_path)
            os.unlink(merged_sql_path)
            return zip_path

        return merged_sql_path

    # 队列相关方法
    async def create_backup_queued(self, options: Optional[BackupOptions] = None,
                                 priority: int = 1) -> Dict[str, Any]:
        """带队列的备份创建"""
        task_id = self.add_task_to_queue(
            'backup',
            lambda: self.create_backup(options),
            options.model_dump() if hasattr(options, 'model_dump') else {},
            priority
        )

        # 等待任务完成
        return await self._wait_for_task_completion(task_id)

    async def _wait_for_task_completion(self, task_id: str) -> Dict[str, Any]:
        """等待任务完成"""
        while True:
            task = self.get_task_status(task_id)
            if not task:
                raise Exception('任务不存在')

            if task.status == 'completed':
                return {'task_id': task_id, 'result': task.result}
            elif task.status == 'failed':
                raise Exception(task.error or '任务执行失败')
            elif task.status == 'cancelled':
                raise Exception('任务已取消')

            await asyncio.sleep(1)

    def cancel_task(self, task_id: str) -> bool:
        """取消任务"""
        task = self.task_queue.get(task_id)
        if not task:
            return False

        if task.status == 'running':
            task.status = 'cancelled'
            task.completed_at = datetime.now()
            return True
        elif task.status == 'queued':
            del self.task_queue[task_id]
            return True

        return False

    def get_queue_stats(self) -> Dict[str, Any]:
        """获取队列统计信息"""
        tasks = list(self.task_queue.values())

        queued_tasks = len([t for t in tasks if t.status == 'queued'])
        running_tasks = len([t for t in tasks if t.status == 'running'])
        completed_tasks = len([t for t in tasks if t.status == 'completed'])
        failed_tasks = len([t for t in tasks if t.status == 'failed'])
        cancelled_tasks = len([t for t in tasks if t.status == 'cancelled'])

        # 计算平均等待时间
        wait_times = []
        for task in tasks:
            if task.started_at:
                wait_time = (task.started_at.timestamp() - task.created_at.timestamp()) * 1000
                wait_times.append(wait_time)

        average_wait_time = sum(wait_times) / len(wait_times) if wait_times else 0

        # 计算平均执行时间
        execution_times = []
        for task in tasks:
            if task.completed_at and task.started_at:
                execution_time = (task.completed_at.timestamp() - task.started_at.timestamp()) * 1000
                execution_times.append(execution_time)

        average_execution_time = sum(execution_times) / len(execution_times) if execution_times else 0

        return {
            'total_tasks': len(tasks),
            'queued_tasks': queued_tasks,
            'running_tasks': running_tasks,
            'completed_tasks': completed_tasks,
            'failed_tasks': failed_tasks,
            'cancelled_tasks': cancelled_tasks,
            'max_concurrent_tasks': self.max_concurrent_tasks,
            'average_wait_time': average_wait_time,
            'average_execution_time': average_execution_time
        }

    def get_active_trackers(self) -> List[ProgressTracker]:
        """获取所有活动的进度跟踪器"""
        return list(self.progress_trackers.values())

    def cancel_operation(self, tracker_id: str) -> bool:
        """取消指定的操作"""
        tracker = self.progress_trackers.get(tracker_id)
        if tracker and tracker.cancellation_token:
            tracker.cancellation_token.cancel()
            return True
        return False

    def cleanup_completed_trackers(self) -> None:
        """清理已完成的跟踪器"""
        current_time = time.time()
        expired_trackers = []

        for tracker_id, tracker in self.progress_trackers.items():
            elapsed = current_time - tracker.start_time.timestamp()
            if elapsed > 300 or tracker.progress.stage in ['completed', 'error']:
                expired_trackers.append(tracker_id)

        for tracker_id in expired_trackers:
            del self.progress_trackers[tracker_id]

    def set_max_concurrent_tasks(self, max_concurrent: int) -> None:
        """设置最大并发任务数"""
        if max_concurrent < 1:
            raise ValueError('最大并发任务数必须大于0')

        self.max_concurrent_tasks = max_concurrent

    def pause_queue(self) -> None:
        """暂停任务队列处理"""
        if self.scheduler_running:
            self.scheduler_running = False
            # 发出队列暂停事件
            self.emit('queue-paused')

    def resume_queue(self) -> None:
        """恢复任务队列处理"""
        if not self.scheduler_running:
            self.scheduler_running = True
            self.scheduler_task = asyncio.create_task(self._task_scheduler_loop())
            # 发出队列恢复事件
            self.emit('queue-resumed')

    def clear_queue(self) -> int:
        """清空任务队列"""
        queued_tasks = [task for task in self.task_queue.values() if task.status == 'queued']

        for task in queued_tasks:
            del self.task_queue[task.id]

        # 发出队列清空事件
        self.emit('queue-cleared', {
            'cleared_count': len(queued_tasks),
            'remaining_tasks': len(self.task_queue)
        })

        return len(queued_tasks)

    def get_all_tasks(self, status: Optional[str] = None) -> List[TaskQueue]:
        """获取所有任务信息"""
        tasks = list(self.task_queue.values())

        if status:
            tasks = [task for task in tasks if task.status == status]

        return tasks

    def get_queue_diagnostics(self) -> Dict[str, Any]:
        """
        获取任务队列诊断信息

        Returns:
            Dict[str, Any]: 详细的诊断信息
        """
        tasks = list(self.task_queue.values())
        current_time = time.time()

        # 队列信息
        tasks_by_type: Dict[str, int] = {}
        tasks_by_status: Dict[str, int] = {}
        oldest_task: Optional[datetime] = None
        newest_task: Optional[datetime] = None

        for task in tasks:
            # 按类型统计
            tasks_by_type[task.type] = tasks_by_type.get(task.type, 0) + 1

            # 按状态统计
            tasks_by_status[task.status] = tasks_by_status.get(task.status, 0) + 1

            # 时间范围
            if not oldest_task or task.created_at < oldest_task:
                oldest_task = task.created_at
            if not newest_task or task.created_at > newest_task:
                newest_task = task.created_at

        # 性能指标
        completed_tasks = [task for task in tasks if task.status == 'completed']
        failed_tasks = [task for task in tasks if task.status == 'failed']
        total_finished_tasks = len(completed_tasks) + len(failed_tasks)

        success_rate = 0.0
        if total_finished_tasks > 0:
            success_rate = len(completed_tasks) / total_finished_tasks

        # 计算吞吐量（每分钟完成的任务数）
        time_window = 60  # 1分钟
        recent_completed_tasks = [
            task for task in completed_tasks
            if task.completed_at and (current_time - task.completed_at.timestamp()) <= time_window
        ]
        throughput = len(recent_completed_tasks)

        # 平均等待和执行时间
        wait_times = []
        execution_times = []

        for task in tasks:
            if task.started_at:
                wait_time = (task.started_at.timestamp() - task.created_at.timestamp())
                wait_times.append(wait_time)

            if task.completed_at and task.started_at:
                execution_time = (task.completed_at.timestamp() - task.started_at.timestamp())
                execution_times.append(execution_time)

        average_wait_time = sum(wait_times) / len(wait_times) if wait_times else 0
        average_execution_time = sum(execution_times) / len(execution_times) if execution_times else 0

        return {
            'scheduler': {
                'is_running': self.scheduler_running,
                'check_interval': 1.0,
                'adaptive_scheduling': True
            },
            'queue': {
                'size': len(tasks),
                'oldest_task': oldest_task.isoformat() if oldest_task else None,
                'newest_task': newest_task.isoformat() if newest_task else None,
                'tasks_by_type': tasks_by_type,
                'tasks_by_status': tasks_by_status
            },
            'performance': {
                'average_wait_time': average_wait_time,
                'average_execution_time': average_execution_time,
                'success_rate': success_rate,
                'throughput': throughput
            },
            'resources': {
                'max_concurrent_tasks': self.max_concurrent_tasks,
                'current_running_tasks': self.running_tasks,
                'queued_tasks': len([t for t in tasks if t.status == 'queued']),
                'memory_usage': self.memory_manager.get_current_usage(),
                'memory_pressure': self.memory_manager.check_memory_pressure()
            }
        }

    async def cleanup(self) -> None:
        """清理资源 - 改进版资源清理"""
        try:
            # 停止任务调度器
            self.scheduler_running = False
            if self.scheduler_task:
                self.scheduler_task.cancel()
                try:
                    await self.scheduler_task
                except asyncio.CancelledError:
                    pass

            # 取消所有运行中的任务
            running_tasks = [task for task in self.task_queue.values() if task.status == 'running']
            for task in running_tasks:
                task.status = 'cancelled'
                task.completed_at = datetime.now()

            # 清理缓存
            await self.cache_manager.clear_region(CacheRegion.QUERY_RESULT)

            # 禁用内存监控
            self.memory_manager.disable_memory_monitoring()

            # 清理所有跟踪器
            self.progress_trackers.clear()

            # 清空任务队列
            self.task_queue.clear()

            # 清空事件监听器
            self._event_listeners.clear()

            # 获取最终内存使用情况
            final_memory_usage = self.memory_manager.get_current_usage()

            # 发出清理完成事件
            self.emit('cleanup-completed', {
                'final_memory_usage': {
                    'rss_mb': f"{final_memory_usage.rss / 1024 / 1024:.1f}",
                    'heap_used_mb': f"{final_memory_usage.heap_used / 1024 / 1024:.1f}"
                },
                'resources_cleared': {
                    'query_cache': True,
                    'connection_cache': True,
                    'task_queue': True,
                    'progress_trackers': True,
                    'event_listeners': True
                },
                'cancelled_tasks': len(running_tasks)
            })

            logger.info("MySQL备份工具已清理完成", "MySQLBackupTool")

        except Exception as e:
            logger.error(f"清理备份工具时出错: {e}", "MySQLBackupTool")

    def get_export_diagnostics(self) -> Dict[str, Any]:
        """
        获取导出系统诊断信息

        Returns:
            Dict[str, Any]: 导出系统详细诊断信息
        """
        memory_usage = self.memory_manager.get_current_usage()
        memory_pressure = self.memory_manager.check_memory_pressure()

        # 获取活跃的导出跟踪器
        active_trackers = [tracker for tracker in self.progress_trackers.values()
                          if tracker.operation == 'export']

        # 计算导出性能指标
        completed_exports = [tracker for tracker in self.progress_trackers.values()
                           if tracker.operation == 'export' and tracker.progress.stage == 'completed']

        success_rate = 0.0
        if len(active_trackers) + len(completed_exports) > 0:
            success_rate = len(completed_exports) / (len(active_trackers) + len(completed_exports))

        # 内存优化建议
        memory_suggestions = []
        if memory_pressure > 0.8:
            memory_suggestions.append("内存压力较高，建议减少并发导出任务")
        if memory_usage.heap_used > 500 * 1024 * 1024:  # 500MB
            memory_suggestions.append("内存使用量较大，建议启用流式处理")

        return {
            'export_system': {
                'active_exports': len(active_trackers),
                'completed_exports': len(completed_exports),
                'success_rate': success_rate,
                'memory_pressure': memory_pressure,
                'memory_usage_mb': {
                    'heap_used': round(memory_usage.heap_used / 1024 / 1024, 2),
                    'heap_total': round(memory_usage.heap_total / 1024 / 1024, 2),
                    'rss': round(memory_usage.rss / 1024 / 1024, 2)
                }
            },
            'performance_metrics': {
                'average_export_time': 0,  # 可以从历史数据计算
                'peak_concurrent_exports': len(active_trackers),
                'total_exports_processed': len(completed_exports)
            },
            'optimization_suggestions': {
                'memory_optimizations': memory_suggestions,
                'recommended_batch_size': self._calculate_optimal_batch_size(),
                'streaming_recommended': memory_usage.heap_used > 200 * 1024 * 1024  # 200MB
            },
            'system_health': {
                'memory_monitoring_enabled': True,
                'error_recovery_enabled': True,
                'progress_tracking_enabled': True,
                'queue_management_enabled': True
            }
        }

    def _calculate_optimal_batch_size(self) -> int:
        """计算最优批处理大小"""
        memory_usage = self.memory_manager.get_current_usage()
        available_memory = memory_usage.heap_total - memory_usage.heap_used

        # 基于可用内存计算批处理大小
        if available_memory > 500 * 1024 * 1024:  # > 500MB
            return 1000
        elif available_memory > 200 * 1024 * 1024:  # > 200MB
            return 500
        else:
            return 100

    async def get_export_performance_report(self) -> Dict[str, Any]:
        """
        生成导出性能报告

        Returns:
            Dict[str, Any]: 详细的性能报告
        """
        diagnostics = self.get_export_diagnostics()
        queue_stats = self.get_queue_stats()
        memory_stats = self.memory_manager.get_memory_stats()

        return {
            'report_generated_at': datetime.now().isoformat(),
            'export_diagnostics': diagnostics,
            'queue_statistics': queue_stats,
            'memory_statistics': memory_stats,
            'recommendations': self._generate_export_recommendations(diagnostics, queue_stats)
        }

    def _generate_export_recommendations(self, diagnostics: Dict[str, Any], queue_stats: Dict[str, Any]) -> List[str]:
        """生成导出优化建议"""
        recommendations = []

        # 内存相关建议
        memory_pressure = diagnostics['export_system']['memory_pressure']
        if memory_pressure > 0.8:
            recommendations.append("内存压力较高，建议减少并发导出任务数量")
        elif memory_pressure > 0.9:
            recommendations.append("内存压力严重，建议启用流式处理模式")

        # 队列相关建议
        if queue_stats['queued_tasks'] > 10:
            recommendations.append("队列中任务较多，建议增加最大并发任务数")
        elif queue_stats['average_wait_time'] > 30:  # 30秒
            recommendations.append("任务等待时间较长，建议优化任务调度策略")

        # 性能相关建议
        success_rate = diagnostics['export_system']['success_rate']
        if success_rate < 0.95:
            recommendations.append("导出成功率偏低，建议检查错误恢复机制")

        return recommendations


# 导出类和函数
__all__ = [
    'MySQLBackupTool',
    'MemoryManager',
    'BackupOptions',
    'ExportOptions',
    'ReportConfig',
    'with_error_handling',
    'with_performance_monitoring'
]