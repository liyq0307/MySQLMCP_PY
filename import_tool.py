"""
MySQL 数据导入工具 - 企业级数据导入解决方案

提供全面的企业级数据库数据导入功能，支持CSV、JSON、Excel、SQL等多种数据格式的高性能批量导入。
集成了智能数据验证、字段映射、事务管理、错误处理、重复检查和性能优化等完整的导入生态系统。

@fileoverview MySQL数据导入工具 - 企业级数据导入的完整解决方案
@author liyq
@version 1.0.0
@since 1.0.0
@updated 2025-09-25
@license MIT
"""

import asyncio
import csv
import io
import json
import os
import re
import time
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

import aiofiles
import openpyxl
from openpyxl import Workbook

from mysql_manager import MySQLManager
from type_utils import (
    ImportOptions,
    ImportResult,
    ValidationResult,
    FieldMapping,
    ImportProgress,
    DuplicateCheckConfig,
    DuplicateCacheItem,
    CandidateKey,
    ColumnDefinition,
    TableSchemaInfo,
    MySQLMCPError,
    ErrorCategory,
    ErrorSeverity
)
from common_utils import TimeUtils
from logger import logger
from cache import CacheManager, CacheRegion
from error_handler import safe_error as safe_error_handler


class MySQLImportTool:
    """
    MySQL 数据导入工具类

    企业级导入工具，集成了智能验证、字段映射、事务管理、
    错误恢复等高级特性。支持CSV、JSON、Excel、SQL等多种数据格式的导入。

    主要组件：
    - 数据验证器：智能数据格式和类型验证
    - 字段映射器：灵活的字段映射和转换
    - 导入引擎：不同格式数据的专用导入器
    - 事务管理器：确保数据一致性的事务控制
    - 错误处理器：详细错误诊断和分类
    """

    # 重复检查缓存
    duplicate_cache: Dict[str, DuplicateCacheItem] = {}
    # 默认缓存大小
    max_cache_size: int = 10000
    # 当前表格的候选键列表
    current_candidate_keys: List[CandidateKey] = []

    def __init__(self, mysql_manager: MySQLManager):
        """初始化导入工具"""
        self.mysql_manager = mysql_manager
        self.session_id = str(uuid.uuid4())

        # 初始化缓存管理器
        self.cache_manager = CacheManager.get_instance()

        # 注册内存跟踪
        self._register_memory_tracking()

    def _register_memory_tracking(self) -> None:
        """注册内存泄漏跟踪"""
        # 这里可以添加内存跟踪逻辑
        pass

    async def import_from_csv(self, options: ImportOptions) -> ImportResult:
        """
        从CSV文件导入数据

        支持标准CSV格式和自定义分隔符，支持表头识别和字段映射。
        提供数据验证、批量插入和错误处理等完整功能。

        Args:
            options: 导入选项配置

        Returns:
            包含导入结果的JSON格式数据

        Raises:
            MySQLMCPError: 当导入失败时抛出
        """
        start_time = TimeUtils.now_in_milliseconds()

        try:
            # 发送开始进度事件
            if options.with_progress:
                initial_progress = self._create_progress(
                    0,
                    'fileReading',
                    '开始读取CSV文件...',
                    0,
                    0,
                    start_time
                )
                await self._emit_progress(initial_progress, options.file_path, 'csv')

            # 验证输入参数
            self._validate_import_options(options, 'csv')

            # 读取CSV文件内容
            file_content = await self._read_file_content(options.file_path, options.encoding or 'utf8')

            # 发送文件读取完成进度
            if options.with_progress:
                progress = self._calculate_overall_progress(
                    {'fileReading': 30, 'dataParsing': 40, 'validation': 20, 'insertion': 10},
                    'fileReading',
                    100,
                    0,
                    0,
                    start_time,
                    f'CSV文件读取完成，大小: {(len(file_content.encode()) / 1024):.1f} KB'
                )
                await self._emit_progress(progress, options.file_path, 'csv')

            # 发送开始解析进度
            if options.with_progress:
                progress = self._calculate_overall_progress(
                    {'fileReading': 30, 'dataParsing': 40, 'validation': 20, 'insertion': 10},
                    'dataParsing',
                    0,
                    0,
                    0,
                    start_time,
                    '开始解析CSV数据...'
                )
                await self._emit_progress(progress, options.file_path, 'csv')

            # 解析CSV数据
            csv_data = await self._parse_csv(file_content, {
                'delimiter': options.delimiter or ',',
                'quote': options.quote or '"',
                'hasHeaders': options.has_headers is not False,
                'encoding': options.encoding or 'utf8'
            })

            # 发送解析完成进度
            if options.with_progress:
                progress = self._calculate_overall_progress(
                    {'fileReading': 30, 'dataParsing': 40, 'validation': 20, 'insertion': 10},
                    'dataParsing',
                    100,
                    len(csv_data),
                    len(csv_data),
                    start_time,
                    f'CSV解析完成，共解析 {len(csv_data)} 行数据'
                )
                await self._emit_progress(progress, options.file_path, 'csv')

            # 执行导入
            result = await self._execute_import(options.table_name, csv_data, options)

            return ImportResult(
                success=True,
                imported_rows=result['imported_rows'],
                skipped_rows=result['skipped_rows'],
                failed_rows=result['failed_rows'],
                updated_rows=result['updated_rows'],
                total_rows=len(csv_data),
                duration=int(TimeUtils.get_duration_in_ms(start_time)),
                batches_processed=result['batches_processed'],
                file_path=options.file_path,
                format='csv',
                table_name=options.table_name
            )

        except Exception as error:
            safe_error = self._safe_error(error, 'importFromCSV')
            return ImportResult(
                success=False,
                imported_rows=0,
                skipped_rows=0,
                failed_rows=0,
                updated_rows=0,
                total_rows=0,
                duration=int(TimeUtils.get_duration_in_ms(start_time)),
                batches_processed=0,
                file_path=options.file_path,
                format='csv',
                table_name=options.table_name,
                error=safe_error.message
            )

    async def import_from_json(self, options: ImportOptions) -> ImportResult:
        """
        从JSON文件导入数据

        支持JSON数组和对象格式，支持嵌套数据结构的扁平化处理。
        提供灵活的字段映射和数据转换功能。

        Args:
            options: 导入选项配置

        Returns:
            包含导入结果的JSON格式数据

        Raises:
            MySQLMCPError: 当导入失败时抛出
        """
        start_time = TimeUtils.now_in_milliseconds()

        try:
            # 发送开始进度事件
            if options.with_progress:
                initial_progress = self._create_progress(
                    0,
                    'fileReading',
                    '开始读取JSON文件...',
                    0,
                    0,
                    start_time
                )
                await self._emit_progress(initial_progress, options.file_path, 'json')

            # 验证输入参数
            self._validate_import_options(options, 'json')

            # 读取JSON文件内容
            file_content = await self._read_file_content(options.file_path, options.encoding or 'utf8')

            # 发送文件读取完成进度
            if options.with_progress:
                progress = self._calculate_overall_progress(
                    {'fileReading': 20, 'dataParsing': 60, 'validation': 15, 'insertion': 5},
                    'fileReading',
                    100,
                    0,
                    0,
                    start_time,
                    f'JSON文件读取完成，大小: {(len(file_content.encode()) / 1024):.1f} KB'
                )
                await self._emit_progress(progress, options.file_path, 'json')

            # 发送开始解析进度
            if options.with_progress:
                progress = self._calculate_overall_progress(
                    {'fileReading': 20, 'dataParsing': 60, 'validation': 15, 'insertion': 5},
                    'dataParsing',
                    0,
                    0,
                    0,
                    start_time,
                    '开始解析JSON数据...'
                )
                await self._emit_progress(progress, options.file_path, 'json')

            # 解析JSON数据
            json_data = json.loads(file_content)

            # 处理不同JSON格式
            processed_data: List[Dict[str, Any]] = []
            if isinstance(json_data, list):
                processed_data = json_data
            elif isinstance(json_data, dict):
                # 如果是单个对象，转换为数组
                processed_data = [json_data]
            else:
                raise MySQLMCPError(
                    'JSON文件必须包含对象数组或单个对象',
                    ErrorCategory.VALIDATION_ERROR,
                    ErrorSeverity.MEDIUM
                )

            # 扁平化嵌套数据（如果需要）
            flatten_start_time = TimeUtils.now_in_milliseconds()
            processed_data = [self._flatten_object(item) for item in processed_data]
            flatten_time = TimeUtils.get_duration_in_ms(flatten_start_time)

            # 发送解析完成进度
            if options.with_progress:
                progress = self._calculate_overall_progress(
                    {'fileReading': 20, 'dataParsing': 60, 'validation': 15, 'insertion': 5},
                    'dataParsing',
                    100,
                    len(processed_data),
                    len(processed_data),
                    start_time,
                    f'JSON解析完成，共解析 {len(processed_data)} 个对象 {f"(扁平化耗时: {flatten_time}ms)" if flatten_time > 0 else ""}'
                )
                await self._emit_progress(progress, options.file_path, 'json')

            # 执行导入
            result = await self._execute_import(options.table_name, processed_data, options)

            return ImportResult(
                success=True,
                imported_rows=result['imported_rows'],
                skipped_rows=result['skipped_rows'],
                failed_rows=result['failed_rows'],
                updated_rows=result['updated_rows'],
                total_rows=len(processed_data),
                duration=int(TimeUtils.get_duration_in_ms(start_time)),
                batches_processed=result['batches_processed'],
                file_path=options.file_path,
                format='json',
                table_name=options.table_name
            )

        except Exception as error:
            safe_error = self._safe_error(error, 'importFromJSON')
            return ImportResult(
                success=False,
                imported_rows=0,
                skipped_rows=0,
                failed_rows=0,
                updated_rows=0,
                total_rows=0,
                duration=int(TimeUtils.get_duration_in_ms(start_time)),
                batches_processed=0,
                file_path=options.file_path,
                format='json',
                table_name=options.table_name,
                error=safe_error.message
            )

    async def import_from_excel(self, options: ImportOptions) -> ImportResult:
        """
        从Excel文件导入数据

        支持多工作表Excel文件，支持样式保留和复杂数据类型处理。
        提供工作表选择和字段映射等高级功能。

        Args:
            options: 导入选项配置

        Returns:
            包含导入结果的JSON格式数据

        Raises:
            MySQLMCPError: 当导入失败时抛出
        """
        start_time = TimeUtils.now_in_milliseconds()

        try:
            # 发送开始进度事件
            if options.with_progress:
                initial_progress = self._create_progress(
                    0,
                    'fileReading',
                    '开始读取Excel文件...',
                    0,
                    0,
                    start_time
                )
                await self._emit_progress(initial_progress, options.file_path, 'excel')

            # 验证输入参数
            self._validate_import_options(options, 'excel')

            # 发送开始加载进度
            if options.with_progress:
                progress = self._calculate_overall_progress(
                    {'fileReading': 40, 'dataParsing': 50, 'validation': 7, 'insertion': 3},
                    'fileReading',
                    10,
                    0,
                    0,
                    start_time,
                    '正在加载Excel工作簿...'
                )
                await self._emit_progress(progress, options.file_path, 'excel')

            # 加载Excel文件
            workbook = openpyxl.load_workbook(options.file_path, data_only=True)

            # 发送加载完成进度
            if options.with_progress:
                progress = self._calculate_overall_progress(
                    {'fileReading': 40, 'dataParsing': 50, 'validation': 7, 'insertion': 3},
                    'fileReading',
                    100,
                    0,
                    0,
                    start_time,
                    'Excel文件加载完成，正在选择工作表...'
                )
                await self._emit_progress(progress, options.file_path, 'excel')

            # 选择工作表
            worksheet_name = options.sheet_name or workbook.sheetnames[0]
            worksheet = workbook[worksheet_name]

            if not worksheet:
                raise MySQLMCPError(
                    f'工作表 \'{worksheet_name}\' 不存在',
                    ErrorCategory.VALIDATION_ERROR,
                    ErrorSeverity.MEDIUM
                )

            # 发送开始解析进度
            if options.with_progress:
                progress = self._calculate_overall_progress(
                    {'fileReading': 40, 'dataParsing': 50, 'validation': 7, 'insertion': 3},
                    'dataParsing',
                    0,
                    0,
                    0,
                    start_time,
                    '开始解析Excel数据...'
                )
                await self._emit_progress(progress, options.file_path, 'excel')

            # 解析Excel数据
            excel_data: List[Dict[str, Any]] = []
            rows = list(worksheet.iter_rows(values_only=True))

            if len(rows) == 0:
                raise MySQLMCPError(
                    'Excel文件为空或不包含数据',
                    ErrorCategory.VALIDATION_ERROR,
                    ErrorSeverity.MEDIUM
                )

            # 处理表头
            has_headers = options.has_headers is not False
            headers: List[str] = []

            if has_headers:
                # 第一行作为表头
                header_row = rows[0]
                headers = [str(cell) if cell is not None else f"column_{i+1}" for i, cell in enumerate(header_row)]
            else:
                # 生成列名
                first_data_row = rows[0]
                headers = [f"column_{i+1}" for i in range(len(first_data_row))]

            # 处理数据行
            start_row_index = 1 if has_headers else 0
            for row_idx in range(start_row_index, len(rows)):
                row = rows[row_idx]
                if not row:
                    continue

                row_data: Dict[str, Any] = {}
                for col_idx, cell_value in enumerate(row):
                    if col_idx < len(headers):
                        row_data[headers[col_idx]] = cell_value

                excel_data.append(row_data)

            # 执行导入
            result = await self._execute_import(options.table_name, excel_data, options)

            return ImportResult(
                success=True,
                imported_rows=result['imported_rows'],
                skipped_rows=result['skipped_rows'],
                failed_rows=result['failed_rows'],
                updated_rows=result['updated_rows'],
                total_rows=len(excel_data),
                duration=int(TimeUtils.get_duration_in_ms(start_time)),
                batches_processed=result['batches_processed'],
                file_path=options.file_path,
                format='excel',
                table_name=options.table_name
            )

        except Exception as error:
            safe_error = self._safe_error(error, 'importFromExcel')
            return ImportResult(
                success=False,
                imported_rows=0,
                skipped_rows=0,
                failed_rows=0,
                updated_rows=0,
                total_rows=0,
                duration=int(TimeUtils.get_duration_in_ms(start_time)),
                batches_processed=0,
                file_path=options.file_path,
                format='excel',
                table_name=options.table_name,
                error=safe_error.message
            )

    async def import_from_sql(self, options: ImportOptions) -> ImportResult:
        """
        从SQL文件导入数据

        支持SQL脚本文件执行，支持事务管理和错误处理。
        提供SQL语句解析和批量执行功能。

        Args:
            options: 导入选项配置

        Returns:
            包含导入结果的JSON格式数据

        Raises:
            MySQLMCPError: 当导入失败时抛出
        """
        start_time = TimeUtils.now_in_milliseconds()

        try:
            # 发送开始进度事件
            if options.with_progress:
                initial_progress = self._create_progress(
                    0,
                    'fileReading',
                    '开始读取SQL文件...',
                    0,
                    0,
                    start_time
                )
                await self._emit_progress(initial_progress, options.file_path, 'sql')

            # 验证输入参数
            self._validate_import_options(options, 'sql')

            # 读取SQL文件内容
            file_content = await self._read_file_content(options.file_path, options.encoding or 'utf8')

            # 发送文件读取完成进度
            if options.with_progress:
                progress = self._calculate_overall_progress(
                    {'fileReading': 20, 'dataParsing': 70, 'validation': 5, 'insertion': 5},
                    'fileReading',
                    100,
                    0,
                    0,
                    start_time,
                    f'SQL文件读取完成，大小: {(len(file_content.encode()) / 1024):.1f} KB'
                )
                await self._emit_progress(progress, options.file_path, 'sql')

            # 发送开始解析进度
            if options.with_progress:
                progress = self._calculate_overall_progress(
                    {'fileReading': 20, 'dataParsing': 70, 'validation': 5, 'insertion': 5},
                    'dataParsing',
                    0,
                    0,
                    0,
                    start_time,
                    '开始解析SQL语句...'
                )
                await self._emit_progress(progress, options.file_path, 'sql')

            # 解析SQL语句
            parsing_start_time = TimeUtils.now_in_milliseconds()
            sql_statements = self._parse_sql_statements(file_content)
            parsing_time = TimeUtils.get_duration_in_ms(parsing_start_time)

            # 发送解析完成进度
            if options.with_progress:
                progress = self._calculate_overall_progress(
                    {'fileReading': 20, 'dataParsing': 70, 'validation': 5, 'insertion': 5},
                    'dataParsing',
                    100,
                    len(sql_statements),
                    len(sql_statements),
                    start_time,
                    f'SQL解析完成，共解析 {len(sql_statements)} 条语句 (耗时: {parsing_time}ms)'
                )
                await self._emit_progress(progress, options.file_path, 'sql')

            # 执行SQL语句
            result = await self._execute_sql_statements(sql_statements, options)

            return ImportResult(
                success=True,
                imported_rows=result['affected_rows'],
                skipped_rows=0,
                failed_rows=result['failed_statements'],
                updated_rows=0,
                total_rows=len(sql_statements),
                duration=int(TimeUtils.get_duration_in_ms(start_time)),
                batches_processed=1,
                file_path=options.file_path,
                format='sql',
                table_name=options.table_name or 'multiple'
            )

        except Exception as error:
            safe_error = self._safe_error(error, 'importFromSQL')
            return ImportResult(
                success=False,
                imported_rows=0,
                skipped_rows=0,
                failed_rows=0,
                updated_rows=0,
                total_rows=0,
                duration=int(TimeUtils.get_duration_in_ms(start_time)),
                batches_processed=0,
                file_path=options.file_path,
                format='sql',
                table_name=options.table_name or 'multiple',
                error=safe_error.message
            )

    async def import_data(self, options: ImportOptions) -> ImportResult:
        """
        通用导入方法 - 根据文件格式自动选择导入方式

        Args:
            options: 导入选项配置

        Returns:
            导入结果
        """
        file_format = options.format or self._detect_file_format(options.file_path)

        if file_format == 'csv':
            return await self.import_from_csv(options)
        elif file_format == 'json':
            return await self.import_from_json(options)
        elif file_format == 'excel':
            return await self.import_from_excel(options)
        elif file_format == 'sql':
            return await self.import_from_sql(options)
        else:
            raise MySQLMCPError(
                f'不支持的文件格式: {file_format}',
                ErrorCategory.VALIDATION_ERROR,
                ErrorSeverity.MEDIUM
            )

    async def validate_import(self, options: ImportOptions) -> ValidationResult:
        """
        验证数据并生成预览

        Args:
            options: 导入选项配置

        Returns:
            验证结果
        """
        try:
            # 检测文件格式
            file_format = options.format or self._detect_file_format(options.file_path)

            # 读取文件样本
            sample_data = await self._read_sample_data(options.file_path, file_format, 10)

            # 验证表结构
            table_schema = await self.mysql_manager.get_table_schema_cached(options.table_name)

            # 验证字段映射
            field_mapping = self._create_field_mapping(
                options.field_mapping or {},
                sample_data,
                table_schema
            )

            # 验证数据类型和约束
            validation_errors: List[str] = []
            validation_warnings: List[str] = []

            for row in sample_data:
                validation = self._validate_data_row(row, field_mapping, table_schema)
                validation_errors.extend(validation['errors'])
                validation_warnings.extend(validation['warnings'])

            return ValidationResult(
                is_valid=len(validation_errors) == 0,
                errors=list(set(validation_errors)),
                warnings=list(set(validation_warnings)),
                suggestions=self._generate_validation_suggestions(validation_errors, validation_warnings),
                validated_rows=len(sample_data),
                invalid_rows=len(validation_errors)
            )

        except Exception as error:
            return ValidationResult(
                is_valid=False,
                errors=[str(error)],
                warnings=[],
                suggestions=['检查文件格式和表结构是否正确'],
                validated_rows=0,
                invalid_rows=1
            )

    # 辅助方法

    def _validate_import_options(self, options: ImportOptions, file_format: str) -> None:
        """验证导入选项"""
        if not options.table_name:
            raise MySQLMCPError(
                '必须指定目标表名',
                ErrorCategory.VALIDATION_ERROR,
                ErrorSeverity.MEDIUM
            )

        if not options.file_path:
            raise MySQLMCPError(
                '必须指定文件路径',
                ErrorCategory.VALIDATION_ERROR,
                ErrorSeverity.MEDIUM
            )

        # 格式特定验证
        if file_format == 'csv':
            if options.delimiter and len(options.delimiter) != 1:
                raise MySQLMCPError(
                    'CSV分隔符必须是单个字符',
                    ErrorCategory.VALIDATION_ERROR,
                    ErrorSeverity.MEDIUM
                )
        elif file_format == 'excel':
            if options.sheet_name and not isinstance(options.sheet_name, str):
                raise MySQLMCPError(
                    'Excel工作表名称必须是字符串',
                    ErrorCategory.VALIDATION_ERROR,
                    ErrorSeverity.MEDIUM
                )

    async def _parse_csv(
        self,
        content: str,
        options: Dict[str, Any]
    ) -> List[Dict[str, Any]]:
        """解析CSV数据"""
        result: List[Dict[str, Any]] = []

        # 使用StringIO处理内容
        csv_reader = csv.reader(
            io.StringIO(content),
            delimiter=options['delimiter'],
            quotechar=options['quote']
        )

        rows = list(csv_reader)
        if len(rows) == 0:
            return result

        # 解析表头
        headers: List[str] = []
        if options['hasHeaders']:
            headers = [str(cell) for cell in rows[0]]

        # 解析数据行
        start_index = 1 if options['hasHeaders'] else 0
        for row in rows[start_index:]:
            if not row:
                continue

            row_data: Dict[str, Any] = {}
            if options['hasHeaders'] and headers:
                # 使用表头作为键
                for i, value in enumerate(row):
                    if i < len(headers):
                        row_data[headers[i]] = value
            else:
                # 生成列名
                for i, value in enumerate(row):
                    row_data[f'column_{i + 1}'] = value

            result.append(row_data)

        return result

    def _flatten_object(
        self,
        obj: Any,
        prefix: str = '',
        result: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """扁平化对象"""
        if result is None:
            result = {}

        if not obj or not isinstance(obj, dict):
            return result

        for key, value in obj.items():
            new_key = f"{prefix}_{key}" if prefix else key

            if isinstance(value, dict) and value:
                self._flatten_object(value, new_key, result)
            else:
                result[new_key] = value

        return result

    async def _read_file_content(self, file_path: str, encoding: str = 'utf8') -> str:
        """读取文件内容"""
        async with aiofiles.open(file_path, 'r', encoding=encoding) as file:
            return await file.read()

    def _parse_sql_statements(self, content: str) -> List[str]:
        """解析SQL语句"""
        statements: List[str] = []
        current_statement = ''
        in_quotes = False
        quote_char = ''

        for i, char in enumerate(content):
            next_char = content[i + 1] if i + 1 < len(content) else ''

            if not in_quotes and (char in ('"', "'")):
                in_quotes = True
                quote_char = char
            elif in_quotes and char == quote_char and content[i - 1] != '\\':
                in_quotes = False
            elif not in_quotes and char == ';' and next_char and not next_char.isspace():
                current_statement += char
                statements.append(current_statement.strip())
                current_statement = ''
                continue

            current_statement += char

        # 添加最后一个语句（如果没有分号）
        if current_statement.strip():
            statements.append(current_statement.strip())

        return [stmt for stmt in statements if stmt]

    async def _execute_sql_statements(
        self,
        statements: List[str],
        options: ImportOptions
    ) -> Dict[str, Any]:
        """执行SQL语句"""
        total_affected_rows = 0
        failed_statements = 0

        if options.use_transaction:
            # 使用事务批量执行
            queries = [{'sql': sql, 'params': []} for sql in statements]
            results = await self.mysql_manager.execute_batch_queries(queries)
            total_affected_rows = sum(
                r.get('affectedRows', 0) if isinstance(r, dict) else 0
                for r in results
            )
        else:
            # 逐条执行
            for statement in statements:
                try:
                    result = await self.mysql_manager.execute_query(statement)
                    if isinstance(result, dict) and 'affectedRows' in result:
                        total_affected_rows += result['affectedRows']
                except Exception:
                    failed_statements += 1

        return {
            'affected_rows': total_affected_rows,
            'failed_statements': failed_statements
        }

    def _detect_file_format(self, file_path: str) -> str:
        """检测文件格式"""
        ext = Path(file_path).suffix.lower()

        if ext == '.csv':
            return 'csv'
        elif ext == '.json':
            return 'json'
        elif ext in ('.xlsx', '.xls'):
            return 'excel'
        elif ext == '.sql':
            return 'sql'
        else:
            raise MySQLMCPError(
                f'无法检测文件格式: {ext}',
                ErrorCategory.VALIDATION_ERROR,
                ErrorSeverity.MEDIUM
            )

    async def _read_sample_data(
        self,
        file_path: str,
        file_format: str,
        sample_size: int
    ) -> List[Dict[str, Any]]:
        """读取样本数据"""
        if file_format == 'csv':
            content = await self._read_file_content(file_path)
            csv_data = await self._parse_csv(content, {
                'delimiter': ',',
                'quote': '"',
                'hasHeaders': True,
                'encoding': 'utf8'
            })
            return csv_data[:sample_size]

        elif file_format == 'json':
            content = await self._read_file_content(file_path)
            json_data = json.loads(content)
            if isinstance(json_data, list):
                return json_data[:sample_size]
            else:
                return [json_data]

        elif file_format == 'excel':
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            worksheet = workbook.active
            sample_rows: List[Dict[str, Any]] = []

            for row in worksheet.iter_rows(max_row=sample_size, values_only=True):
                row_data: Dict[str, Any] = {}
                for i, cell_value in enumerate(row):
                    row_data[f'column_{i+1}'] = cell_value
                sample_rows.append(row_data)

            return sample_rows

        return []

    def _create_field_mapping(
        self,
        user_mapping: Dict[str, str],
        sample_data: List[Dict[str, Any]],
        table_schema: Optional[List[Dict[str, Any]]]
    ) -> List[FieldMapping]:
        """创建字段映射"""
        mappings: List[FieldMapping] = []
        table_columns = table_schema or []

        # 如果提供了用户映射，使用用户映射
        if user_mapping:
            for source_field, target_field in user_mapping.items():
                column = next((col for col in table_columns if col.get('COLUMN_NAME') == target_field), None)
                if column:
                    mappings.append(FieldMapping(
                        source_field=source_field,
                        target_field=target_field,
                        type_conversion='auto',
                        required=column.get('IS_NULLABLE') == 'NO'
                    ))
        else:
            # 自动映射：使用相同的字段名
            if sample_data:
                source_fields = list(sample_data[0].keys())
                for source_field in source_fields:
                    column = next((col for col in table_columns if col.get('COLUMN_NAME') == source_field), None)
                    if column:
                        mappings.append(FieldMapping(
                            source_field=source_field,
                            target_field=source_field,
                            type_conversion='auto',
                            required=column.get('IS_NULLABLE') == 'NO'
                        ))

        return mappings

    def _validate_data_row(
        self,
        row: Dict[str, Any],
        field_mapping: List[FieldMapping],
        table_schema: Optional[List[Dict[str, Any]]]
    ) -> Dict[str, List[str]]:
        """验证数据行"""
        errors: List[str] = []
        warnings: List[str] = []
        table_columns = table_schema or []

        for mapping in field_mapping:
            source_value = row.get(mapping.source_field)
            column = next((col for col in table_columns if col.get('COLUMN_NAME') == mapping.target_field), None)

            if not column:
                errors.append(f'目标字段不存在: {mapping.target_field}')
                continue

            # 检查必填字段
            if mapping.required and (source_value is None or source_value == ''):
                errors.append(f'必填字段为空: {mapping.target_field}')

            # 类型验证
            if source_value is not None:
                type_error = self._validate_data_type(source_value, column.get('DATA_TYPE', ''), mapping.target_field)
                if type_error:
                    errors.append(type_error)

            # 长度验证
            if isinstance(source_value, str) and 'varchar' in column.get('DATA_TYPE', '').lower():
                max_length_match = re.search(r'varchar\((\d+)\)', column.get('DATA_TYPE', ''))
                if max_length_match:
                    max_length = int(max_length_match.group(1))
                    if len(source_value) > max_length:
                        errors.append(f'字段长度超过限制: {mapping.target_field} ({len(source_value)} > {max_length})')

        return {'errors': errors, 'warnings': warnings}

    def _validate_data_type(self, value: Any, column_type: str, field_name: str) -> Optional[str]:
        """验证数据类型"""
        value_type = type(value).__name__

        if 'int' in column_type.lower() and value_type != 'int' and not str(value).isdigit():
            return f'字段类型不匹配: {field_name} 期望数字，得到 {value_type}'

        if 'varchar' in column_type.lower() and value_type != 'str':
            return f'字段类型不匹配: {field_name} 期望字符串，得到 {value_type}'

        if 'decimal' in column_type.lower() and value_type != 'float' and value_type != 'int':
            return f'字段类型不匹配: {field_name} 期望数字，得到 {value_type}'

        return None

    def _generate_validation_suggestions(self, errors: List[str], warnings: List[str]) -> List[str]:
        """生成验证建议"""
        suggestions: List[str] = []

        if any('类型不匹配' in e for e in errors):
            suggestions.append('检查数据类型是否与表结构匹配')

        if any('必填字段为空' in e for e in errors):
            suggestions.append('确保所有必填字段都有值')

        if any('长度超过限制' in e for e in errors):
            suggestions.append('调整字段长度或截断数据')

        if warnings:
            suggestions.append('考虑清理或转换警告的数据')

        return suggestions

    def _create_progress(
        self,
        progress: float,
        stage: str,
        message: str,
        processed_rows: int,
        total_rows: int,
        start_time: float
    ) -> ImportProgress:
        """创建进度信息对象"""
        current_time = TimeUtils.now_in_milliseconds()
        elapsed = current_time - start_time

        estimated_time_remaining = None
        current_speed = None

        if progress > 0 and progress < 100:
            remaining_progress = 100 - progress
            time_per_percent = elapsed / progress
            estimated_time_remaining = time_per_percent * remaining_progress
            current_speed = (processed_rows / elapsed) * 1000 if elapsed > 0 else None  # 行/秒

        return ImportProgress(
            progress=progress,
            stage=stage,
            message=message,
            processed_rows=processed_rows,
            total_rows=total_rows,
            start_time=datetime.fromtimestamp(start_time / 1000),
            estimated_time_remaining=int(estimated_time_remaining) if estimated_time_remaining else None,
            current_speed=current_speed
        )

    async def _emit_progress(self, progress: ImportProgress, file_path: str, file_format: str) -> None:
        """发送进度事件"""
        # 这里可以实现事件发射逻辑
        logger.info('import_progress', {
            'progress': progress.progress,
            'stage': progress.stage,
            'message': progress.message,
            'processed_rows': progress.processed_rows,
            'total_rows': progress.total_rows,
            'file_path': file_path,
            'format': file_format
        })

    def _calculate_overall_progress(
        self,
        stage_weights: Dict[str, float],
        current_stage: str,
        stage_progress: float,
        processed_rows: int,
        total_rows: int,
        start_time: float,
        message: str
    ) -> ImportProgress:
        """计算总体进度"""
        weights = list(stage_weights.values())
        stages = list(stage_weights.keys())

        overall_progress = 0.0
        for i, stage in enumerate(stages):
            weight = weights[i]
            if stage == current_stage:
                overall_progress += weight * (stage_progress / 100)
                break
            else:
                overall_progress += weight

        return self._create_progress(
            overall_progress,
            current_stage,
            message,
            processed_rows,
            total_rows,
            start_time
        )

    def _safe_error(self, error: Exception, operation: str) -> MySQLMCPError:
        """安全地处理错误"""
        return safe_error_handler(error, operation=operation, session_id=self.session_id)

    async def _execute_import(
        self,
        table_name: str,
        data: List[Dict[str, Any]],
        options: ImportOptions
    ) -> Dict[str, Any]:
        """
        执行导入操作 - 核心导入引擎

        实现完整的导入流程：
        1. 获取表结构信息
        2. 创建字段映射
        3. 数据验证和清洗
        4. 重复检查（如果启用）
        5. 批量插入数据
        6. 处理冲突数据更新
        7. 进度监控和统计

        Args:
            table_name: 目标表名
            data: 要导入的数据
            options: 导入选项

        Returns:
            导入结果统计信息
        """
        start_time = TimeUtils.now_in_milliseconds()

        # 发送开始进度事件
        if options.with_progress:
            initial_progress = self._create_progress(
                0,
                'initialization',
                '开始导入操作...',
                0,
                len(data),
                start_time
            )
            await self._emit_progress(initial_progress, options.file_path, options.format)

        # 获取表结构
        table_schema = await self.mysql_manager.get_table_schema_cached(table_name)

        # 发送表结构获取完成进度
        if options.with_progress:
            progress = self._calculate_overall_progress(
                {'fileReading': 10, 'dataParsing': 20, 'validation': 30, 'insertion': 40},
                'validation',
                10,
                0,
                len(data),
                start_time,
                f'已获取表 "{table_name}" 的结构信息'
            )
            await self._emit_progress(progress, options.file_path, options.format)

        # 初始化候选键信息（用于重复检查优化）
        duplicate_config = options.duplicate_check or {
            'enable': options.skip_duplicates is True,
            'useCache': True,
            'candidateKeys': [key.fields for key in self.current_candidate_keys]
        }

        if duplicate_config.get('enable', False):
            await self.initialize_candidate_keys(table_name)

        # 创建字段映射
        field_mapping = self._create_field_mapping(
            options.field_mapping or {},
            data,
            table_schema
        )

        # 发送字段映射完成进度
        if options.with_progress:
            progress = self._calculate_overall_progress(
                {'fileReading': 10, 'dataParsing': 20, 'validation': 30, 'insertion': 40},
                'validation',
                30,
                0,
                len(data),
                start_time,
                '字段映射配置完成'
            )
            await self._emit_progress(progress, options.file_path, options.format)

        # 验证所有数据并处理冲突
        validated_data: List[Dict[str, Any]] = []
        duplicate_data: List[Dict[str, Any]] = []
        skipped_rows = 0
        failed_rows = 0

        # 首先进行数据验证
        validated_rows_with_index: List[Dict[str, Any]] = []
        validation_errors: List[Dict[str, Any]] = []

        for i, row in enumerate(data):
            # 发送验证进度（每处理10%或100行发送一次）
            if options.with_progress and (i % max(100, len(data) // 20) == 0 or i == len(data) - 1):
                validation_progress = (i + 1) / len(data) * 50  # 验证占50%的进度
                progress = self._calculate_overall_progress(
                    {'fileReading': 10, 'dataParsing': 20, 'validation': 70, 'insertion': 20},
                    'validation',
                    validation_progress,
                    i + 1,
                    len(data),
                    start_time,
                    f'正在验证数据格式... ({i + 1}/{len(data)})'
                )
                await self._emit_progress(progress, options.file_path, options.format)

            # 执行数据验证
            validation_result = self._validate_data_row(row, field_mapping, table_schema)

            # 如果数据验证失败，根据策略处理
            if validation_result['errors']:
                if options.conflict_strategy == 'error':
                    raise MySQLMCPError(
                        f'数据验证失败（第{i + 1}行）: {", ".join(validation_result["errors"])}',
                        ErrorCategory.VALIDATION_ERROR,
                        ErrorSeverity.MEDIUM
                    )
                else:
                    failed_rows += 1
                    validation_errors.append({
                        'index': i,
                        'errors': validation_result['errors'],
                        'row': row
                    })
                    continue

            # 记录验证通过的行及其索引
            validated_rows_with_index.append({'row': row, 'index': i})

        # 执行批量重复检查
        duplicate_results: List[bool] = []
        if duplicate_config.get('enable', False) and validated_rows_with_index:
            # 发送重复检查进度
            if options.with_progress:
                progress = self._calculate_overall_progress(
                    {'fileReading': 10, 'dataParsing': 20, 'validation': 70, 'insertion': 20},
                    'validation',
                    60,
                    len(validated_rows_with_index),
                    len(data),
                    start_time,
                    '正在进行重复数据检查...'
                )
                await self._emit_progress(progress, options.file_path, options.format)

            try:
                rows_to_check = [item['row'] for item in validated_rows_with_index]
                duplicate_results = await self.batch_check_duplicates(
                    table_name,
                    rows_to_check,
                    field_mapping,
                    {
                        'config': duplicate_config,
                        'batchSize': options.batch_size or 200
                    }
                )
            except Exception as error:
                # 记录重复检查错误但不阻止导入
                if duplicate_config.get('precisionLevel') == 'exact':
                    raise

        # 处理验证结果
        for i, item in enumerate(validated_rows_with_index):
            row = item['row']
            index = item['index']
            is_duplicate = duplicate_results[i] if i < len(duplicate_results) else False

            if is_duplicate:
                # 根据冲突处理策略执行相应操作
                if options.conflict_strategy == 'skip':
                    skipped_rows += 1
                    continue
                elif options.conflict_strategy == 'update':
                    duplicate_data.append(self.map_data_row(row, field_mapping))
                elif options.conflict_strategy == 'error':
                    raise MySQLMCPError(
                        f'检测到重复数据（第{index + 1}行），策略设置为停止导入',
                        ErrorCategory.DATA_INTEGRITY_ERROR,
                        ErrorSeverity.HIGH
                    )
            else:
                # 非重复数据，直接添加
                validated_data.append(self.map_data_row(row, field_mapping))

        # 批量插入新数据和更新重复数据
        batch_size = options.batch_size or 1000
        imported_rows = 0
        updated_rows = 0
        batches_processed = 0

        # 发送验证完成，开始插入阶段
        if options.with_progress:
            progress = self._calculate_overall_progress(
                {'fileReading': 10, 'dataParsing': 20, 'validation': 50, 'insertion': 20},
                'insertion',
                0,
                len(data),
                len(data),
                start_time,
                f'数据验证完成，开始插入数据... ({len(validated_data)} 行待插入)'
            )
            await self._emit_progress(progress, options.file_path, options.format)

        # 首先处理新数据的插入
        if validated_data:
            if options.use_transaction:
                # 使用事务
                data_rows = [list(row.values()) for row in validated_data]
                result = await self.mysql_manager.execute_batch_queries([
                    {
                        'sql': f'INSERT INTO `{table_name}` ({", ".join(f"`{k}`" for k in validated_data[0].keys())}) VALUES ({", ".join(["%s"] * len(validated_data[0]))})',
                        'params': data_rows
                    }
                ])
                imported_rows = len(data_rows)
                batches_processed = 1

                # 发送事务插入完成进度
                if options.with_progress:
                    progress = self._calculate_overall_progress(
                        {'fileReading': 10, 'dataParsing': 20, 'validation': 50, 'insertion': 20},
                        'insertion',
                        100,
                        len(data),
                        len(data),
                        start_time,
                        f'事务中成功插入 {imported_rows} 行数据'
                    )
                    await self._emit_progress(progress, options.file_path, options.format)
            else:
                # 逐批插入
                for i in range(0, len(validated_data), batch_size):
                    batch = validated_data[i:i + batch_size]
                    data_rows = [list(row.values()) for row in batch]
                    columns = list(batch[0].keys())

                    # 构建批量插入SQL
                    placeholders = ', '.join([f'({", ".join(["%s"] * len(columns))})' for _ in batch])
                    sql = f'INSERT INTO `{table_name}` ({", ".join(f"`{col}`" for col in columns)}) VALUES {placeholders}'

                    try:
                        await self.mysql_manager.execute_query(sql, [param for row in data_rows for param in row])
                        imported_rows += len(batch)
                        batches_processed += 1

                        # 发送批量插入进度
                        if options.with_progress:
                            insertion_progress = (i + len(batch)) / len(validated_data) * 100
                            progress = self._calculate_overall_progress(
                                {'fileReading': 10, 'dataParsing': 20, 'validation': 50, 'insertion': 20},
                                'insertion',
                                insertion_progress,
                                len(data) - (len(validated_data) - (i + len(batch))),
                                len(data),
                                start_time,
                                f'正在插入数据... ({insertion_progress:.0f}%) - 已处理 {i + len(batch)}/{len(validated_data)} 行'
                            )
                            await self._emit_progress(progress, options.file_path, options.format)
                    except Exception as error:
                        logger.error(f'批量插入失败: {error}', 'MySQLImportTool')
                        # 继续处理下一批

        # 处理重复数据的更新
        if duplicate_data:
            if options.with_progress:
                progress = self._calculate_overall_progress(
                    {'fileReading': 10, 'dataParsing': 20, 'validation': 50, 'insertion': 20},
                    'insertion',
                    90,
                    len(data),
                    len(data),
                    start_time,
                    f'正在更新 {len(duplicate_data)} 行重复数据...'
                )
                await self._emit_progress(progress, options.file_path, options.format)

            updated_rows = await self.handle_duplicate_updates(table_name, duplicate_data, field_mapping, options)

        # 发送完成进度事件
        if options.with_progress:
            final_progress = self._create_progress(
                100,
                'completed',
                f'导入完成! 总共处理 {len(data)} 行数据，成功导入 {imported_rows + updated_rows} 行',
                len(data),
                len(data),
                start_time
            )
            await self._emit_progress(final_progress, options.file_path, options.format)

        return {
            'imported_rows': imported_rows,
            'skipped_rows': skipped_rows,
            'failed_rows': failed_rows,
            'updated_rows': updated_rows,
            'batches_processed': batches_processed
        }

    async def check_duplicate(
        self,
        table_name: str,
        row: Dict[str, Any],
        field_mapping: List[FieldMapping],
        options: Optional[Dict[str, Any]] = None
    ) -> bool:
        """
        检查重复数据

        根据配置的候选键检查指定行是否在数据库中已存在。
        支持缓存机制以提高性能。

        Args:
            table_name: 表名
            row: 要检查的行数据
            field_mapping: 字段映射配置
            options: 重复检查选项

        Returns:
            如果存在重复数据返回True，否则返回False
        """
        start_time = TimeUtils.now_in_milliseconds()
        use_cache = options.get('useCache', True) if options else True
        config = options.get('config') if options else None

        try:
            # 获取用于重复检查的字段列表
            duplicate_fields = self._get_duplicate_check_fields(field_mapping, config)

            if not duplicate_fields:
                return False

            # 生成查询条件
            condition_data = self._build_duplicate_condition(row, duplicate_fields)

            # 没有任何可用于检查的条件
            if not condition_data['queryConditions']:
                return False

            # 检查缓存
            if use_cache:
                cache_hit = self._check_duplicate_cache(condition_data['queryKey'])
                if cache_hit is not None:
                    return cache_hit

            # 执行数据库查询检查
            is_duplicate = await self._execute_duplicate_query(table_name, condition_data)

            # 缓存查询结果
            if use_cache:
                self._set_duplicate_cache(condition_data['queryKey'], is_duplicate)

            # 记录重复检查性能
            duration = TimeUtils.get_duration_in_ms(start_time)
            # 检查耗时超过100ms时记录
            if duration > 100:
                logger.warning('重复检查耗时较长', 'MySQLImportTool', {
                    'table_name': table_name,
                    'duration': duration,
                    'field_count': len(duplicate_fields),
                    'query_conditions': len(condition_data['queryConditions'])
                })

            return is_duplicate

        except Exception as error:
            logger.error('重复检查失败', 'MySQLImportTool', error, {
                'table_name': table_name,
                'error': str(error)
            })

            # 根据配置决定是否在出错时返回true
            return config.get('precisionLevel') == 'exact' if config else False

    async def batch_check_duplicates(
        self,
        table_name: str,
        rows: List[Dict[str, Any]],
        field_mapping: List[FieldMapping],
        options: Optional[Dict[str, Any]] = None
    ) -> List[bool]:
        """
        批量重复检查

        批量执行重复检查以提高性能。支持分批处理和错误恢复。

        Args:
            table_name: 表名
            rows: 要检查的行数据列表
            field_mapping: 字段映射配置
            options: 重复检查选项

        Returns:
            每个行的重复检查结果列表
        """
        results: List[bool] = []
        batch_size = options.get('batchSize', 100) if options else 100

        for i in range(0, len(rows), batch_size):
            batch = rows[i:i + batch_size]

            # 并发检查当前批次
            batch_results = await asyncio.gather(
                *[self.check_duplicate(table_name, row, field_mapping, options) for row in batch],
                return_exceptions=True
            )

            # 处理结果
            for result in batch_results:
                if isinstance(result, Exception):
                    # 出错时默认返回不重复
                    results.append(False)
                else:
                    results.append(result)

        return results

    async def initialize_candidate_keys(self, table_name: str) -> None:
        """
        初始化表格的候选键信息

        从数据库中获取表的索引信息，构建候选键列表用于重复检查优化。
        按优先级排序：主键 > 唯一索引 > 普通索引

        Args:
            table_name: 表名
        """
        try:
            schema = await self.mysql_manager.get_table_schema_cached(table_name)

            if not schema:
                return

            candidate_keys: List[CandidateKey] = []

            # 首先收集主键
            primary_key_columns = [col for col in schema if col.get('Key') == 'PRI']
            if primary_key_columns:
                primary_key = CandidateKey(
                    key_name='PRIMARY',
                    fields=[col['Field'] for col in primary_key_columns],
                    is_unique=True,
                    priority=100
                )
                candidate_keys.append(primary_key)

            # 收集唯一索引
            unique_indexes = [col for col in schema if col.get('Key') == 'UNI']
            for i, col in enumerate(unique_indexes):
                unique_key = CandidateKey(
                    key_name=f'UNIQUE_{i}',
                    fields=[col['Field']],
                    is_unique=True,
                    priority=90 - i
                )
                candidate_keys.append(unique_key)

            # 收集普通索引（优先级较低）
            regular_indexes = [col for col in schema if col.get('Key') == 'MUL']
            for i, col in enumerate(regular_indexes):
                regular_key = CandidateKey(
                    key_name=f'INDEX_{i}',
                    fields=[col['Field']],
                    is_unique=False,
                    priority=50 - i
                )
                candidate_keys.append(regular_key)

            candidate_keys.sort(key=lambda x: x.priority, reverse=True)
            self.current_candidate_keys = candidate_keys

            logger.info('候选键初始化完成', 'MySQLImportTool', {
                'table_name': table_name,
                'candidate_keys_count': len(candidate_keys),
                'keys': [{'name': k.key_name, 'fields': k.fields, 'unique': k.is_unique} for k in candidate_keys]
            })

        except Exception as error:
            logger.error('候选键初始化失败', 'MySQLImportTool', error, {
                'table_name': table_name
            })

    def clear_duplicate_cache(self) -> None:
        """清理重复检查缓存"""
        self.duplicate_cache.clear()
        self.current_candidate_keys.clear()

    def get_duplicate_cache_stats(self) -> Dict[str, Any]:
        """获取重复检查缓存统计信息"""
        return {
            'cache_size': len(self.duplicate_cache),
            'max_cache_size': self.max_cache_size,
            'candidate_keys_count': len(self.current_candidate_keys)
        }

    def map_data_row(
        self,
        row: Dict[str, Any],
        field_mapping: List[FieldMapping]
    ) -> Dict[str, Any]:
        """映射数据行"""
        mapped_row: Dict[str, Any] = {}

        for mapping in field_mapping:
            value = row.get(mapping.source_field)

            # 类型转换
            converted_value = value
            if mapping.type_conversion == 'number' and isinstance(value, str):
                try:
                    converted_value = float(value)
                except ValueError:
                    converted_value = value
            elif mapping.type_conversion == 'boolean' and isinstance(value, str):
                converted_value = value.lower() in ('true', '1')

            mapped_row[mapping.target_field] = converted_value

        return mapped_row

    async def handle_duplicate_updates(
        self,
        table_name: str,
        duplicate_data: List[Dict[str, Any]],
        field_mapping: List[FieldMapping],
        options: ImportOptions
    ) -> int:
        """
        处理重复数据更新

        根据冲突处理策略更新重复的记录。

        Args:
            table_name: 表名
            duplicate_data: 重复数据列表
            field_mapping: 字段映射配置
            options: 导入选项

        Returns:
            更新的记录数量
        """
        updated_rows = 0

        # 获取表的键字段（用于WHERE条件）
        primary_key_fields = [mapping for mapping in field_mapping if mapping.required]

        if not primary_key_fields:
            logger.error('无法更新记录：未找到主键或必需字段', 'MySQLImportTool')
            return 0

        # 逐条更新重复数据
        for row in duplicate_data:
            try:
                where_conditions: List[str] = []
                where_params: List[Any] = []
                set_values: Dict[str, Any] = {}

                # 构建WHERE条件（基于主键）
                for mapping in primary_key_fields:
                    value = row.get(mapping.target_field)
                    if value is not None and value != '':
                        where_conditions.append(f'`{mapping.target_field}` = %s')
                        where_params.append(value)

                # 构建SET值（排除主键字段）
                for mapping in field_mapping:
                    if not any(pk.target_field == mapping.target_field for pk in primary_key_fields):
                        set_values[mapping.target_field] = row.get(mapping.target_field)

                if not where_conditions or not set_values:
                    continue  # 跳过无法更新的行

                set_clause = ', '.join([f'`{field}` = %s' for field in set_values.keys()])
                set_params = list(set_values.values())
                query = f'UPDATE `{table_name}` SET {set_clause} WHERE {" AND ".join(where_conditions)}'

                await self.mysql_manager.execute_query(query, [*set_params, *where_params])
                updated_rows += 1

                logger.debug('重复数据更新成功', 'MySQLImportTool', {
                    'table_name': table_name,
                    'updated_fields': list(set_values.keys()),
                    'where_conditions': ' AND '.join(where_conditions)
                })

            except Exception as error:
                logger.error('重复数据更新失败', 'MySQLImportTool', error, {
                    'table_name': table_name,
                    'row': row
                })
                # 根据策略处理更新错误
                if options.conflict_strategy == 'error':
                    raise MySQLMCPError(
                        f'更新重复数据失败: {str(error)}',
                        ErrorCategory.DATA_ERROR,
                        ErrorSeverity.MEDIUM
                    )

        return updated_rows

    def _check_duplicate_cache(self, condition_key: str) -> Optional[bool]:
        """
        检查重复检查缓存

        检查指定的查询条件是否在缓存中，如果存在且未过期则返回结果。

        Args:
            condition_key: 查询条件键

        Returns:
            如果缓存命中返回True/False，如果未命中或过期返回None
        """
        if condition_key not in self.duplicate_cache:
            return None

        cache_item = self.duplicate_cache[condition_key]

        # 检查缓存是否过期（1小时）
        current_time = TimeUtils.now_in_milliseconds()
        if current_time - cache_item.timestamp > 3600000:  # 1小时
            del self.duplicate_cache[condition_key]
            return None

        return cache_item.exists

    def _set_duplicate_cache(self, condition_key: str, exists: bool) -> None:
        """
        设置重复检查缓存

        将查询结果存储到缓存中，如果缓存已满则使用LRU策略清除最旧的条目。

        Args:
            condition_key: 查询条件键
            exists: 是否存在重复数据
        """
        # 检查缓存大小限制
        if len(self.duplicate_cache) >= self.max_cache_size:
            # 使用LRU策略清除最旧的项目
            oldest_key = min(self.duplicate_cache.keys(),
                           key=lambda k: self.duplicate_cache[k].timestamp)
            del self.duplicate_cache[oldest_key]

        # 添加新缓存项
        self.duplicate_cache[condition_key] = DuplicateCacheItem(
            condition_key=condition_key,
            exists=exists,
            timestamp=int(TimeUtils.now_in_milliseconds())
        )

    def _get_duplicate_check_fields(
        self,
        field_mapping: List[FieldMapping],
        config: Optional[Dict[str, Any]] = None
    ) -> List[FieldMapping]:
        """
        获取用于重复检查的字段列表

        根据配置和候选键选择最适合的字段用于重复检查。
        优先级：配置的候选键 > 显式标记的字段 > 必填字段 > 所有字段

        Args:
            field_mapping: 字段映射列表
            config: 重复检查配置

        Returns:
            用于重复检查的字段列表
        """
        # 首先尝试使用配置中指定的候选键
        if config and config.get('candidateKeys'):
            candidate_fields: List[FieldMapping] = []

            for key_fields in config['candidateKeys']:
                key_mappings: List[FieldMapping] = []

                for field_name in key_fields:
                    mapping = next((f for f in field_mapping if f.target_field == field_name), None)
                    if mapping:
                        key_mappings.append(mapping)

                if len(key_mappings) == len(key_fields):
                    candidate_fields.extend(key_mappings)
                    break  # 使用第一个完整匹配的候选键

            if candidate_fields:
                return candidate_fields

        # 使用映射中指定用于重复检查的字段
        explicit_fields = [mapping for mapping in field_mapping if mapping.check_duplicate]
        if explicit_fields:
            return explicit_fields

        # 最后使用主键或非空必填字段
        primary_key_fields = [mapping for mapping in field_mapping if mapping.required]
        if primary_key_fields:
            return primary_key_fields

        # 作为最后的备选，使用所有映射字段
        return [mapping for mapping in field_mapping if mapping.source_field and mapping.target_field]

    def _build_duplicate_condition(
        self,
        row: Dict[str, Any],
        duplicate_fields: List[FieldMapping]
    ) -> Dict[str, Any]:
        """
        构建重复检查查询条件

        根据行数据和字段映射构建SQL查询条件。

        Args:
            row: 行数据
            duplicate_fields: 用于重复检查的字段列表

        Returns:
            包含查询条件、参数和键的字典
        """
        query_conditions: List[str] = []
        query_params: List[Any] = []
        condition_parts: List[str] = []

        for mapping in duplicate_fields:
            value = row.get(mapping.source_field)

            if value is not None and value != '':
                # 根据精度等级处理值
                param_value = value
                if isinstance(value, str):
                    param_value = value.strip()

                query_conditions.append(f'`{mapping.target_field}` = %s')
                query_params.append(param_value)
                condition_parts.append(f'{mapping.target_field}:{str(param_value)}')

        # 生成查询键用于缓存
        query_key = '|'.join(condition_parts)

        return {
            'queryConditions': query_conditions,
            'queryParams': query_params,
            'queryKey': query_key
        }

    async def _execute_duplicate_query(
        self,
        table_name: str,
        condition_data: Dict[str, Any]
    ) -> bool:
        """
        执行重复检查数据库查询

        执行数据库查询检查是否存在重复记录。

        Args:
            table_name: 表名
            condition_data: 查询条件数据

        Returns:
            如果存在重复记录返回True，否则返回False
        """
        if not condition_data['queryConditions']:
            return False

        query = f'SELECT 1 FROM `{table_name}` WHERE {" AND ".join(condition_data["queryConditions"])} LIMIT 1'
        result = await self.mysql_manager.execute_query(query, condition_data['queryParams'])

        return bool(result)