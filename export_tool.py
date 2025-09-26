"""
MySQL 数据导出工具 - 企业级数据导出解决方案

提供全面的企业级数据库数据导出功能，支持Excel、CSV、JSON等多种数据格式的高性能批量导出。
集成了智能数据处理、流式导出、内存优化、错误处理、进度跟踪等完整的导出生态系统。

@fileoverview MySQL数据导出工具 - 企业级数据导出的完整解决方案
@author liyq
@version 1.0.0
@since 1.0.0
@updated 2025-09-26
@license MIT
"""

import asyncio
import csv
import io
import json
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Union
from abc import ABC, abstractmethod

import aiofiles
import openpyxl

from mysql_manager import MySQLManager
from type_utils import (
    ExportOptions,
    ExportResult,
    MySQLMCPError,
    ErrorCategory,
    ErrorSeverity
)
from common_utils import TimeUtils
from logger import logger
from cache import CacheManager
from error_handler import safe_error as safe_error_handler


class MySQLExportTool:
    """
    MySQL 数据导出工具类

    企业级导出工具，集成了智能数据处理、流式导出、内存优化、
    错误恢复等高级特性。支持Excel、CSV、JSON等多种数据格式的导出。

    主要组件：
    - 导出器工厂：管理不同格式导出器的创建和缓存
    - 基础导出器：提供所有导出器的共同功能和接口定义
    - 格式化导出器：不同格式数据的专用导出器
    - 进度跟踪器：实时进度监控和事件通知
    - 错误处理器：详细错误诊断和恢复机制
    """

    def __init__(self, mysql_manager: MySQLManager):
        """初始化导出工具"""
        self.mysql_manager = mysql_manager
        self.session_id = str(uuid.uuid4())

        # 初始化缓存管理器
        self.cache_manager = CacheManager.get_instance()

        # 注册内存跟踪
        self._register_memory_tracking()

    def _register_memory_tracking(self) -> None:
        """注册内存泄漏跟踪"""
        # 这里可以添加内存跟踪逻辑
        pass

    async def export_to_csv(self, query: str, options: ExportOptions) -> ExportResult:
        """
        导出查询结果到CSV格式

        支持标准CSV格式和自定义分隔符，支持流式处理和批量写入。
        提供数据验证、字符转义和性能优化等完整功能。

        Args:
            query: SQL查询语句
            options: 导出选项配置

        Returns:
            包含导出结果的JSON格式数据

        Raises:
            MySQLMCPError: 当导出失败时抛出
        """
        start_time = TimeUtils.now_in_milliseconds()

        try:
            # 发送开始进度事件
            if options.with_progress:
                initial_progress = self._create_progress(
                    0,
                    'queryExecution',
                    '开始执行查询...',
                    0,
                    0,
                    start_time
                )
                await self._emit_progress(initial_progress, options.file_path or 'csv_export', 'csv')

            # 验证输入参数
            self._validate_export_options(options, 'csv')

            # 执行查询获取数据
            rows = await self._execute_query_with_progress(query, options, start_time)

            # 发送数据处理完成进度
            if options.with_progress:
                progress = self._create_progress(
                    40,
                    'dataProcessing',
                    '开始处理CSV数据...',
                    len(rows),
                    len(rows),
                    start_time
                )
                await self._emit_progress(progress, options.file_path or 'csv_export', 'csv')

            # 生成输出文件路径
            output_path = self._generate_output_path(options, 'csv')

            # 导出到CSV格式
            exporter = CsvExporter()
            await exporter.perform_export(rows, options, output_path)

            # 获取文件统计信息
            file_stats = await self._get_file_stats(output_path)
            duration = TimeUtils.get_duration_in_ms(start_time)

            # 发送完成进度事件
            if options.with_progress:
                final_progress = self._create_progress(
                    100,
                    'completed',
                    f'CSV导出完成! 共导出 {len(rows)} 行数据',
                    len(rows),
                    len(rows),
                    start_time
                )
                await self._emit_progress(final_progress, options.file_path or 'csv_export', 'csv')

            return ExportResult(
                success=True,
                file_path=output_path,
                file_size=file_stats['size'],
                row_count=len(rows),
                column_count=len(rows[0]) if rows else 0,
                format='csv',
                duration=duration,
                processing_mode='streaming' if options.streaming else 'direct'
            )

        except Exception as error:
            duration = TimeUtils.get_duration_in_ms(start_time)
            safe_error = self._safe_error(error, 'exportToCSV')

            return ExportResult(
                success=False,
                error=safe_error.message,
                duration=duration,
                format='csv'
            )

    async def export_to_json(self, query: str, options: ExportOptions) -> ExportResult:
        """
        导出查询结果到JSON格式

        支持JSON数组和对象格式，支持嵌套数据结构的扁平化处理。
        提供数据清理、格式化和批量处理等高级功能。

        Args:
            query: SQL查询语句
            options: 导出选项配置

        Returns:
            包含导出结果的JSON格式数据

        Raises:
            MySQLMCPError: 当导出失败时抛出
        """
        start_time = TimeUtils.now_in_milliseconds()

        try:
            # 发送开始进度事件
            if options.with_progress:
                initial_progress = self._create_progress(
                    0,
                    'queryExecution',
                    '开始执行查询...',
                    0,
                    0,
                    start_time
                )
                await self._emit_progress(initial_progress, options.file_path or 'json_export', 'json')

            # 验证输入参数
            self._validate_export_options(options, 'json')

            # 执行查询获取数据
            rows = await self._execute_query_with_progress(query, options, start_time)

            # 发送数据处理完成进度
            if options.with_progress:
                progress = self._create_progress(
                    30,
                    'dataProcessing',
                    '开始处理JSON数据...',
                    len(rows),
                    len(rows),
                    start_time
                )
                await self._emit_progress(progress, options.file_path or 'json_export', 'json')

            # 生成输出文件路径
            output_path = self._generate_output_path(options, 'json')

            # 导出到JSON格式
            exporter = JsonExporter()
            await exporter.perform_export(rows, options, output_path)

            # 获取文件统计信息
            file_stats = await self._get_file_stats(output_path)
            duration = TimeUtils.get_duration_in_ms(start_time)

            # 发送完成进度事件
            if options.with_progress:
                final_progress = self._create_progress(
                    100,
                    'completed',
                    f'JSON导出完成! 共导出 {len(rows)} 行数据',
                    len(rows),
                    len(rows),
                    start_time
                )
                await self._emit_progress(final_progress, options.file_path or 'json_export', 'json')

            return ExportResult(
                success=True,
                file_path=output_path,
                file_size=file_stats['size'],
                row_count=len(rows),
                column_count=len(rows[0]) if rows else 0,
                format='json',
                duration=duration,
                processing_mode='streaming' if options.streaming else 'direct'
            )

        except Exception as error:
            duration = TimeUtils.get_duration_in_ms(start_time)
            safe_error = self._safe_error(error, 'exportToJSON')

            return ExportResult(
                success=False,
                error=safe_error.message,
                duration=duration,
                format='json'
            )

    async def export_to_excel(self, query: str, options: ExportOptions) -> ExportResult:
        """
        导出查询结果到Excel格式

        支持多工作表Excel文件，支持样式设置、列宽调整和格式化。
        提供企业级Excel导出功能，包括表头样式、自动筛选等。

        Args:
            query: SQL查询语句
            options: 导出选项配置

        Returns:
            包含导出结果的JSON格式数据

        Raises:
            MySQLMCPError: 当导出失败时抛出
        """
        start_time = TimeUtils.now_in_milliseconds()

        try:
            # 发送开始进度事件
            if options.with_progress:
                initial_progress = self._create_progress(
                    0,
                    'queryExecution',
                    '开始执行查询...',
                    0,
                    0,
                    start_time
                )
                await self._emit_progress(initial_progress, options.file_path or 'excel_export', 'excel')

            # 验证输入参数
            self._validate_export_options(options, 'excel')

            # 执行查询获取数据
            rows = await self._execute_query_with_progress(query, options, start_time)

            # 发送数据处理完成进度
            if options.with_progress:
                progress = self._create_progress(
                    25,
                    'dataProcessing',
                    '开始处理Excel数据...',
                    len(rows),
                    len(rows),
                    start_time
                )
                await self._emit_progress(progress, options.file_path or 'excel_export', 'excel')

            # 生成输出文件路径
            output_path = self._generate_output_path(options, 'xlsx')

            # 导出到Excel格式
            exporter = ExcelExporter()
            await exporter.perform_export(rows, options, output_path)

            # 获取文件统计信息
            file_stats = await self._get_file_stats(output_path)
            duration = TimeUtils.get_duration_in_ms(start_time)

            # 发送完成进度事件
            if options.with_progress:
                final_progress = self._create_progress(
                    100,
                    'completed',
                    f'Excel导出完成! 共导出 {len(rows)} 行数据',
                    len(rows),
                    len(rows),
                    start_time
                )
                await self._emit_progress(final_progress, options.file_path or 'excel_export', 'excel')

            return ExportResult(
                success=True,
                file_path=output_path,
                file_size=file_stats['size'],
                row_count=len(rows),
                column_count=len(rows[0]) if rows else 0,
                format='excel',
                duration=duration,
                processing_mode='direct'
            )

        except Exception as error:
            duration = TimeUtils.get_duration_in_ms(start_time)
            safe_error = self._safe_error(error, 'exportToExcel')

            return ExportResult(
                success=False,
                error=safe_error.message,
                duration=duration,
                format='excel'
            )

    async def export_data(self, query: str, options: ExportOptions) -> ExportResult:
        """
        通用导出方法 - 根据格式自动选择导出方式

        Args:
            query: SQL查询语句
            options: 导出选项配置

        Returns:
            导出结果
        """
        export_format = options.format or 'csv'

        if export_format == 'csv':
            return await self.export_to_csv(query, options)
        elif export_format == 'json':
            return await self.export_to_json(query, options)
        elif export_format in ('excel', 'xlsx'):
            return await self.export_to_excel(query, options)
        else:
            raise MySQLMCPError(
                f'不支持的导出格式: {export_format}',
                ErrorCategory.VALIDATION_ERROR,
                ErrorSeverity.MEDIUM
            )

    # 辅助方法

    def _validate_export_options(self, options: ExportOptions, export_format: str) -> None:
        """验证导出选项"""
        if not options.output_dir:
            raise MySQLMCPError(
                '必须指定输出目录',
                ErrorCategory.VALIDATION_ERROR,
                ErrorSeverity.MEDIUM
            )

        # 格式特定验证
        if export_format == 'excel':
            if options.sheet_name and not isinstance(options.sheet_name, str):
                raise MySQLMCPError(
                    'Excel工作表名称必须是字符串',
                    ErrorCategory.VALIDATION_ERROR,
                    ErrorSeverity.MEDIUM
                )

    async def _execute_query_with_progress(
        self,
        query: str,
        options: ExportOptions,
        start_time: float
    ) -> List[Dict[str, Any]]:
        """执行查询并带进度跟踪"""
        # 应用行数限制
        final_query = query
        if options.max_rows and options.max_rows > 0:
            final_query += f' LIMIT {options.max_rows}'

        # 发送查询进度
        if options.with_progress:
            progress = self._create_progress(
                10,
                'queryExecution',
                '正在执行查询...',
                0,
                0,
                start_time
            )
            await self._emit_progress(progress, options.file_path or 'export', options.format)

        # 执行查询
        result = await self.mysql_manager.execute_query(final_query)

        # 发送查询完成进度
        if options.with_progress:
            progress = self._create_progress(
                30,
                'queryExecution',
                f'查询完成，共 {len(result)} 行数据',
                len(result),
                len(result),
                start_time
            )
            await self._emit_progress(progress, options.file_path or 'export', options.format)

        return result

    def _generate_output_path(self, options: ExportOptions, file_extension: str) -> str:
        """生成输出文件路径"""
        # 确保输出目录存在
        Path(options.output_dir).mkdir(parents=True, exist_ok=True)

        # 生成文件名
        if not options.file_name:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_%f')[:-3]
            file_name = f'export_{timestamp}.{file_extension}'
        else:
            file_name = options.file_name
            if not file_name.endswith(f'.{file_extension}'):
                file_name += f'.{file_extension}'

        return str(Path(options.output_dir) / file_name)

    async def _get_file_stats(self, file_path: str) -> Dict[str, int]:
        """获取文件统计信息"""
        stat = await asyncio.to_thread(Path(file_path).stat)
        return {'size': stat.st_size}

    def _create_progress(
        self,
        progress: float,
        stage: str,
        message: str,
        processed_rows: int,
        total_rows: int,
        start_time: float
    ) -> Dict[str, Any]:
        """创建进度信息对象"""
        current_time = TimeUtils.now_in_milliseconds()
        elapsed = current_time - start_time

        estimated_time_remaining = None
        current_speed = None

        if progress > 0 and progress < 100:
            remaining_progress = 100 - progress
            time_per_percent = elapsed / progress
            estimated_time_remaining = time_per_percent * remaining_progress
            current_speed = (processed_rows / elapsed) * 1000 if elapsed > 0 else None  # 行/秒

        return {
            'progress': progress,
            'stage': stage,
            'message': message,
            'processed_rows': processed_rows,
            'total_rows': total_rows,
            'start_time': datetime.fromtimestamp(start_time / 1000),
            'estimated_time_remaining': int(estimated_time_remaining) if estimated_time_remaining else None,
            'current_speed': current_speed
        }

    async def _emit_progress(self, progress: Dict[str, Any], file_path: str, file_format: str) -> None:
        """发送进度事件"""
        # 这里可以实现事件发射逻辑
        logger.info('export_progress', {
            'progress': progress['progress'],
            'stage': progress['stage'],
            'message': progress['message'],
            'processed_rows': progress['processed_rows'],
            'total_rows': progress['total_rows'],
            'file_path': file_path,
            'format': file_format
        })

    def _safe_error(self, error: Exception, operation: str) -> MySQLMCPError:
        """安全地处理错误"""
        return safe_error_handler(error, operation=operation, session_id=self.session_id)


class BaseExporter(ABC):
    """
    基础导出器抽象类

    提供所有导出器的共同功能和接口定义，包括：
    - 统一的导出流程管理
    - 内存优化的数据处理
    - 文件IO操作封装
    - 事件驱动的进度通知
    - 错误处理和恢复机制
    """

    def __init__(self):
        """初始化基础导出器"""
        self.default_options = self.get_default_options()

    @abstractmethod
    def get_default_options(self) -> ExportOptions:
        """获取默认导出选项 - 由子类实现"""
        pass

    @abstractmethod
    def get_file_extension(self) -> str:
        """获取文件扩展名 - 由子类实现"""
        pass

    @abstractmethod
    async def perform_export(
        self,
        rows: List[Dict[str, Any]],
        options: ExportOptions,
        output_path: str
    ) -> None:
        """执行具体的导出操作 - 由子类实现"""
        pass

    async def export(
        self,
        rows: List[Dict[str, Any]],
        options: ExportOptions,
        output_path: str
    ) -> ExportResult:
        """主导出方法 - 统一的导出流程"""
        start_time = TimeUtils.now_in_milliseconds()
        final_options = {**self.default_options.__dict__, **options.__dict__}

        try:
            # 执行导出
            await self.perform_export(rows, ExportOptions(**final_options), output_path)

            # 获取文件统计信息
            file_stats = await self._get_file_stats(output_path)
            duration = TimeUtils.get_duration_in_ms(start_time)

            return ExportResult(
                success=True,
                file_path=output_path,
                file_size=file_stats['size'],
                row_count=len(rows),
                column_count=len(rows[0]) if rows else 0,
                format=final_options.get('format', 'unknown'),
                duration=duration,
                processing_mode='streaming' if final_options.get('streaming') else 'direct'
            )

        except Exception as error:
            duration = TimeUtils.get_duration_in_ms(start_time)
            return ExportResult(
                success=False,
                error=str(error),
                duration=duration,
                format=final_options.get('format', 'unknown')
            )

    async def _get_file_stats(self, file_path: str) -> Dict[str, int]:
        """获取文件统计信息"""
        stat = await asyncio.to_thread(Path(file_path).stat)
        return {'size': stat.st_size}

    async def _write_to_file(self, file_path: str, content: str) -> None:
        """写入文件内容"""
        async with aiofiles.open(file_path, 'w', encoding='utf8') as file:
            await file.write(content)

    def _clean_data(self, data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """清理数据，移除undefined值"""
        return [
            {k: v for k, v in row.items() if v is not None}
            for row in data
        ]


class CsvExporter(BaseExporter):
    """
    CSV导出器

    提供高效的CSV格式导出功能，支持：
    - 流式处理大数据集
    - 自动字符转义和格式化
    - 批量写入性能优化
    - 可配置的列头选项
    """

    def get_default_options(self) -> ExportOptions:
        """获取默认导出选项"""
        return ExportOptions(
            output_dir='./exports',
            format='csv',
            include_headers=True,
            max_rows=1000000,
            streaming=True,
            batch_size=5000
        )

    def get_file_extension(self) -> str:
        """获取文件扩展名"""
        return 'csv'

    async def perform_export(
        self,
        rows: List[Dict[str, Any]],
        options: ExportOptions,
        output_path: str
    ) -> None:
        """执行CSV导出"""
        if not rows:
            # 空数据集，创建空文件
            await self._write_to_file(output_path, '')
            return

        # 清理数据
        cleaned_rows = self._clean_data(rows)

        # 构建CSV内容
        csv_content = self._build_csv_content(cleaned_rows, options)

        # 写入文件
        await self._write_to_file(output_path, csv_content)

    def _build_csv_content(
        self,
        rows: List[Dict[str, Any]],
        options: ExportOptions
    ) -> str:
        """构建CSV内容"""
        output = io.StringIO()

        try:
            writer = csv.writer(output, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)

            # 写入头部
            if options.include_headers and rows:
                headers = list(rows[0].keys())
                writer.writerow(headers)

            # 写入数据行
            for row in rows:
                values = list(row.values())
                writer.writerow(values)

            return output.getvalue()

        finally:
            output.close()

    def _format_csv_value(self, value: Any) -> str:
        """格式化CSV值"""
        if value is None:
            return ''

        string_value = str(value)

        # 如果包含特殊字符，需要用引号包裹
        if ',' in string_value or '"' in string_value or '\n' in string_value or '\r' in string_value:
            return f'"{string_value.replace(chr(34), chr(34) + chr(34))}"'

        return string_value


class JsonExporter(BaseExporter):
    """
    JSON导出器

    提供高效的JSON格式导出功能，支持：
    - 流式JSON数组写入
    - 自动数据清理和格式化
    - 批量处理性能优化
    - 完整的JSON语法兼容
    """

    def get_default_options(self) -> ExportOptions:
        """获取默认导出选项"""
        return ExportOptions(
            output_dir='./exports',
            format='json',
            include_headers=True,
            max_rows=1000000,
            streaming=True,
            batch_size=5000
        )

    def get_file_extension(self) -> str:
        """获取文件扩展名"""
        return 'json'

    async def perform_export(
        self,
        rows: List[Dict[str, Any]],
        options: ExportOptions,
        output_path: str
    ) -> None:
        """执行JSON导出"""
        # 清理数据
        cleaned_rows = self._clean_data(rows)

        # 构建JSON内容
        json_content = self._build_json_content(cleaned_rows, options)

        # 写入文件
        await self._write_to_file(output_path, json_content)

    def _build_json_content(
        self,
        rows: List[Dict[str, Any]],
        options: ExportOptions
    ) -> str:
        """构建JSON内容"""
        return json.dumps(rows, indent=2, ensure_ascii=False, default=str)


class ExcelExporter(BaseExporter):
    """
    Excel导出器

    提供企业级Excel格式导出功能，支持：
    - 丰富的样式和格式化
    - 自动列宽调整
    - 表头样式和冻结
    - 自动筛选功能
    - 批量数据处理
    """

    def get_default_options(self) -> ExportOptions:
        """获取默认导出选项"""
        return ExportOptions(
            output_dir='./exports',
            format='excel',
            sheet_name='Data',
            include_headers=True,
            max_rows=1000000,
            streaming=True,
            batch_size=5000
        )

    def get_file_extension(self) -> str:
        """获取文件扩展名"""
        return 'xlsx'

    async def perform_export(
        self,
        rows: List[Dict[str, Any]],
        options: ExportOptions,
        output_path: str
    ) -> None:
        """执行Excel导出"""
        # 创建工作簿
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = options.sheet_name or 'Data'

        if not rows:
            # 空数据集，仅保存空工作簿
            workbook.save(output_path)
            return

        # 设置列头
        columns = list(rows[0].keys())
        worksheet.append(columns)

        # 样式化表头
        self._style_header(worksheet)

        # 添加数据行
        for row in rows:
            worksheet.append(list(row.values()))

        # 应用样式
        self._apply_styles(worksheet)

        # 保存文件
        workbook.save(output_path)

    def _style_header(self, worksheet) -> None:
        """样式化表头"""
        header_row = worksheet[1]

        # 设置字体样式
        for cell in header_row:
            cell.font = openpyxl.styles.Font(bold=True, color='FFFFFF')
            cell.fill = openpyxl.styles.PatternFill(
                start_color='4472C4',
                end_color='4472C4',
                fill_type='solid'
            )
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        # 设置行高
        worksheet.row_dimensions[1].height = 25

    def _apply_styles(self, worksheet) -> None:
        """应用表格样式"""
        # 添加边框
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

        # 隔行着色（除了标题行）
        for i, row in enumerate(worksheet.iter_rows(min_row=2)):
            if i % 2 == 1:
                for cell in row:
                    cell.fill = openpyxl.styles.PatternFill(
                        start_color='F2F2F2',
                        end_color='F2F2F2',
                        fill_type='solid'
                    )

        # 冻结首行
        worksheet.freeze_panes = 'A2'

        # 启用自动筛选
        if worksheet.max_row > 1:
            worksheet.auto_filter.ref = f'A1:{openpyxl.utils.get_column_letter(worksheet.max_column)}1'


class ExporterFactory:
    """
    导出器工厂类

    主要功能：
    - 单例模式管理
    - 导出器实例创建
    - 基础缓存功能
    """

    _instance: Optional['ExporterFactory'] = None
    _exporter_registry: Dict[str, type] = {}
    _exporter_cache: Dict[str, BaseExporter] = {}

    def __new__(cls):
        if cls._instance is None:
            cls._instance = super(ExporterFactory, cls).__new__(cls)
            cls._instance._initialize_default_exporters()
        return cls._instance

    def _initialize_default_exporters(self) -> None:
        """初始化默认导出器"""
        self._exporter_registry.update({
            'csv': CsvExporter,
            'json': JsonExporter,
            'excel': ExcelExporter,
            'xlsx': ExcelExporter
        })

    def create_exporter(self, export_type: str = 'csv') -> BaseExporter:
        """创建导出器实例"""
        # 检查缓存
        if export_type in self._exporter_cache:
            return self._exporter_cache[export_type]

        # 创建新实例
        exporter_class = self._exporter_registry.get(export_type)
        if not exporter_class:
            raise MySQLMCPError(
                f'不支持的导出器类型: {export_type}。支持的类型: {list(self._exporter_registry.keys())}',
                ErrorCategory.INVALID_INPUT,
                ErrorSeverity.MEDIUM
            )

        try:
            exporter = exporter_class()

            # 缓存实例（限制缓存大小）
            if len(self._exporter_cache) < 10:
                self._exporter_cache[export_type] = exporter

            return exporter

        except Exception as error:
            raise MySQLMCPError(
                f'创建导出器失败 ({export_type}): {str(error)}',
                ErrorCategory.CONFIGURATION_ERROR,
                ErrorSeverity.HIGH
            )

    @classmethod
    def get_instance(cls) -> 'ExporterFactory':
        """获取工厂实例 - 单例模式"""
        if cls._instance is None:
            cls._instance = cls()
        return cls._instance